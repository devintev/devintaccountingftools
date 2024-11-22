# ╔══════════════════════════════════════════════════════════════╗
# ║                      ★ DEPENDENCY GRAPH ★                    ║
# ║                     BEGIN DEPENDENCY GRAPH                   ║
# ╚══════════════════════════════════════════════════════════════╝

# ========== [ Class: BytesIOWrapper ] ==========
# BytesIOWrapper.__init__ ( <- 0 x)
# BytesIOWrapper.__getattr__ ( <- 0 x)
# BytesIOWrapper.read ( <- 1 x)
#   -> BytesIOWrapper.read
# BytesIOWrapper.write ( <- 1 x)
#   -> BytesIOWrapper.write

# ========== [ Class: SecretsAndSettingsManager ] ==========
# SecretsAndSettingsManager.__init__ ( <- 0 x)
# SecretsAndSettingsManager.get_secret ( <- 1 x)
#   -> SecretsAndSettingsManager.get_secret
# SecretsAndSettingsManager.get_function_app_url ( <- 0 x)

# ========== [ Class: HTMLListHandler ] ==========
# HTMLListHandler.__init__ ( <- 1 x)
#   -> HTMLListHandler.__init__
# HTMLListHandler.emit ( <- 0 x)
# HTMLListHandler.get_html_log ( <- 0 x)

# ========== [ Class: DevIntConnector ] ==========
# DevIntConnector.__init__ ( <- 0 x)
# DevIntConnector._colchar ( <- 7 x)
# DevIntConnector._check_depth ( <- 3 x)
#   -> DevIntConnector._check_depth
# DevIntConnector.load_settings ( <- 0 x)
# DevIntConnector.get_settings ( <- 0 x)
# DevIntConnector.set_key_vault_access ( <- 1 x)
# DevIntConnector.setup ( <- 0 x)
#   -> DevIntConnector.set_key_vault_access
# DevIntConnector.build_clients ( <- 0 x)
# DevIntConnector.analyse_received_http_request ( <- 0 x)
# DevIntConnector.download_all_sheets ( <- 2 x)
# DevIntConnector.read_time_slots ( <- 1 x)
# DevIntConnector.read_kontenrahmen ( <- 1 x)
# DevIntConnector.read_kostenstellenplan ( <- 1 x)
# DevIntConnector.read_reports_overview ( <- 1 x)
# DevIntConnector.read_report_schema_into ( <- 1 x)
# DevIntConnector.read_distribution_instructions ( <- 1 x)
# DevIntConnector.read_instruction_files ( <- 0 x)
#   -> DevIntConnector.read_time_slots
#   -> DevIntConnector.read_kontenrahmen
#   -> DevIntConnector.read_kostenstellenplan
#   -> DevIntConnector.read_reports_overview
#   -> DevIntConnector.read_distribution_instructions
#   -> DevIntConnector.download_all_sheets
#   -> DevIntConnector.read_report_schema_into
# DevIntConnector.get_all_bb_posts ( <- 0 x)
# DevIntConnector.get_all_bb_accounts ( <- 0 x)
# DevIntConnector.read_expected_bookings ( <- 0 x)
#   -> DevIntConnector.download_all_sheets
# DevIntConnector.add_more_account_information_to_bookings ( <- 1 x)
# DevIntConnector.collect_all_accounts ( <- 1 x)
# DevIntConnector.add_up_bookings ( <- 2 x)
# DevIntConnector.collect_all_costlocations ( <- 1 x)
#   -> DevIntConnector.add_up_bookings
# DevIntConnector.fill_bookings_to_sheet ( <- 2 x)
#   -> DevIntConnector._colchar
# DevIntConnector.fill_costlocations_to_sheet ( <- 1 x)
# DevIntConnector.fill_accounts_to_sheet ( <- 1 x)
# DevIntConnector.get_bookings_sheet_name ( <- 4 x)
# DevIntConnector.get_listings_sheet_name ( <- 4 x)
# DevIntConnector.fill_report_row_to_listings_sheet ( <- 1 x)
# DevIntConnector.select_bookings_by_costlocation ( <- 1 x)
# DevIntConnector.fill_listings_sheet ( <- 1 x)
#   -> DevIntConnector._colchar
#   -> DevIntConnector._check_depth
#   -> DevIntConnector.select_bookings_by_costlocation
#   -> DevIntConnector.fill_report_row_to_listings_sheet
# DevIntConnector.fill_bookings_listings_sheet ( <- 0 x)
# DevIntConnector._add_slot_header ( <- 1 x)
#   -> DevIntConnector._colchar
# DevIntConnector.build_group_summation_formula ( <- 1 x)
#   -> DevIntConnector._colchar
# DevIntConnector._get_column_letter_by_column_header ( <- 1 x)
# DevIntConnector.fill_row_to_sheet ( <- 1 x)
#   -> DevIntConnector.get_bookings_sheet_name
#   -> DevIntConnector.build_group_summation_formula
#   -> DevIntConnector._get_column_letter_by_column_header
#   -> DevIntConnector._colchar
# DevIntConnector._set_border_to_area ( <- 1 x)
# DevIntConnector.fill_report_to_sheet ( <- 1 x)
#   -> DevIntConnector.fill_bookings_to_sheet
#   -> DevIntConnector._add_slot_header
#   -> DevIntConnector._colchar
#   -> DevIntConnector.fill_row_to_sheet
#   -> DevIntConnector._set_border_to_area
#   -> DevIntConnector.get_listings_sheet_name
#   -> DevIntConnector.fill_listings_sheet
# DevIntConnector.fill_personnel_bookings_to_sheet ( <- 1 x)
#   -> DevIntConnector._colchar
# DevIntConnector.build_personnel_bookings ( <- 1 x)
# DevIntConnector.compile_all_xls_sheets ( <- 1 x)
#   -> DevIntConnector.fill_bookings_to_sheet
#   -> DevIntConnector.fill_costlocations_to_sheet
#   -> DevIntConnector.fill_accounts_to_sheet
#   -> DevIntConnector.fill_personnel_bookings_to_sheet
#   -> DevIntConnector.fill_report_to_sheet
#   -> DevIntConnector.get_bookings_sheet_name
#   -> DevIntConnector.get_listings_sheet_name
# DevIntConnector.build_reports ( <- 0 x)
#   -> DevIntConnector.add_more_account_information_to_bookings
#   -> DevIntConnector.build_personnel_bookings
#   -> DevIntConnector.collect_all_accounts
#   -> DevIntConnector.collect_all_costlocations
#   -> DevIntConnector.compile_all_xls_sheets
#   -> DevIntConnector.add_up_bookings
#   -> DevIntConnector._check_depth
# DevIntConnector.get_report_summary ( <- 3 x)
# DevIntConnector.send_email_sending ( <- 1 x)
#   -> DevIntConnector.get_report_summary
#   -> DevIntConnector.get_bookings_sheet_name
#   -> DevIntConnector.get_listings_sheet_name
# DevIntConnector.send_teams_sending ( <- 1 x)
#   -> DevIntConnector.get_report_summary
# DevIntConnector.send_azure_blob_sending ( <- 1 x)
#   -> DevIntConnector.get_report_summary
#   -> DevIntConnector.get_bookings_sheet_name
#   -> DevIntConnector.get_listings_sheet_name
# DevIntConnector.send_cosmosdb_sending ( <- 1 x)
# DevIntConnector.send_reports ( <- 0 x)
#   -> DevIntConnector.send_email_sending
#   -> DevIntConnector.send_teams_sending
#   -> DevIntConnector.send_azure_blob_sending
#   -> DevIntConnector.send_cosmosdb_sending

# ========== [ Non-Class Functions ] ==========
# read_html_page_template ( <- 0 x)
# extract_data_from_received_http_request ( <- 0 x)
# replace_and_format_html_template ( <- 0 x)
# get_db_config_from_keyvault ( <- 0 x)
# save_virtual_workbook ( <- 0 x)

# ╔══════════════════════════════════════════════════════════════╗
# ║                     END DEPENDENCY GRAPH                     ║
# ╚══════════════════════════════════════════════════════════════╝

# last updated 2024-11-14
import logging
import yaml
import os
from os import getenv
import re
import datetime
import requests
import json
import base64
from typing import List
from io import StringIO, BytesIO
from pytz import timezone
import pandas as pd
from bs4 import BeautifulSoup

from azure.keyvault.secrets import SecretClient
from azure.functions import HttpRequest
from azure.identity import DefaultAzureCredential
from azure.mgmt.web import WebSiteManagementClient
from azure.storage.blob import BlobServiceClient  # , BlobClient, ContainerClient
from azure.cosmos import CosmosClient
import azure.cosmos.exceptions as cdbexceptions

from tempfile import NamedTemporaryFile
from openpyxl import Workbook, styles, load_workbook  # , workbook
from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.utils.dataframe import dataframe_to_rows

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName, FileType, Disposition)  # ContentId
import pymsteams


class BytesIOWrapper:
    def __init__(self, string_buffer, encoding='utf-8'):
        self.string_buffer = string_buffer
        self.encoding = encoding

    def __getattr__(self, attr):
        return getattr(self.string_buffer, attr)

    def read(self, size=-1):
        content = self.string_buffer.read(size)
        return content.encode(self.encoding)

    def write(self, b):
        content = b.decode(self.encoding)
        return self.string_buffer.write(content)


class SecretsAndSettingsManager:
    """
    This class manages secrets and settings, specifically for Azure Key Vault.

    Attributes:
    logger (logging.Logger): The logger to use for logging messages.
    key_vault_client (azure.keyvault.secrets.SecretClient): The client to use for accessing the Azure Key Vault.
    """

    def __init__(self, parent_logger=None):
        """
        Initializes the SecretsAndSettingsManager.

        Parameters:
        parent_logger (logging.Logger, optional): The parent logger. If none is provided, a new logger is created.
        """

        try:
            if parent_logger is not None:
                self.logger = parent_logger.getChild(__class__.__name__)
            else:
                self.logger = logging.getLogger(
                    f"{__name__}.{__class__.__name__}")
        except:
            self.logger = logging.getLogger(__name__)
        try:
            self.credential = DefaultAzureCredential(
                exclude_developer_cli_credential=False)
            self.logger.debug(
                "Obtaining Azure Default Credentials succeeded.")
        except:
            self.logger.error("Obtaining Azure Default Credentials failed.")
            self.credential = None
        key_vault_name = getenv("KEY_VAULT_NAME", None)
        if key_vault_name:
            self.logger.debug(
                "KEY_VAULT_NAME was successfully identified in the environment variables: " + key_vault_name)
            key_vault_Uri = f"https://{key_vault_name}.vault.azure.net"
            if self.credential is not None:
                try:
                    self.key_vault_client = SecretClient(
                        vault_url=key_vault_Uri, credential=self.credential)
                    self.logger.debug(
                        "SecretClient was successfully created.")
                except:
                    self.logger.critical("Creating SecretClient failed.")
                    self.key_vault_client = None
            else:
                self.logger.critical(
                    "KEY_VAULT_NAME is was found in environment variables but credential is None.")
                self.key_vault_client = None
        else:
            self.logger.critical(
                "KEY_VAULT_NAME is not found in environment variables.")
            self.key_vault_client = None

    def get_secret(self, secret_name) -> str:
        """
        Retrieves a secret from the Azure Key Vault.

        Parameters:
        secret_name (str): The name of the secret to retrieve.

        Returns:
        str: The value of the secret. If the secret could not be retrieved, returns None.
        """
        secret_value = ""
        if self.key_vault_client is not None:
            try:
                secret = self.key_vault_client.get_secret(secret_name)
                secret_value = secret.value
            except:
                self.logger.error(
                    f"Getting secret {secret_name} from key vault failed.")
                secret = None
                try:
                    secrets = self.key_vault_client.list_properties_of_secrets()
                    secret_names = list(
                        [secret_item.name for secret_item in secrets])
                    if secret_name in secret_names:
                        self.logger.error(
                            f"Secret {secret_name} is in list of secrets but could not be retrieved. Check the keyvault's access policies in Azure Portal.")
                    else:
                        self.logger.error(
                            f"Secret {secret_name} is not in list of {len(secret_names)} secrets of the keyvault. Check the keyvault's access policies in Azure Portal.")
                except:
                    self.logger.error(
                        "Getting list of secrets failed. Check access policies of keyvault in Azure Portal.")
                    secret_names = None
        else:
            self.logger.error(
                "KEY_VAULT_NAME is not set in environment variables.")
            secret = None
        return secret_value

    def get_function_app_url(self) -> str:
        """
        Retrieves the URL of the Azure Function App.

        This method first attempts to retrieve the Function App URL from the 'AZURE_FUNCTION_APP_URL' environment variable.

        If 'AZURE_FUNCTION_APP_URL' is not set, it then checks if 'AZURE_SUBSCRIPTION_ID', 'AZURE_RESOURCE_GROUP_NAME', and 'AZURE_FUNCTION_APP_NAME' environment variables are set. If they are, it uses the Azure SDK's WebSiteManagementClient with the provided Azure credentials to fetch the Function App's details and construct the URL.

        If any of the required environment variables are not set, or if an error occurs while creating the WebSiteManagementClient, it logs an error and returns an empty string.

        Returns:
        str: The URL of the Function App. If the URL could not be retrieved, returns an empty string.
        """

        function_app_url = getenv('AZURE_FUNCTION_APP_URL')
        if function_app_url is None:
            function_app_url = ""
            self.logger.warning(
                "AZURE_FUNCTION_APP_URL is not set in environment variables. Trying to obtain function app URL from Azure SDK.")
            subscription_id = getenv("AZURE_SUBSCRIPTION_ID")
            resource_group_name = getenv("AZURE_RESOURCE_GROUP_NAME")
            function_app_name = getenv("AZURE_FUNCTION_APP_NAME")
            if subscription_id is not None and resource_group_name is not None and function_app_name is not None:
                self.logger.debug(
                    "Found AZURE_SUBSCRIPTION_ID, AZURE_RESOURCE_GROUP_NAME, and AZURE_FUNCTION_APP_NAME in environment variables.")
                if self.credential is not None:
                    self.logger.debug("Creating WebSiteManagementClient.")
                    try:
                        client = WebSiteManagementClient(
                            self.credential, subscription_id)
                        function_app = client.web_apps.get(
                            resource_group_name, function_app_name)
                        function_app_url = f"https://{function_app.default_host_name}"
                    except Exception as e:
                        self.logger.error(
                            f"Creating WebSiteManagementClient failed: {e}")
                else:
                    self.logger.error(
                        "Creating WebSiteManagementClient failed because: azure credential is missing and AZURE_FUNCTION_APP_URL is also not set in environment variables.")
            else:
                self.logger.error(
                    "Neither AZURE_FUNCTION_APP_URL nor all of AZURE_SUBSCRIPTION_ID, AZURE_RESOURCE_GROUP_NAME, AZURE_FUNCTION_APP_NAME are set in environment variables. Cannot obtain function app URL.")
        self.logger.debug(f"Obtained function app URL: '{function_app_url}'")
        return function_app_url


class HTMLListHandler(logging.Handler):
    """
    This class is a custom logging handler that stores log records in a list and can output them as an HTML table.

    Attributes:
    log_list (list): A list that stores the log records.
    """

    def __init__(self):
        super().__init__()
        self.log_list = []

    def emit(self, record):
        self.log_list.append(record)

    def get_html_log(self,
                     min_include_level=logging.DEBUG,
                     show_columns: List[str] = [
                         "Level", "Message", "Time", "Location", "Function"]
                     ) -> str:
        """
        Returns the log records as an HTML table.

        Parameters:
        min_include_level (int, optional): The minimum log level to include in the output. Default is logging.DEBUG.
        show_columns (list, optional): A list of column names to include in the output. Default is ["Level", "Message", "Time", "Location", "Function"].

        Returns:
        str: The log records formatted as an HTML table.
        """
        print(
            f"min_include_level: {min_include_level}, type: {type(min_include_level)}")
        # LogRecord-Attribute
        #
        # name: Der Name des Loggers, der den Log-Record erstellt hat.
        # msg: Die Log-Nachricht.
        # args: Die Argumente, die in die Log-Nachricht eingefügt werden.
        # levelname: Der Text, der den Schweregrad der Log-Nachricht repräsentiert, z.B. 'DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'.
        # levelno: Die numerische Darstellung des Schweregrads der Log-Nachricht.
        # pathname: Der vollständige Pfadname der Quelldatei, in der der Logger aufgerufen wurde.
        # filename: Der Dateiname der Quelldatei, in der der Logger aufgerufen wurde.
        # module: Der Modulname, in dem der Logger aufgerufen wurde.
        # lineno: Die Zeilennummer in der Quelldatei, an der der Logger aufgerufen wurde.
        # funcName: Der Name der Funktion oder Methode, in der der Logger aufgerufen wurde.
        # created: Die Zeit, zu der der Log-Record erstellt wurde (als Zeitstempel).
        # msecs: Die Millisekunden-Teil der Erstellungszeit.
        # relativeCreated: Die Zeit in Millisekunden zwischen der Erstellung des Log-Records und der Erstellung des Root-Loggers.
        # thread: Die Thread-ID.
        # threadName: Der Thread-Name.
        # process: Die Prozess-ID.
        # processName: Der Prozessname.
        # exception: Informationen über eine Ausnahme (falls vorhanden).

        table_columns = {"Level": "levelname", "Location": "name", "Line": "lineno",
                         "Function": "funcName", "Path": "pathname", "Time": "created", "Message": "msg"}
        show_columns = ["Level", "Message",
                        "Time", "Location", "Line", "Function"]

        html_log = ''
        records = [
            record for record in self.log_list if record.levelno >= min_include_level]
        print(
            f"all records: {len(self.log_list)}, filtered records: {len(records)}")
        if records:
            html_log += '<table class="logtable">\n'
            html_log += "<tr>\n"
            for col in show_columns:
                html_log += f"<th>{col}</th>\n"
            html_log += "</tr>\n"
        for record in records:
            html_log += "<tr>\n"
            for col in show_columns:
                cell_class = f"logcell_{col.lower()} logcell_{record.levelname.lower()}"
                if col == "Time":
                    value = datetime.datetime.fromtimestamp(
                        getattr(record, table_columns[col])).astimezone()
                    value = value.strftime(
                        "%H:%M:%S") + "." + str(value.microsecond // 100)[:4]
                else:
                    value = getattr(record, table_columns[col])
                html_log += f'<td class="{cell_class}">{value}</td>\n'
            html_log += "</tr>\n"
        if records:
            html_log += "</table>\n"
        return html_log


def read_html_page_template(
        filename_html,
        filename_css="assets/styles.css",
        filename_js="assets/includes.js"
) -> str:
    """
    This function reads an HTML page template and optionally a CSS file,
    and returns the HTML string with the CSS embedded.

    Parameters:
    filename_html (str): The path to the HTML file.
    filename_css (str, optional): The path to the CSS file. If none is provided,
                            no styles are embedded. Default is "assets/styles.css".

    Returns:
    str: The HTML string with the CSS embedded. If a CSS file was provided,
        its contents are embedded within a <style> tag. If no CSS file was
        provided, the HTML string is returned as is.
    """

    template_html_string = ""
    styles_string = ""
    replace_data: dict[str, str] = {}
    with open(filename_html, "r") as template_file:
        template_html_string = template_file.read()
    if filename_css is not None:
        with open(filename_css, "r") as styles_file:
            styles_string = styles_file.read()
        replace_data["styles"] = styles_string
    else:
        replace_data["styles"] = ""
    if filename_js is not None:
        with open(filename_js, "r") as js_file:
            js_string = js_file.read()
        replace_data["js_script"] = js_string
    else:
        replace_data["js_script"] = ""
    for key, value in replace_data.items():
        template_html_string = template_html_string.replace(
            "{{" + key + "}}", value)
    soup = BeautifulSoup(template_html_string, 'html.parser')
    return soup.prettify()


def extract_data_from_received_http_request(http_request: HttpRequest, parent_logger=None) -> dict:
    """
    This function extracts data from a received HTTP request.

    Parameters:
    http_request (HttpRequest): The received HTTP request.
    parent_logger (logging.Logger, optional): The parent logger. If none is provided, a new logger is created.

    Returns:
    dict: A dictionary containing the extracted data. For a POST request, the data is extracted from the request body and URL-encoded characters are replaced. For a GET request, the data is extracted from the request parameters and numbers are converted to int or float.

    Raises:
    Exception: If there is an error in decoding the request body or converting the numbers.
    """

    try:
        if parent_logger is not None:
            logger = parent_logger.getChild(__name__)
        else:
            logger = logging.getLogger(f"{__name__}")
    except:
        logger = logging.getLogger(__name__)

    def replace_post_chars(string):
        # see https://www.w3schools.com/tags/ref_urlencode.ASP
        rs = {"+": " ", "%E2%80%9C": "\"", "%2C": ",", "%3A": ":", "%2F": "/", "%3F": "?", "%3D": "=", "%26": "&", "%23": "#",
              "%25": "%", "%22": "\"", "%27": "'", }
        for key in rs:
            string = string.replace(key, rs[key])
        return string

    data = {
        "method": http_request.method,
        "url": http_request.url,
        "header": type(http_request.headers),
        "data": {}
    }
    post_data = {}
    if http_request.method == "POST":
        post_data_raw = http_request.get_body().decode("utf-8").split('&')
        if post_data_raw and post_data_raw[0]:
            logger.debug(
                f"Received data via http.{data['method']} method: {post_data_raw}")
            found_variables = set()
            for post_data_item in post_data_raw:
                [variable_name, value] = post_data_item.split('=')
                if variable_name in found_variables:
                    if type(post_data[variable_name]) is not list:
                        post_data[variable_name] = [post_data[variable_name]]
                    post_data[variable_name].append(replace_post_chars(value))
                else:
                    post_data[variable_name] = replace_post_chars(value)
                found_variables.add(variable_name)
    elif http_request.method == "GET":
        post_data = dict(http_request.params)

    # loop over all post_data items and convert numbers to int or float if they are of type string using try/except
    for key, value in post_data.items():
        if isinstance(value, str):
            try:
                post_data[key] = int(value)
            except ValueError:
                try:
                    post_data[key] = float(value)
                except ValueError:
                    pass
    return post_data


def replace_and_format_html_template(html_template, replace_data):
    """
    Replaces placeholders in an HTML template with corresponding values and formats the resulting HTML string.

    Parameters:
    html_template (str): The HTML template containing placeholders in the form "{{key}}".
    replace_data (dict): A dictionary containing key-value pairs. Each key corresponds to a placeholder in the HTML template and its value is the value to replace the placeholder with.

    Returns:
    str: The formatted HTML string where all placeholders have been replaced with their corresponding values.
    """
    for key, value in replace_data.items():
        html_template = html_template.replace("{{" + key + "}}", value)
    soup = BeautifulSoup(html_template, 'html.parser')
    html_template = soup.prettify(formatter="html5")
    return html_template
# .prettify(formatter="html5")


def get_db_config_from_keyvault(config: SecretsAndSettingsManager, logger: logging.Logger) -> dict | None:  # noqa
    db_con_data = None
    try:
        db_con_data = {
            "host": config.get_secret("wordpress-db-main-host"),
            "database": config.get_secret("wordpress-db-main-database-name"),
            "username": config.get_secret("wordpress-db-main-user"),
            "password": config.get_secret("wordpress-db-main-password")
        }
    except:
        logger.error(
            "Getting database connection data from keyvault failed.")
    return db_con_data


def save_virtual_workbook(workbook):
    """Save an openpyxl workbook in memory."""

    try:
        # Save workbook to a temporary file and read its content
        with NamedTemporaryFile() as f:
            workbook.save(f.name)
            # self.logger.debug("Workbook saved to a temporary file.")
            f.seek(0)
            content = f.read()
            # self.logger.debug("Workbook read into memory.")
        return content

    except Exception as e:  # noqa: e is intentionally unused
        # self.logger.error(f"Error in save_virtual_workbook: {e}")
        raise


class DevIntConnector:

    def __init__(self, parent_logger=None, settings_file='devint_settings.yaml'):
        try:
            if parent_logger is not None:
                self.logger = parent_logger.getChild(__class__.__name__)
            else:
                self.logger = logging.getLogger(
                    f"{__name__}.{__class__.__name__}")
        except:
            self.logger = logging.getLogger(__name__)

        self.key_vault: SecretsAndSettingsManager | None = None
        self.conn_clients = None
        self.settings = self.load_settings(settings_file)

        # possibly remove later
        # self.keep_personnel_columns = self.settings['keep_personnel_columns']
        # self.personnel_monthly_sum_columns = self.settings['personnel_monthly_sum_columns']
        # self.expense_term = self.settings["expense_term"]
        # self.income_term = self.settings["income_term"]

    def _colchar(self, col=0):
        """Convert a column number to an Excel-style column letter."""

        base = ord('A')
        rounds = (col - 1) // 26
        letters = chr(base + ((col - 1) % 26))

        if rounds > 0:
            letters = chr(base - 1 + rounds) + letters

        return letters

    def _check_depth(self, row, depth=1):
        mdepth = depth
        if 'children' in row and len(row['children']) > 0:
            cdepth = depth + 1
            for child in row['children']:
                ndepth = self._check_depth(child, cdepth)
                if ndepth > mdepth:
                    mdepth = ndepth
        return mdepth

    def load_settings(self, settings_file):
        try:
            with open(settings_file, 'r') as file:
                settings = yaml.safe_load(file)
                self.logger.debug(
                    f"Successfully loaded settings from {os.path.abspath(settings_file)}.")
                additional_settings = {
                    'slotTitleFont': styles.Font(bold=True, size=10),
                    'slotTitleAlignment': styles.Alignment(horizontal='center', vertical='top', wrap_text=True),
                    'slotDateFont': styles.Font(bold=False, size=6, color='333333'),
                    'titleFont': styles.Font(bold=True, size=18),
                    'depth1Font': styles.Font(name='Cambria', bold=True, size=12, ),
                    'depth2Font': styles.Font(name='Cambria', bold=True, size=14, ),
                    'depth3Font': styles.Font(name='Cambria', bold=True, size=16, ),
                    'depth4Font': styles.Font(name='Cambria', bold=True, size=18, ),
                    'depth5Font': styles.Font(name='Cambria', bold=True, size=20, ),
                    'level0Fill': styles.PatternFill(fill_type='solid', start_color='85bfe6', end_color='85bfe6'),
                    'level1Fill': styles.PatternFill(fill_type='solid', start_color='96c9eb', end_color='96c9eb'),
                    'level2Fill': styles.PatternFill(fill_type='solid', start_color='a7d4f2', end_color='a7d4f2'),
                    'level3Fill': styles.PatternFill(fill_type='solid', start_color='b5dbf5', end_color='b5dbf5'),
                    'listingsHeaderFill': styles.PatternFill(fill_type='solid', start_color='bee0ec', end_color='bee0ec')
                }
                settings.update(additional_settings)
                return settings
        except FileNotFoundError:
            self.logger.error(
                f"Settings file '{settings_file}' not found. Python tried to find it in directory: {os.path.abspath(settings_file)}.")
        except yaml.YAMLError as e:
            self.logger.error(
                f"Error parsing YAML file '{settings_file}': {str(e)}")
        except Exception as e:
            self.logger.error(
                f"An unexpected error occurred while loading settings from '{settings_file}': {str(e)}")
        return {}

    def get_settings(self):
        return self.settings

    def set_key_vault_access(self, config: SecretsAndSettingsManager):
        self.key_vault = config

    def setup(self, config: SecretsAndSettingsManager | None = None):
        if config:
            self.set_key_vault_access(config)
        if self.key_vault:
            self.conn_clients = self.build_clients()

    def build_clients(self):
        # Load settings
        settings = self.settings

        # Azure Blob Storage Information and Client
        blob_storage_account_name = settings.get(
            'azure_blob_storage_account_name')
        abs_acct_url = f"https://{blob_storage_account_name}.blob.core.windows.net/"
        try:
            blob_service_client = BlobServiceClient(
                abs_acct_url, credential=self.key_vault.credential)
            self.logger.debug(
                f"Successfully created BlobServiceClient for account: {blob_storage_account_name}.")
        except Exception as e:
            self.logger.error(
                f"Failed to create BlobServiceClient for account: {blob_storage_account_name}. Error: {str(e)}")
            return None

        # Blob Container Names
        templates_container_name = settings['blob_containers']['templates']
        financialplanning_container_name = settings['blob_containers']['financialplanning']
        financial_reports_container_name = settings['blob_containers']['financial_reports']
        working_time_reports_container_name = settings['blob_containers']['working_time_reports']
        wcc_invoicing_container_name = settings['blob_containers']['wcc_invoicing']
        od_invoicing_container_name = settings['blob_containers']['od_invoicing']

        # Buchhaltungsbuttler API Information
        try:
            bb_api_key = self.key_vault.get_secret(
                settings['secrets']['bb_api_key'])
            self.logger.debug(
                "Successfully retrieved Buchhaltungsbuttler API key.")
        except Exception as e:
            self.logger.error(
                f"Failed to retrieve Buchhaltungsbuttler API key. Error: {str(e)}")
            return None
        bb_authorization = self.key_vault.get_secret(
            settings['secrets']['bb_authorization'])
        bb_cookie = self.key_vault.get_secret(settings['secrets']['bb_cookie'])

        # Sendgrid API Information
        sendgrid_api_key = self.key_vault.get_secret(
            settings['secrets']['sendgrid_api_key'])
        try:
            sendgrid_client = SendGridAPIClient(sendgrid_api_key)
            self.logger.debug("Successfully created SendGridAPIClient.")
        except Exception as e:
            self.logger.error(
                f"Failed to create SendGridAPIClient. Error: {str(e)}")
            return None

        # Azure Cosmos DB Information and Client
        cosmos_db_url = settings.get('cosmos_db_url')
        db_access_key = self.key_vault.get_secret(
            settings['secrets']['cosmos_db_key'])
        try:
            cosmos_client = CosmosClient(
                url=cosmos_db_url, credential=db_access_key)
            self.logger.debug(
                f"Successfully created CosmosClient for URL: {cosmos_db_url}.")
        except Exception as e:
            self.logger.error(
                f"Failed to create CosmosClient for URL: {cosmos_db_url}. Error: {str(e)}")
            return None
        db_id = settings.get('cosmos_db_id')

        try:
            try:
                db = cosmos_client.create_database(id=db_id)
                self.logger.debug(
                    f"Successfully created database with id: {db_id}.")
            except cdbexceptions.CosmosResourceExistsError:
                db = cosmos_client.get_database_client(db_id)
                self.logger.debug(
                    f"Database with id: {db_id} already exists. Using existing database client.")
            except Exception as e:
                self.logger.error(
                    f"Failed to create or get Cosmos DB database with id: {db_id}. Error: {str(e)}")
                return None
        except cdbexceptions.CosmosResourceExistsError:
            db = cosmos_client.get_database_client(db_id)

        # Teams Webhook and Report Request Backlink
        try:
            teams_webhook = self.key_vault.get_secret(
                settings['secrets']['teams_webhook'])
            self.logger.debug("Successfully retrieved Teams webhook URL.")
        except Exception as e:
            self.logger.error(
                f"Failed to retrieve Teams webhook URL. Error: {str(e)}")
            return None
        try:
            reportrequest_backlink = self.key_vault.get_secret(
                settings['secrets']['reportrequest_backlink'])
            self.logger.debug(
                "Successfully retrieved report request backlink.")
        except Exception as e:
            self.logger.error(
                f"Failed to retrieve report request backlink. Error: {str(e)}")
            return None

        # Clients Dictionary
        self.conn_clients = {
            "blob_service": blob_service_client,
            "templates_folder": blob_service_client.get_container_client(container=templates_container_name),
            "financialplanning_folder": blob_service_client.get_container_client(container=financialplanning_container_name),
            "financial_reports_folder": blob_service_client.get_container_client(container=financial_reports_container_name),
            "working_time_reports_folder": blob_service_client.get_container_client(container=working_time_reports_container_name),
            "wcc_invoicing_folder": blob_service_client.get_container_client(container=wcc_invoicing_container_name),
            "od_invoicing_folder": blob_service_client.get_container_client(container=od_invoicing_container_name),
            "key_vault": self.key_vault,
            "cosmos": cosmos_client,
            "sendgrid": sendgrid_client,
            "teams_webhook": teams_webhook,
            "reportrequest_backlink": reportrequest_backlink,
            "bb_api_key": bb_api_key,
            "bb_authorization": bb_authorization,
            "bb_cookie": bb_cookie,
            "db": db
        }
        return self.conn_clients

    def analyse_received_http_request(self, http_request):

        def replace_post_chars(string):
            # see https://www.w3schools.com/tags/ref_urlencode.ASP
            rs = {"+": " ", "%E2%80%9C": "\"", "%2C": ",", "%3A": ":", "%2F": "/", "%3F": "?", "%3D": "=", "%26": "&", "%23": "#",
                  "%25": "%", "%22": "\"", "%27": "'", }
            for key in rs:
                string = string.replace(key, rs[key])
            return string

        data = {"method": http_request.method,
                "url": http_request.url,
                "header": type(http_request.headers),
                "data": {}
                }
        post_data = {}
        if http_request.method == "POST":
            post_data_raw = http_request.get_body().decode("utf-8").split('&')
            # case_tag_options = []
            if post_data_raw and post_data_raw[0]:
                for post_data_item in post_data_raw:
                    post_data_item_parts = post_data_item.split('=')
                    post_data[post_data_item_parts[0]] = replace_post_chars(
                        post_data_item_parts[1])
        elif http_request.method == "GET":
            post_data = dict(http_request.params)
        self.logger.log(f"received data with method: {data['method']}<br>")
        return post_data

    def download_all_sheets(self, blob_client):
        self.logger.debug(
            f"Starting download_all_sheets from {blob_client.container_name}/{blob_client.blob_name}.")

        try:
            # Initiate download stream
            self.logger.debug("Initializing blob download stream...")
            stream = BytesIO()
            blob_downloader = blob_client.download_blob()

            # Download blob data to stream
            blob_downloader.download_to_stream(stream)
            self.logger.debug(
                "Blob downloaded successfully into memory stream.")

            # Load Excel data from stream
            get_blob_data = pd.ExcelFile(stream, engine='openpyxl')
            self.logger.debug(
                f"Excel file loaded. Sheets found: {get_blob_data.sheet_names}")

            # Process each sheet and store in dictionary
            dataframes = {}
            for sheet in get_blob_data.sheet_names:
                # self.logger.debug(f"start processing sheet: {sheet}")
                new_df = pd.read_excel(
                    get_blob_data, engine='openpyxl', sheet_name=sheet)
                dataframes[sheet] = new_df
                # self.logger.debug(
                #     f"Sheet '{sheet}' read into DataFrame with {new_df.shape[0]} rows and {new_df.shape[1]} columns.")

            # self.logger.info(
            #     "All sheets downloaded and processed successfully.")
            return dataframes

        except Exception as e:
            self.logger.error(f"Error in download_all_sheets: {e}")
            raise

    def read_time_slots(self, df: pd.DataFrame):
        self.logger.debug("Starting read_time_slots function.")

        try:
            # Drop empty rows and log the result
            time_slots_df = df.dropna(how='all')
            self.logger.debug(
                f"Dropped empty rows. Remaining rows: {len(time_slots_df)}")

            # Rename columns and log the new column names
            rename_list = {'Slot Start': 'start', 'Slot End': 'end',
                           'Slot Title': 'name', 'Slot ID': 'id'}
            time_slots_df = time_slots_df.rename(
                columns=rename_list, inplace=False)
            self.logger.debug(f"Columns renamed: {rename_list}")

            # Initialize dictionary to store results
            dict_list = {}
            self.logger.debug(
                "Processing each row to extract time slot information.")

            for time_slot_ID in time_slots_df.index:
                mdict = time_slots_df.loc[time_slot_ID].to_dict()
                # self.logger.debug(f"Processing time slot ID: {time_slot_ID}")

                # Convert start and end to strings and store
                if pd.notnull(mdict.get('start')) and pd.notnull(mdict.get('end')):
                    mdict['startString'] = mdict['start'].strftime('%d.%m.%Y')
                    mdict['endString'] = mdict['end'].strftime('%d.%m.%Y')
                else:
                    self.logger.warning(
                        f"Missing start or end date for time slot ID {time_slot_ID}. Skipping conversion.")

                # Store time slot information in dictionary
                dict_list[str(mdict['id'])] = mdict

            self.logger.info("All time slots processed successfully.")
            return dict_list

        except Exception as e:
            self.logger.error(f"Error in read_time_slots: {e}")
            raise

    def read_kontenrahmen(self, dfs: dict):
        self.logger.debug("Starting read_kontenrahmen function.")

        try:
            # Initialize kontenrahmen dictionary to store processed data
            kontenrahmen = {}
            self.logger.debug(
                f"Processing {len(dfs)} dataframes for 'kontenrahmen'.")

            for key, df in dfs.items():
                self.logger.debug(f"Processing dataframe with key: {key}")
                entries = []

                for index in df.index:
                    mdict = df.loc[index].to_dict()

                    # Construct entry dictionary with required fields
                    entry = {
                        'accountRangeStart': mdict.get('Start'),
                        'accountRangeEnd': mdict.get('Ende'),
                        'type1': mdict.get('Typ 1'),
                        'type2': mdict.get('Typ 2')
                    }

                    # Check and add optional categories if they are present
                    if not pd.isna(mdict.get('Kategorie 1')):
                        entry['category1'] = mdict['Kategorie 1']
                    if not pd.isna(mdict.get('Kategorie 2')):
                        entry['category2'] = mdict['Kategorie 2']
                    if not pd.isna(mdict.get('Kategorie 3')):
                        entry['category3'] = mdict['Kategorie 3']
                    entries.append(entry)

                # Store entries list in kontenrahmen dictionary with a modified key name
                kontenrahmen[key[13:]] = entries
                self.logger.debug(
                    f"Processed {len(entries)} entries for key '{key[13:]}'.")

            # Set a default kontenrahmen entry
            kontenrahmen['default'] = kontenrahmen[list(
                kontenrahmen.keys())[0]]
            self.logger.debug(
                "Default kontenrahmen set to the first processed entry.")

            self.logger.info(
                "Completed processing all dataframes for 'kontenrahmen'.")
            return kontenrahmen

        except Exception as e:
            self.logger.error(f"Error in read_kontenrahmen: {e}")
            raise

    def read_kostenstellenplan(self, df: pd.DataFrame):
        self.logger.debug("Starting read_kostenstellenplan function.")

        try:
            # Initialize costlocations dictionary to store processed data
            costlocations = {}

            # Drop empty rows and rename relevant columns
            ksp_df = df.dropna(how='all')
            self.logger.info(
                f"Dropped empty rows. Remaining rows: {len(ksp_df)}")
            ksp_df = ksp_df.rename(
                columns={
                    'Unnamed: 5': 'sectionname', 'Unnamed: 6': 'groupName',
                    'Unnamed: 7': 'subGroupName', 'Unnamed: 8': 'name', 'Unnamed: 9': 'type'
                },
                inplace=False
            )
            # self.logger.debug("Renamed columns as per mapping.")

            # Drop rows missing key identifier
            ksp_df = ksp_df.dropna(subset=['Unnamed: 1'])
            max_costlocation_digits = 4
            self.logger.info("Starting to process each row in the dataframe.")

            for ksp_df_index in ksp_df.index:
                item = {"limits": {}}
                mdict = ksp_df.loc[ksp_df_index].to_dict()
                # self.logger.debug(f"Processing row at index {ksp_df_index}")

                # Determine item type based on type column
                if not pd.isna(mdict['type']):
                    if mdict['type'] in ['income', 'Einnahmen', self.settings["income_term"]]:
                        item['type'] = 'income'
                    elif mdict['type'] in ['expense', 'Ausgaben', self.settings["expense_term"]]:
                        item['type'] = 'expense'
                    elif mdict['type'] == 'budget':
                        item['type'] = 'budget'
                    # self.logger.debug(f"Set item type for index {ksp_df_index}: {item['type']}")

                # Process limits
                if not pd.isna(mdict.get('Unnamed: 10')):
                    item['limits']['max'] = float(mdict['Unnamed: 10'])
                    # self.logger.debug(f"Set max limit for index {ksp_df_index}: {item['limits']['max']}")
                if not pd.isna(mdict.get('Unnamed: 11')):
                    item['limits']['extendedMax'] = float(mdict['Unnamed: 11'])
                    # self.logger.debug(f"Set extended max limit for index {ksp_df_index}: {item['limits']['extendedMax']}")

                # Handle group items
                if not pd.isna(mdict.get('groupName')) or not pd.isna(mdict.get('subGroupName')):
                    item['type'] = 'group'
                    item['name'] = mdict['subGroupName'] if not pd.isna(
                        mdict.get('subGroupName')) else mdict['groupName']
                    # self.logger.debug(f"Set item as group with name '{item['name']}'")

                    # Construct the ID and range of the cost location
                    starting_digits = "".join(str(int(mdict[col])) for col in [
                                              'Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3'] if not pd.isna(mdict[col]))
                    min_digits = starting_digits + '0' * \
                        (max_costlocation_digits - len(starting_digits))
                    max_digits = starting_digits + '9' * \
                        (max_costlocation_digits - len(starting_digits))

                    item.update({
                        'id': starting_digits,
                        'costLocations': {
                            'startingDigits': starting_digits,
                            'min': int(min_digits),
                            'max': int(max_digits)
                        },
                        'children': {}
                    })
                    costlocations[item['id']] = item
                    # self.logger.debug(f"Group item with ID '{item['id']}' has range {min_digits} - {max_digits}")

                # Handle individual items
                elif not pd.isna(mdict.get('name')):
                    item['type'] = 'item'
                    item['name'] = mdict['name']
                    digits = "".join(str(int(mdict[col])) for col in [
                                     'Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3'] if not pd.isna(mdict[col]))
                    item['number'] = int(digits)
                    item['id'] = digits
                    costlocations[item['id']] = item
                    # self.logger.debug(f"Item with ID '{item['id']}' and name '{item['name']}' added.")

            # Assign children to groups based on their ranges
            self.logger.debug(
                "Assigning children to groups based on cost location ranges.")
            for id, item in costlocations.items():
                if item['type'] == 'group':
                    for child_id, child_item in costlocations.items():
                        if child_item['type'] == 'item' and item['costLocations']['min'] <= child_item['number'] <= item['costLocations']['max']:
                            item['children'][child_id] = child_item
                            # self.logger.debug(f"Added item '{child_id}' as a child of group '{id}'")
                        elif child_item['type'] == 'group' and child_id != id and item['costLocations']['min'] <= child_item['costLocations']['min'] and item['costLocations']['max'] >= child_item['costLocations']['max']:
                            item['children'][child_id] = child_item
                            # self.logger.debug(f"Added group '{child_id}' as a subgroup of '{id}'")

            self.logger.info("Completed processing kostenstellenplan.")
            return costlocations

        except Exception as e:
            self.logger.error(f"Error in read_kostenstellenplan: {e}")
            raise

    def read_reports_overview(self, df: pd.DataFrame):
        self.logger.debug("Starting read_reports_overview function.")

        # Initialize dictionary to store reports and counter for ignored reports
        reports = {}
        ignored_reports = 0
        self.logger.debug("Processing report entries in the dataframe.")

        for report_ID in df.index:
            mdict = df.loc[report_ID].to_dict()
            report = {
                'id': mdict["Plan"],
                'name': mdict['Name'],
                'costLocationsRange': {
                    'min': mdict['Kostenstellenstart'],
                    'max': mdict['Kostenstellenende']
                }
            }

            # Include only reports with version 2
            if mdict['Version'] == 2:
                reports[mdict["Plan"]] = report
                self.logger.debug(
                    f"Added report with ID '{report['id']}' and name '{report['name']}'.")
            else:
                ignored_reports += 1
                self.logger.debug(
                    f"Ignored report with ID '{report['id']}' due to deprecated version.")

        # Log ignored reports if any
        if ignored_reports > 0:
            self.logger.warning(
                f"{ignored_reports} report plans were not read because of their deprecated version.")

        self.logger.debug("Completed processing all report entries.")
        return reports

    def read_report_schema_into(self, report: dict, schema_df: pd.DataFrame, time_slot_templates: dict):
        self.logger.debug("Starting read_report_schema_into function.")

        try:
            # Clean the DataFrame by dropping empty rows and columns
            schema_df.dropna(how='all', inplace=True)
            schema_df.dropna(axis=1, how="all", inplace=True)
            self.logger.debug(
                "Dropped empty rows and columns from schema DataFrame.")

            # Identify slot columns and section columns
            slot_columns = [
                col for col in schema_df.columns if col.startswith('s:')]
            slot_names = [col.split(':')[1] for col in slot_columns]
            section_column_names = [
                col for col in schema_df.columns if col.startswith('Section ')]
            self.logger.debug(
                f"Found {len(slot_columns)} slot columns and {len(section_column_names)} section columns.")

            # Process slot data
            slots_data = []
            listings_data = []
            order_number = 0
            for slot_column in slot_columns:
                slot_data = {
                    'timeSlotId': slot_column.split(':')[1],
                    'buildListing': len(slot_column.split(':')) > 2 and slot_column.split(':')[2] == "listing",
                    'orderNumber': order_number
                }
                order_number += 1

                # Adjust slot ID if necessary and retrieve slot info
                if re.findall(r'\.\d$', slot_data['timeSlotId']) and slot_data['timeSlotId'][:-2] in slot_names:
                    slot_data['timeSlotId'] = slot_data['timeSlotId'][:-2]
                slot_template = time_slot_templates.get(
                    slot_data['timeSlotId'], {})
                slot_data.update({
                    'name': slot_template.get('name'),
                    'start': slot_template.get('start'),
                    'end': slot_template.get('end')
                })
                slots_data.append(slot_data)
                if slot_data['buildListing']:
                    listings_data.append(slot_data)
                # self.logger.debug(f"Processed slot column '{slot_column}' with order number {slot_data['orderNumber']}.")

            report['slots'] = slots_data
            report['numberSectionLevels'] = len(section_column_names)

            # Process items in schema
            items = []
            order_number = 0
            for row_id in schema_df.index:
                mdict = schema_df.loc[row_id].to_dict()
                item = {'orderNumber': order_number}
                order_number += 1
                hierarchy_location = [
                    mdict[col] for col in section_column_names if not pd.isna(mdict[col])
                ]
                item['hierarchyLocation'] = hierarchy_location

                # Populate item details based on type
                item['name'] = mdict.get('Name', '') if not pd.isna(
                    mdict.get('Name')) else ''
                item['type'] = mdict['Type'].lower() if mdict['Type'].lower() in [
                    'income', 'expense', 'group'] else 'unknown'

                if item['type'] in ['expense', 'income']:
                    cost_loc_str = mdict['Cost Location'] if isinstance(
                        mdict['Cost Location'], str) else f"{int(mdict['Cost Location']):04d}"
                    item['costLocationsStrings'] = cost_loc_str.split(',')
                    item['costLocations'] = [
                        int(cl) for cl in item['costLocationsStrings']]
                    item['hierarchyLocation'].append(item['name'])
                elif item['type'] == 'group' and hierarchy_location:
                    item['name'] = hierarchy_location[-1]

                # Process slot-specific info
                slots_info = []
                for report_slot in report['slots']:
                    row_slot_info = {
                        'timeSlotId': report_slot['timeSlotId'],
                        'buildListing': report_slot['buildListing'],
                        'slotDetails': report_slot,
                        'orderNumber': report_slot['orderNumber']
                    }
                    column_name = f"s:{report_slot['timeSlotId']}"
                    if row_slot_info['buildListing']:
                        column_name += ":listing"

                    # Parse values and set type/limits
                    value = mdict.get(column_name)
                    if not pd.isna(value):
                        if isinstance(value, str):
                            row_slot_info['type'] = 'sum' if value.lower() in ['fieldsum', 'cellsum', 'sum', 'summe'] else 'bookings' if value.lower() in [
                                'bookings', 'buchungen'] else 'budget'
                            if ':' in value:
                                limit, *extra = value.split(':')
                                row_slot_info['limit'] = float(limit)
                                row_slot_info['limitType'] = 'absolute' if extra and extra[0] == '!' else 'exceedable' if extra and extra[0].startswith(
                                    '+') else 'relative'
                                if extra and extra[0].startswith('+') and extra[0].endswith('%'):
                                    row_slot_info['extendedLimit'] = row_slot_info['limit'] * (
                                        1.0 + float(extra[0][1:-1]) / 100.0)
                        elif isinstance(value, (int, float)):
                            row_slot_info.update(
                                {'type': 'budget', 'limit': float(value), 'limitType': 'relative'})
                    else:
                        row_slot_info.update(
                            {'type': 'empty', 'limit': 0.0, 'limitType': 'relative'})

                    slots_info.append(row_slot_info)
                    # self.logger.debug(f"Processed slot info for item {item['name']} with slot ID '{report_slot['timeSlotId']}'.")

                item['slots'] = slots_info
                items.append(item)

            # Collect cost locations and strings
            report['costLocations'] = list({loc for i in items if i['type'] in [
                                           'expense', 'income'] for loc in i['costLocations']})
            report['costLocationsStrings'] = list({str for i in items if i['type'] in [
                                                  'expense', 'income'] for str in i['costLocationsStrings']})

            # Assign children to groups
            for row in items:
                if row['type'] == 'group':
                    row['children'] = [
                        item for item in items
                        if len(item['hierarchyLocation']) == len(row['hierarchyLocation']) + 1
                        and item['hierarchyLocation'][:-1] == row['hierarchyLocation']
                    ]
                    # if row['children']:
                    #     self.logger.debug(f"Assigned {len(row['children'])} children to group '{row['name']}'.")

            report['rows'] = items
            if listings_data:
                report['listings'] = listings_data
            self.logger.debug("Completed processing report schema.")

        except Exception as e:
            self.logger.error(f"Error in read_report_schema_into: {e}")
            raise

    def read_distribution_instructions(self, df: pd.DataFrame):
        self.logger.debug("Starting read_distribution_instructions function.")

        try:
            # Initialize distribution list
            dist = []
            self.logger.debug("Processing each row in the dataframe.")

            for row_id in df.index:
                mdict = df.loc[row_id].to_dict()
                entry = {'channelType': mdict['type'], 'packages': []}
                # self.logger.debug(
                #     f"Processing distribution entry with channel type '{mdict['type']}'.")

                # Check and add recipient if available
                if not pd.isna(mdict.get('recipient')):
                    entry['recipient'] = mdict['recipient']
                    # self.logger.debug(f"Added recipient: {entry['recipient']}")

                # Process trigger information if available
                if not pd.isna(mdict.get('trigger')):
                    trigger_parts = mdict['trigger'].split(':')
                    entry['trigger'] = {'type': trigger_parts[0].strip()}
                    if len(trigger_parts) > 1:
                        entry['trigger']['condition'] = trigger_parts[1].strip()
                    # self.logger.debug(f"Added trigger: {entry['trigger']}")

                # Process content information
                if not pd.isna(mdict.get('content')):
                    content = mdict['content'].split('),')
                    # self.logger.debug(
                    #     f"Processing content for entry with {len(content)} content items.")

                    for item in content:
                        # Parse job type and job list
                        subitems = item.split('(')
                        job_type = subitems[0].strip().lower()
                        job_list_raw = subitems[1].rstrip(')') if len(
                            subitems) > 1 else ""
                        content_items = []

                        # Parse individual jobs within the job list
                        for job in job_list_raw.split(','):
                            job_parts = job.split(':')
                            if len(job_parts) == 2:
                                citem = {'type': job_parts[0].strip(
                                ), 'scope': job_parts[1].strip()}
                                content_items.append(citem)
                                # self.logger.debug(
                                #     f"Added content item: {citem}")
                            else:
                                self.logger.warning(
                                    f"Invalid job format in content item: '{job}'")

                        # Append processed package to entry packages
                        entry['packages'].append(
                            {'packageType': job_type, 'content': content_items})
                        self.logger.debug(
                            f"Added package '{job_type}' with {len(content_items)} content items to entry.")

                # Append completed entry to the distribution list
                dist.append(entry)
                self.logger.debug(
                    f"Completed processing entry with channel type '{mdict['type']}'.")

            self.logger.debug(
                "Completed processing all distribution instructions.")
            return dist

        except Exception as e:
            self.logger.error(f"Error in read_distribution_instructions: {e}")
            raise

    def read_instruction_files(self, container=None):
        self.logger.debug("Starting read_instruction_files function.")

        # Determine container
        container = container or self.conn_clients["templates_folder"]
        self.logger.debug(f"Using container: {container.container_name}")

        # Initialize instructions structure
        instructions = {
            "sheets": {},
            "unprocessedSheets": {},
            "processedSheets": {}
        }

        # List and process blobs in container
        self.logger.debug("Listing blobs in container...")
        instruction_files_list = container.list_blobs()
        file_count = 0

        for instruction_file in instruction_files_list:
            file_count += 1
            blob_client = container.get_blob_client(instruction_file.name)
            self.logger.debug(f"Processing file {instruction_file.name}")

            # Process only .xlsx files
            if instruction_file.name.endswith(".xlsx"):
                self.logger.debug(
                    f"Identified Excel file: {instruction_file.name}")
                dataframes = self.download_all_sheets(blob_client)
                for key, df in dataframes.items():
                    instructions["sheets"][key] = df
                    instructions["unprocessedSheets"][key] = df

        if file_count == 0:
            self.logger.warning("No files found in the container.")
        else:
            self.logger.info(
                f"Total instruction files processed: {file_count}")

        # Time Slots
        if "Time Slots" in instructions["unprocessedSheets"]:
            self.logger.debug("Found 'Time Slots' sheet.")
            instructions["timeSlotTemplates"] = self.read_time_slots(
                instructions["unprocessedSheets"]["Time Slots"]
            )
            instructions["processedSheets"]["Time Slots"] = instructions["unprocessedSheets"]["Time Slots"]
            del instructions["unprocessedSheets"]["Time Slots"]
        else:
            self.logger.error("No 'Time Slots' sheet found in any Excel file.")

        # Kontenrahmen
        kontenrahmen = {k: v for k, v in instructions["unprocessedSheets"].items(
        ) if k.startswith('Kontenrahmen ')}
        if kontenrahmen:
            self.logger.info(
                f"Found {len(kontenrahmen)} 'Kontenrahmen' sheets.")
            instructions["kontenrahmen"] = self.read_kontenrahmen(kontenrahmen)
            instructions["processedSheets"].update(kontenrahmen)
            for key in kontenrahmen:
                del instructions["unprocessedSheets"][key]
        else:
            self.logger.error(
                "No 'Kontenrahmen' sheet found in any Excel file.")

        # Kostenstellenplan
        if "Kostenstellenplan" in instructions["unprocessedSheets"]:
            self.logger.info("Found 'Kostenstellenplan' sheet.")
            instructions["kostenstellenplan"] = self.read_kostenstellenplan(
                instructions["unprocessedSheets"]["Kostenstellenplan"]
            )
            instructions["processedSheets"]["Kostenstellenplan"] = instructions["unprocessedSheets"]["Kostenstellenplan"]
            del instructions["unprocessedSheets"]["Kostenstellenplan"]
        else:
            self.logger.error(
                "No 'Kostenstellenplan' sheet found in any Excel file.")

        # Reports Overview
        if "Reports Overview" in instructions["unprocessedSheets"]:
            self.logger.info("Found 'Reports Overview' sheet.")
            # Uncomment below if reports overview processing is needed
            instructions["reports"] = self.read_reports_overview(
                instructions["unprocessedSheets"]["Reports Overview"]
            )
            instructions["processedSheets"]["Reports Overview"] = instructions["unprocessedSheets"]["Reports Overview"]
            del instructions["unprocessedSheets"]["Reports Overview"]
        else:
            self.logger.error(
                "No 'Reports Overview' sheet found in any Excel file.")

        # Budget Plans
        if "reports" in instructions:
            self.logger.info("Processing reports for budget plans.")
            for key, report in instructions['reports'].items():
                sheet_name = f"Budget Plan {key}"
                if sheet_name in instructions["unprocessedSheets"]:
                    self.read_report_schema_into(
                        report, instructions["unprocessedSheets"][sheet_name], instructions["timeSlotTemplates"]
                    )
                    instructions["processedSheets"][sheet_name] = instructions["unprocessedSheets"][sheet_name]
                    del instructions["unprocessedSheets"][sheet_name]
                else:
                    self.logger.warning(
                        f"Budget Plan {key} not found in unprocessed sheets.")
        else:
            self.logger.error(
                "No 'reports' key found in instructions dictionary.")

        # Flow Plan
        if "Flow Plan" in instructions["unprocessedSheets"]:
            self.logger.info("Found 'Flow Plan' sheet.")
            instructions["distribution"] = self.read_distribution_instructions(
                instructions["unprocessedSheets"]["Flow Plan"]
            )
            instructions["processedSheets"]["Flow Plan"] = instructions["unprocessedSheets"]["Flow Plan"]
            del instructions["unprocessedSheets"]["Flow Plan"]
        else:
            self.logger.error("No 'Flow Plan' sheet found in any Excel file.")

        self.logger.info("Completed reading instruction files.")
        return instructions

    def get_all_bb_posts(self, start_date="2021-01-01", end_date="2039-12-31"):
        self.logger.debug("Starting get_all_bb_posts function.")

        try:
            # Initialize variables for pagination
            base_url = "https://webapp.buchhaltungsbutler.de/api/v1"
            request = '/postings/get'
            url = base_url + request
            offset = 0
            limit = 1000
            retrieved_posts = []

            # Loop to fetch posts until returned rows are less than the limit
            while True:
                # Prepare payload and headers for the request
                payload = json.dumps({
                    "api_key": self.conn_clients['bb_api_key'],
                    "date_from": start_date,
                    "date_to": end_date,
                    "offset": offset,
                    "limit": limit
                })
                headers = {
                    'Content-Type': 'application/json',
                    'Authorization': self.conn_clients['bb_authorization'],
                    'Cookie': 'bbutler=' + self.conn_clients['bb_cookie']
                }

                # Send request to BB API
                self.logger.debug(f"Sending request with offset {offset}.")
                response = requests.request(
                    "POST", url, headers=headers, data=payload)
                self.logger.debug("Request sent, awaiting response.")

                # Parse response
                result = json.loads(response.text)
                if 'rows' in result and 'data' in result:
                    rows = result['rows']
                    retrieved_posts.extend(result['data'])
                    self.logger.debug(
                        f"Retrieved {rows} rows; total collected: {len(retrieved_posts)}.")

                    # If rows returned are fewer than limit, exit loop
                    if rows < limit:
                        self.logger.debug(
                            "Final batch retrieved; exiting loop.")
                        break

                    # Increment offset for next batch
                    offset += limit
                else:
                    # If unexpected response structure, log warning and exit
                    self.logger.warning(
                        "Unexpected response structure received from API.")
                    break

            self.logger.debug("Completed fetching all posts.")
            return retrieved_posts

        except Exception as e:
            self.logger.error(f"Error in get_all_bb_posts: {e}")
            raise

    def get_all_bb_accounts(self):
        self.logger.debug("Starting get_all_bb_accounts function.")

        try:
            # Set up API URL and endpoint
            base_url = "https://webapp.buchhaltungsbutler.de/api/v1"
            request = '/accounts/get'
            url = base_url + request
            self.logger.debug(f"Request URL: {url}")

            # Prepare payload and headers for the request
            payload = json.dumps({
                "api_key": self.conn_clients['bb_api_key']
            })
            headers = {
                'Content-Type': 'application/json',
                'Authorization': self.conn_clients['bb_authorization'],
                'Cookie': 'bbutler=' + self.conn_clients['bb_cookie']
            }

            # Send request to BB API
            self.logger.debug("Sending request to download all BB accounts.")
            response = requests.request(
                "POST", url, headers=headers, data=payload)
            self.logger.debug("Request sent, awaiting response.")

            # Parse response
            result = json.loads(response.text)
            if 'rows' in result and 'data' in result:
                self.logger.debug(
                    f"Downloaded {result['rows']} accounts successfully.")
            else:
                self.logger.warning(
                    "Unexpected response structure received from API.")

            return result.get('data', [])

        except Exception as e:
            self.logger.error(f"Error in get_all_bb_accounts: {e}")
            raise

    def read_expected_bookings(self, container=None):
        self.logger.debug("Starting read_expected_bookings function.")

        try:
            # Initialize the expected bookings structure
            container = container or self.conn_clients["financialplanning_folder"]
            expected_bookings = {
                "sheets": {},
                "unprocessedSheets": {},
                "processedSheets": {}
            }

            # List and check files in the container
            files_list = container.list_blobs()
            filenames = [file.name for file in files_list]
            self.logger.debug(f"Files found in container: {filenames}")

            # Check for "Expected Bookings.xlsx" and process if available
            if "Expected Bookings.xlsx" in filenames:
                blob_client = container.get_blob_client(
                    "Expected Bookings.xlsx")
                self.logger.debug("Found 'Expected Bookings.xlsx'.")

                # Download all sheets from the Excel file
                expected_bookings_dfs = self.download_all_sheets(blob_client)
                for key, df in expected_bookings_dfs.items():
                    expected_bookings["sheets"][key] = df
                    expected_bookings["unprocessedSheets"][key] = df
                    # self.logger.debug(f"Sheet '{key}' added to expected bookings.")

            else:
                self.logger.warning(
                    "'Expected Bookings.xlsx' not found in container.")

            self.logger.debug("Completed reading expected bookings.")
            return expected_bookings

        except Exception as e:
            self.logger.error(f"Error in read_expected_bookings: {e}")
            raise

    def add_more_account_information_to_bookings(self, bookings, kontenrahmen):
        self.logger.debug(
            "Starting add_more_account_information_to_bookings function.")

        for booking in bookings:
            debit_account = int(booking['debit_postingaccount_number'])
            credit_account = int(booking['credit_postingaccount_number'])

            # Process debit account information
            # self.logger.debug(f"Processing debit account information for account number {debit_account}.")
            for account in kontenrahmen:
                if account['accountRangeStart'] <= debit_account <= account['accountRangeEnd']:
                    booking['debit_booking_type_1'] = account['type1']
                    booking['debit_booking_type_2'] = account['type2']
                    categories = [
                        account.get('category1'),
                        account.get('category2'),
                        account.get('category3')
                    ]
                    booking['debit_booking_categories'] = [
                        cat for cat in categories if cat]
                    # self.logger.debug(f"Assigned debit types and categories for account {debit_account}.")
                    break

            # Process credit account information
            # self.logger.debug(f"Processing credit account information for account number {credit_account}.")
            for account in kontenrahmen:
                if account['accountRangeStart'] <= credit_account <= account['accountRangeEnd']:
                    booking['credit_booking_type_1'] = account['type1']
                    booking['credit_booking_type_2'] = account['type2']
                    categories = [
                        account.get('category1'),
                        account.get('category2'),
                        account.get('category3')
                    ]
                    booking['credit_booking_categories'] = [
                        cat for cat in categories if cat]
                    # self.logger.debug(f"Assigned credit types and categories for account {credit_account}.")
                    break

        self.logger.debug(
            "Completed adding account information to all bookings.")

    def collect_all_accounts(self, bookings, kontenrahmen):
        self.logger.debug("Starting collect_all_accounts function.")

        # Initialize dictionary to store account information
        all_accounts = {}

        # Collect all unique debit and credit accounts from bookings
        for booking in bookings:
            debit_account = int(booking['debit_postingaccount_number'])
            credit_account = int(booking['credit_postingaccount_number'])

            # Initialize account entry if it doesn't exist
            if debit_account not in all_accounts:
                all_accounts[debit_account] = {
                    'bookings': [],
                    'type 1': booking['debit_booking_type_1'],
                    'type 2': booking['debit_booking_type_2']
                }
                # self.logger.debug(f"Added new debit account {debit_account}.")

            if credit_account not in all_accounts:
                all_accounts[credit_account] = {
                    'bookings': [],
                    'type 1': booking['credit_booking_type_1'],
                    'type 2': booking['credit_booking_type_2']
                }
                # self.logger.debug(f"Added new credit account {credit_account}.")

        # Sort accounts by account number
        all_accounts = {i: all_accounts[i] for i in sorted(all_accounts)}
        accounts_list = list(all_accounts.keys())
        self.logger.debug(
            f"Found {len(all_accounts)} accounts from {accounts_list[0]} to {accounts_list[-1]}.")

        # Append each booking to its respective debit and credit account
        for booking in bookings:
            all_accounts[int(booking['debit_postingaccount_number'])
                         ]['bookings'].append(booking)
            all_accounts[int(booking['credit_postingaccount_number'])
                         ]['bookings'].append(booking)

        # Add category information and calculate balances for each account
        for account_number, account in all_accounts.items():
            # Find and assign categories from kontenrahmen
            for account_info in kontenrahmen:
                if account_info['accountRangeStart'] <= account_number <= account_info['accountRangeEnd']:
                    categories = [
                        account_info.get('category1'),
                        account_info.get('category2'),
                        account_info.get('category3')
                    ]
                    account['booking_categories'] = [
                        cat for cat in categories if cat]
                    # self.logger.debug(f"Assigned categories to account {account_number}: {account['booking_categories']}")
                    break

            # Sort bookings by date
            account['bookings'].sort(key=lambda x: x['date'])
            # self.logger.debug(f"Sorted bookings for account {account_number} by date.")

            # Initialize financial fields
            account['saldo'] = 0.0
            account['soll'] = 0.0
            account['haben'] = 0.0

            # Calculate soll (debit) and haben (credit) values
            for booking in account['bookings']:
                amount = float(booking['amount'])
                if booking['debit_postingaccount_number'] == str(account_number):
                    account['soll'] += amount
                if booking['credit_postingaccount_number'] == str(account_number):
                    account['haben'] += amount

            # Calculate saldo (balance)
            account['saldo'] = account['soll'] - account['haben']
            # self.logger.debug(
            #     f"Calculated saldo for account {account_number}: {account['saldo']}.")

        self.logger.debug("Completed collecting all accounts.")
        return all_accounts

    def add_up_bookings(self, bookings):
        # self.logger.debug("Starting add_up_bookings function.")

        # Initialize result and category totals
        result = {}
        einnahmen, ausgaben, activa, passiva = 0.0, 0.0, 0.0, 0.0

        # Sum up values based on booking types
        for booking in bookings:
            amount = float(booking['amount'])
            if booking['debit_booking_type_2'] == self.settings["income_term"]:
                einnahmen -= amount
            if booking['credit_booking_type_2'] == self.settings["income_term"]:
                einnahmen += amount
            if booking['debit_booking_type_2'] == self.settings["expense_term"]:
                ausgaben += amount
            if booking['credit_booking_type_2'] == self.settings["expense_term"]:
                ausgaben -= amount
            if booking['debit_booking_type_2'] == 'Activa':
                activa += amount
            if booking['credit_booking_type_2'] == 'Activa':
                activa -= amount
            if booking['debit_booking_type_2'] == 'Passiva':
                passiva -= amount
            if booking['credit_booking_type_2'] == 'Passiva':
                passiva += amount

        # Store results in the result dictionary
        result['revenue'] = einnahmen
        result['expense'] = ausgaben
        result['asset'] = activa
        result['liability'] = passiva
        result['stock'] = activa - passiva
        result['profitAndLoss'] = einnahmen - ausgaben

        # Log calculated results
        # self.logger.debug(f"Calculated totals - Revenue: {result['revenue']}, Expense: {result['expense']}, "
        #                   f"Asset: {result['asset']}, Liability: {result['liability']}, "
        #                   f"Stock: {result['stock']}, Profit and Loss: {result['profitAndLoss']}.")

        # self.logger.debug("Completed add_up_bookings function.")
        return result

    def collect_all_costlocations(self, bookings: list, costlocations: dict):
        self.logger.debug("Starting collect_all_costlocations function.")
        self.logger.debug(
            f"Collecting all cost locations with {len(bookings)} bookings and {len(costlocations)} cost locations.")

        # Initialize dictionary to store cost location information
        all_costlocations = {}

        # Iterate over each booking to categorize by cost location
        for booking in bookings:
            cl = booking['cost_location'] if booking['cost_location'] else 'without'

            # Initialize cost location entry if it doesn't exist
            if cl not in all_costlocations:
                all_costlocations[cl] = {'bookings': []}

                # Assign cost location details if available in the costlocations dictionary
                if str(cl) in costlocations:
                    costlocation_info = costlocations[cl]
                    all_costlocations[cl].update({
                        "limits": costlocation_info.get("limits"),
                        "type": costlocation_info.get("type"),
                        "name": costlocation_info.get("name"),
                        "number": costlocation_info.get("number")
                    })
                    # self.logger.debug(f"Added cost location '{cl}' with details: {all_costlocations[cl]}")

        # Sort cost locations by key
        all_costlocations = {i: all_costlocations[i]
                             for i in sorted(all_costlocations)}
        costlocations_list = list(all_costlocations.keys())
        self.logger.debug(
            f"Found {len(all_costlocations)} cost locations, ranging from '{costlocations_list[0]}' to '{costlocations_list[-1]}'.")

        # Append bookings to their respective cost location
        for booking in bookings:
            if booking['cost_location']:
                all_costlocations[booking['cost_location']
                                  ]['bookings'].append(booking)
            else:
                all_costlocations['without']['bookings'].append(booking)

        # Log the number of bookings without a cost location, if present
        if 'without' in all_costlocations:
            self.logger.debug(
                f"Found {len(all_costlocations['without']['bookings'])} bookings without a cost location.")

        # Sort bookings by date within each cost location and calculate summary information
        for cl, costlocation in all_costlocations.items():
            costlocation['bookings'].sort(key=lambda x: x['date'])
            # self.logger.debug(f"Sorted bookings for cost location '{cl}' by date.")

            # Calculate summary statistics for each cost location
            result = self.add_up_bookings(costlocation['bookings'])
            costlocation.update(result)
            # self.logger.debug(f"Calculated summary for cost location '{cl}'.")

        self.logger.debug("Completed collecting all cost locations.")
        return all_costlocations

    def fill_bookings_to_sheet(self, bookings: list, sheet):
        self.logger.debug("Starting fill_bookings_to_sheet function.")

        def move_a_after_b(lst, a, b):
            if a in lst and b in lst:
                lst.remove(a)
                lst.insert(lst.index(b) + 1, a)

        if bookings:
            # Prepare headers and adjust their order
            headers = list(bookings[0].keys())
            # self.logger.debug("Original headers: " + ", ".join(headers))

            # Reorder headers for better readability in the sheet
            move_a_after_b(headers, 'debit_booking_type_1',
                           'debit_postingaccount_number')
            move_a_after_b(headers, 'debit_booking_type_2',
                           'debit_booking_type_1')
            move_a_after_b(headers, 'debit_booking_categories',
                           'debit_booking_type_2')
            move_a_after_b(headers, 'credit_booking_type_1',
                           'credit_postingaccount_number')
            move_a_after_b(headers, 'credit_booking_type_2',
                           'credit_booking_type_1')
            move_a_after_b(headers, 'credit_booking_categories',
                           'credit_booking_type_2')
            move_a_after_b(headers, 'booking_number', 'id_by_customer')
            move_a_after_b(headers, 'date_delivery', 'comment')
            move_a_after_b(headers, 'date_vat_effective', 'comment')
            move_a_after_b(headers, 'tax_key', 'comment')
            move_a_after_b(headers, 'cost_location', 'date')
            # self.logger.debug("Reordered headers: " + ", ".join(headers))

            # Write headers to the sheet with styling
            crow, ccol = 1, 1
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)
                cell.value = header
                cell.font = styles.Font(bold=True)
                if header in self.settings.get('bookingsHeaderWidth', {}):
                    column_width = self.settings['bookingsHeaderWidth'][header]
                    sheet.column_dimensions[self._colchar(
                        ccol)].width = column_width
                    # self.logger.debug(
                    #     f"Set column width for '{header}' to {column_width}.")
                ccol += 1

            # Write booking data to the sheet
            crow += 1
            for booking in bookings:
                ccol = 1
                for header in headers:
                    cell = sheet.cell(row=crow, column=ccol)

                    try:
                        if header in ['amount', 'vat', 'receipts_assigned_vat_rates', 'receipts_assigned_assigned_amounts']:
                            cell.style = 'Comma'
                            cell.value = float(booking[header])
                        elif header == 'date':
                            cell.value = datetime.datetime.strptime(
                                booking[header], "%Y-%m-%d %H:%M:%S")
                            cell.number_format = 'DD.MM.YY'
                        elif header == 'date_vat_effective':
                            cell.value = datetime.datetime.strptime(
                                booking[header], "%Y-%m-%d")
                            cell.number_format = 'DD.MM.YY'
                        elif header in ['id_by_customer', 'debit_postingaccount_number', 'credit_postingaccount_number', 'tax_key', 'booking_number', 'transactions_id_by_customer']:
                            cell.value = int(booking[header])
                        elif header in ['debit_booking_categories', 'credit_booking_categories']:
                            cell.value = ' - '.join(booking[header])
                        else:
                            cell.value = str(booking[header])
                    except (ValueError, TypeError) as e:  # noqa: e is intentionally ignored
                        cell.value = str(booking[header])
                        # self.logger.warning(
                        #     f"Error formatting cell for '{header}': {e}")

                    ccol += 1
                crow += 1

        self.logger.debug("Completed filling bookings to sheet.")

    def fill_costlocations_to_sheet(self, costlocations: dict, sheet):
        self.logger.debug("Starting fill_costlocations_to_sheet function.")

        # Extract headers from cost locations and adjust their order
        headers = list(costlocations[list(costlocations.keys())[0]].keys())
        headers.remove('bookings')
        headers.remove('type')
        headers.insert(0, 'cost_location')
        self.logger.debug(f"Headers for cost locations: {headers}")

        # Write headers to the sheet with styling
        crow, ccol = 1, 1
        for header in headers:
            cell = sheet.cell(row=crow, column=ccol)
            cell.value = header.replace('type', 'category') if header in [
                'type 1', 'type 2'] else header
            cell.font = styles.Font(bold=True)
            # self.logger.debug(f"Added header '{cell.value}' at row {crow}, column {ccol}.")
            ccol += 1

        # Reset starting row for data population
        crow += 1

        # Populate rows with cost location data
        for number, costlocation in costlocations.items():
            # Skip non-item types
            if 'type' in costlocation and costlocation['type'] != 'item':
                self.logger.debug(
                    f"Skipping cost location '{number}' due to non-item type.")
                continue

            ccol = 1
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)

                # Handle specific headers with customized cell values
                if header == 'cost_location':
                    cell.value = '' if number == 'without' else int(number)
                elif header == 'limits':
                    if header in costlocation:
                        cell.value = ' - '.join([f"{k}: {v}" for k,
                                                v in costlocation[header].items()])
                    else:
                        cell.value = ''
                elif header in ['type', 'name', 'number']:
                    cell.value = costlocation.get(header, '')
                elif header in ['revenue', 'expense', 'asset', 'liability', 'stock', 'profitAndLoss']:
                    try:
                        cell.style = 'Comma'
                        cell.value = float(costlocation[header])
                    except (ValueError, TypeError):
                        cell.value = str(costlocation.get(header, ''))
                        self.logger.warning(
                            f"Non-numeric value for '{header}' in cost location '{number}': {cell.value}")
                else:
                    cell.value = str(costlocation.get(header, ''))

                # self.logger.debug(f"Set value for header '{header}' at row {crow}, column {ccol}.")
                ccol += 1

            crow += 1

        self.logger.debug("Completed filling cost locations to sheet.")

    def fill_accounts_to_sheet(self, accounts: dict, sheet):
        self.logger.debug("Starting fill_accounts_to_sheet function.")

        # Prepare headers and reorder them
        headers = list(accounts[list(accounts.keys())[0]].keys())
        headers.remove('bookings')
        headers.insert(0, 'account')

        # Move 'booking_categories' to follow 'type 2', if present
        if 'booking_categories' in headers and 'type 2' in headers:
            headers.remove('booking_categories')
            headers.insert(headers.index('type 2') + 1, 'booking_categories')
        self.logger.debug(f"Headers for accounts: {headers}")

        # Write headers to the sheet with styling
        crow, ccol = 1, 1
        for header in headers:
            cell = sheet.cell(row=crow, column=ccol)
            cell.value = header.replace('type', 'category') if header in [
                'type 1', 'type 2'] else header
            cell.font = styles.Font(bold=True)
            # self.logger.debug(f"Added header '{cell.value}' at row {crow}, column {ccol}.")
            ccol += 1

        # Reset starting row for data
        crow += 1

        # Populate rows with account data
        for number, account in accounts.items():
            ccol = 1
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)

                # Handle specific headers with customized cell values
                if header == 'account':
                    cell.value = int(number)
                elif header == 'booking_categories':
                    cell.value = ' - '.join(account[header])
                elif header in ['saldo', 'soll', 'haben']:
                    try:
                        cell.style = 'Comma'
                        cell.value = float(account[header])
                    except (ValueError, TypeError):
                        cell.value = str(account[header])
                        self.logger.warning(
                            f"Non-numeric value for '{header}' in account '{number}': {cell.value}")
                else:
                    cell.value = str(account.get(header, ''))

                # self.logger.debug(f"Set value for header '{header}' at row {crow}, column {ccol}.")
                ccol += 1

            crow += 1

        self.logger.debug("Completed filling accounts to sheet.")

    def get_bookings_sheet_name(self, report_id: str):
        """Generate the sheet name for bookings based on report ID."""
        sheet_name = f"{report_id} Buchungen"
        return sheet_name

    def get_listings_sheet_name(self, report_id: str, listing: dict):
        """Generate the sheet name for listings based on report ID and listing name."""
        sheet_name = f"{report_id} Kst-Bericht {listing['name']}"
        return sheet_name

    def fill_report_row_to_listings_sheet(self, report: dict, report_row: dict, sheet: Worksheet, selected_headers: list, current_row: int = 5, first_col: int = 1, depth: int = 1, sub_rows: int = 0):
        # self.logger.debug(f"Starting fill_report_row_to_listings_sheet for report row '{report_row['name']}' with type '{report_row['type']}'.")

        crow = current_row
        col = first_col

        # Handle rows representing budget groups
        if report_row['type'] == 'group':
            level = len(report_row['hierarchyLocation'])
            font = self.settings.get(
                f'depth{depth - level}Font', self.settings['depth5Font'])
            fill = self.settings.get(
                f'level{level}Fill', self.settings['level0Fill'])

            # Set group row styling and name
            cell = sheet.cell(row=crow, column=col + level)
            cell.value = report_row['name']
            cell.font = font
            cell.fill = fill
            # self.logger.debug(f"Set group row '{report_row['name']}' with font and fill at row {crow}, column {col + level}.")

            # Apply font and fill across the entire row span
            for i in range(level, depth + len(selected_headers)):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = font
                cell.fill = fill
            # self.logger.debug(f"Applied group row style across columns for row '{report_row['name']}'.")

        # Handle rows representing actual budget lines (expense or income)
        elif report_row['type'] in ['expense', 'income']:
            col = first_col

            # Apply styling across the budget line row
            for i in range(first_col + depth - 2, depth + len(selected_headers)):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = self.settings['depth1Font']
                cell.fill = self.settings['listingsHeaderFill']
            # self.logger.debug(f"Applied budget line style for row '{report_row['name']}'.")

            # Set name and cost location info
            info_string = f"{report_row['name']} ( cost locations: {', '.join(report_row['costLocationsStrings'])} )" if report_row[
                'costLocationsStrings'] else report_row['name']
            sheet.cell(row=crow, column=col + depth - 1).value = info_string
            # self.logger.debug(f"Set info string for row '{report_row['name']}': {info_string}.")

            # Calculate and set SUM formula for amount column
            if sub_rows > 0:
                amount_col = first_col + depth + len(selected_headers) - 4
                start_cell = sheet.cell(row=crow + 1, column=amount_col)
                end_cell = sheet.cell(row=crow + sub_rows, column=amount_col)
                sum_formula = f"=SUM({start_cell.coordinate}:{end_cell.coordinate})"
                amount_cell = sheet.cell(row=crow, column=amount_col)
                amount_cell.value = sum_formula
                amount_cell.number_format = self.settings['euroFormatPrecise']
            # self.logger.debug(f"Set SUM formula '{sum_formula}' in amount column for row '{report_row['name']}'.")

        # self.logger.debug(f"Completed filling report row '{report_row['name']}' in listings sheet.")

    def select_bookings_by_costlocation(self, cost_locations: list, bookings: list, start_date=None, end_date=None, nur_erfolg=False):
        # self.logger.debug(f"Starting select_bookings_by_costlocation with {len(cost_locations)} cost locations and {len(bookings)} bookings.")

        # Convert cost locations to strings for comparison
        string_cost_locations = [str(cl) for cl in cost_locations]
        selected_bookings = []
        non_erfolg_bookings: list = []
        # Select bookings that match any of the specified cost locations

        for booking in bookings:
            if start_date is not None or end_date is not None:
                booking_date = datetime.datetime.strptime(
                    booking['date'], "%Y-%m-%d %H:%M:%S")
                if start_date is not None:
                    if booking_date < start_date:
                        continue
                if end_date is not None:
                    if booking_date > end_date:
                        continue
                if nur_erfolg and booking['cost_location'] in string_cost_locations or booking['cost_location'] in cost_locations:
                    if not (booking['credit_booking_type_1'] == "Erfolg" or booking['debit_booking_type_1'] == "Erfolg"):
                        non_erfolg_bookings.append(booking)
                        continue
            if booking['cost_location'] in string_cost_locations or booking['cost_location'] in cost_locations:
                selected_bookings.append(booking)

        if nur_erfolg and len(non_erfolg_bookings) > 0:
            self.logger.warning(str(len(non_erfolg_bookings))
                                + " Kostenstellenbuchung ohne Erfolgskonto in Kst. " + str(cost_locations))
            for b in non_erfolg_bookings:
                self.logger.error(
                    "Kostenstellenbuchung ohne Erfolgskonto!!" + str(b))
        # self.logger.debug(f"Selected {len(selected_bookings)} bookings matching specified cost locations.")
        return selected_bookings

    def fill_listings_sheet(self, report: dict, sheet: Worksheet, listing: dict):
        self.logger.debug(
            f"Starting fill_listings_sheet for listing: '{listing['name']}' in report '{report['name']}'.")

        first_report_column = 1
        report_structure_levels = max(self._check_depth(row, 1)
                                      for row in report['rows'])
        self.logger.debug(
            f"Calculated report structure levels: {report_structure_levels}.")

        crow = 1
        selected_headers = [
            'date', 'booking_number', 'cost_location',
            'debit_postingaccount_number', 'credit_postingaccount_number',
            'postingtext', 'amount', 'currency', 'debit_booking_type_2', 'credit_booking_type_2',
        ]
        header_width = self.settings['bookingsHeaderWidth']

        # Add information about the time window
        cell = sheet.cell(row=crow, column=report_structure_levels)
        start_date_string = listing['start'].strftime('%Y-%m-%d')
        end_date_string = listing['end'].strftime('%Y-%m-%d')
        cell.value = f"Time window: from {start_date_string} to {end_date_string}"
        cell.font = styles.Font(bold=True)
        self.logger.debug(
            f"Added time window info: '{cell.value}' at row {crow}.")
        crow += 2

        # Add table headers
        ccol = report_structure_levels + 1
        for header in selected_headers:
            col_letter = self._colchar(ccol)
            if header in header_width:
                sheet.column_dimensions[col_letter].width = header_width[header]
            elif 'default' in header_width and header_width['default'] is not False:
                sheet.column_dimensions[col_letter].width = header_width['default']

            cell = sheet.cell(row=crow, column=ccol)
            cell.value = header
            cell.font = styles.Font(bold=True)
            # self.logger.debug(
            #     f"Added header '{header}' at column {col_letter} with width setting.")
            ccol += 1
        crow += 1

        # Adjust column widths for report structure
        for i in range(first_report_column, report_structure_levels + 1):
            col_letter = self._colchar(i)
            sheet.column_dimensions[col_letter].width = self.settings['groupTitleRowWidth']
        # self.logger.debug(
        #     "Adjusted column widths for report structure levels.")

        # Populate rows with report data
        for row in report['rows']:
            selected_bookings = []
            if row['type'] in ['expense', 'income']:
                selected_bookings = self.select_bookings_by_costlocation(
                    row['costLocations'], report['bookings'],
                    start_date=listing['start'], end_date=listing['end'])
                # self.logger.debug(
                #     f"Selected {len(selected_bookings)} bookings for row '{row['name']}' with type '{row['type']}'.")

            # Fill in report row
            if not (row['type'] == 'group' and row['name'] == ''):
                self.fill_report_row_to_listings_sheet(
                    report, row, sheet, selected_headers, crow,
                    first_col=first_report_column, depth=report_structure_levels, sub_rows=len(
                        selected_bookings)
                )
                # self.logger.debug(
                #     f"Filled report row '{row['name']}' at row {crow}.")
                crow += 1

                # Fill in selected bookings if available
                if row['type'] in ['expense', 'income'] and selected_bookings:
                    self.fill_bookings_listings_sheet(
                        bookings=selected_bookings,
                        sheet=sheet, current_row=crow, selected_headers=selected_headers,
                        first_col=first_report_column + report_structure_levels,
                        row_type=row['type']
                    )
                    crow += len(selected_bookings)
                    # self.logger.debug(
                    #     f"Filled {len(selected_bookings)} bookings for row '{row['name']}'.")

        self.logger.debug(
            f"Completed filling listings sheet for '{listing['name']}' in report '{report['name']}'.")

    def fill_bookings_listings_sheet(self, bookings: list, sheet: Worksheet, current_row: int,
                                     selected_headers: list, first_col: int = 1,
                                     row_type: str = 'expense'):
        self.logger.debug(
            f"Starting fill_bookings_listings_sheet with {len(bookings)} bookings and headers: {selected_headers}.")

        crow = current_row
        # listing_sign = + \
        #     1.0 if self.settings.get("listing_type") == 'expense' else 1.0

        for booking in bookings:
            ccol = first_col
            for header in selected_headers:
                cell = sheet.cell(row=crow, column=ccol)

                # Set cell value and format based on header type
                try:
                    if header in ['amount', 'vat', 'receipts_assigned_vat_rates', 'receipts_assigned_assigned_amounts']:
                        cell.style = 'Comma'
                        sign = 1.0
                        if (booking.get("debit_booking_type_2") == booking.get("credit_booking_type_2") or
                            (booking.get("debit_booking_type_2") == self.settings["income_term"] and
                            booking.get("credit_booking_type_2") == self.settings["expense_term"]) or
                            (booking.get("debit_booking_type_2") == self.settings["expense_term"] and
                            booking.get("credit_booking_type_2") == self.settings["income_term"]) or
                            (booking.get("debit_booking_type_2") not in [self.settings["expense_term"], self.settings["income_term"]] and
                                booking.get("credit_booking_type_2") not in [self.settings["expense_term"], self.settings["income_term"]])):
                            sign = 0.0
                        else:
                            if booking.get("debit_booking_type_2") == self.settings["expense_term"]:
                                if row_type == 'expense':
                                    sign = 1.0
                                if row_type == 'income':
                                    sign = -1.0
                            elif booking.get("debit_booking_type_2") == self.settings["income_term"]:
                                if row_type == 'expense':
                                    sign = 1.0
                                if row_type == 'income':
                                    sign = -1.0
                            elif booking.get("credit_booking_type_2") == self.settings["expense_term"]:
                                if row_type == 'expense':
                                    sign = -1.0
                                if row_type == 'income':
                                    sign = 1.0
                            elif booking.get("credit_booking_type_2") == self.settings["income_term"]:
                                if row_type == 'expense':
                                    sign = -1.0
                                if row_type == 'income':
                                    sign = 1.0
                            else:
                                self.logger.warning(
                                    "tried to add booking to listings sheet with no expense or income booking account type.")
                        if header == 'amount':
                            cell.value = sign * float(booking[header])
                    elif header == 'date':
                        cell.value = datetime.datetime.strptime(
                            booking[header], "%Y-%m-%d %H:%M:%S")
                        cell.number_format = 'DD.MM.YY'
                    elif header == 'date_vat_effective':
                        cell.value = datetime.datetime.strptime(
                            booking[header], "%Y-%m-%d")
                        cell.number_format = 'DD.MM.YY'
                    elif header in ['id_by_customer', 'debit_postingaccount_number', 'credit_postingaccount_number', 'tax_key', 'booking_number', 'transactions_id_by_customer']:
                        cell.value = int(booking[header])
                    elif header in ['debit_booking_categories', 'credit_booking_categories']:
                        cell.value = ' - '.join(booking[header])
                    elif header in ['debit_booking_type_2']:
                        debit_type = booking.get(header, '')
                        if debit_type == self.settings["expense_term"]:
                            debit_type = "Aufwand"
                        elif debit_type == self.settings["income_term"]:
                            debit_type = "Ertrag"
                        cell.value = debit_type + " (Soll) <- an"
                        cell.alignment = styles.Alignment(horizontal='right')
                    elif header in ['credit_booking_type_2']:
                        credit_type = booking.get(header, '')
                        if credit_type == self.settings["expense_term"]:
                            credit_type = "Aufwand"
                        elif credit_type == self.settings["income_term"]:
                            credit_type = "Ertrag"
                        cell.value = credit_type + " (Haben)"
                        cell.alignment = styles.Alignment(horizontal='left')
                    else:
                        cell.value = str(booking[header])
                except (ValueError, TypeError) as e:
                    cell.value = str(booking.get(header, ''))
                    self.logger.warning(
                        f"Error processing '{header}' in booking at row {crow}: {e}")

                # self.logger.debug(f"Set value for header '{header}' at row {crow}, column {ccol}.")
                ccol += 1
            crow += 1

        # self.logger.debug("Completed filling bookings listings sheet.")

    def _add_slot_header(self, sheet, slot: dict, col: int = 1, row: int = 1, depth: int = 1):
        # self.logger.debug(f"Starting _add_slot_header for slot '{slot['name']}' at column {col}, row {row} with depth {depth}.")

        # Set slot title and row height
        sheet.cell(row=row, column=col).value = slot['name']
        sheet.row_dimensions[row].height = self.settings['slotTitleRowHeight']
        # self.logger.debug(f"Set slot title '{slot['name']}' and row height.")

        # Configure start and end date rows
        slot['startDateRow'] = row + 1
        slot['endDateRow'] = slot['startDateRow'] + 1
        # self.logger.debug(f"Configured startDateRow as {slot['startDateRow']} and endDateRow as {slot['endDateRow']}.")

        # Set column widths based on depth
        for i in range(depth):
            column_width = self.settings['slotListColumnWidth'] if i == (
                depth - 1) else self.settings['slotGroupColumnWidth']
            sheet.column_dimensions[self._colchar(
                col + i)].width = column_width
            # self.logger.debug(
            #     f"Set column width for column {self._colchar(col + i)} to {column_width}.")

        # Merge cells based on depth
        if depth > 1:
            sheet.merge_cells(start_row=row, start_column=col,
                              end_row=row, end_column=col + depth - 1)
        if depth > 2:
            sheet.merge_cells(start_row=slot['startDateRow'], start_column=col,
                              end_row=slot['startDateRow'], end_column=col + depth - 2)
            sheet.merge_cells(start_row=slot['endDateRow'], start_column=col,
                              end_row=slot['endDateRow'], end_column=col + depth - 2)
        # self.logger.debug(f"Merged cells for slot '{slot['name']}' based on depth.")

        # Apply font and alignment settings
        title_cell = sheet.cell(row=row, column=col)
        title_cell.alignment = self.settings['slotTitleAlignment']
        title_cell.font = self.settings['slotTitleFont']
        # self.logger.debug(f"Applied title font and alignment for slot '{slot['name']}'.")

        # Apply date fonts and set row heights for date rows
        for cell in sheet[str(slot['startDateRow']) + ":" + str(slot['startDateRow'])]:
            cell.font = self.settings['slotDateFont']
        for cell in sheet[str(slot['endDateRow']) + ":" + str(slot['endDateRow'])]:
            cell.font = self.settings['slotDateFont']
        sheet.row_dimensions[slot['startDateRow']
                             ].height = self.settings['slotDateRowHeight']
        sheet.row_dimensions[slot['endDateRow']
                             ].height = self.settings['slotDateRowHeight']
        # self.logger.debug(f"Set date fonts and row heights for start and end date rows.")

        # Set "from" and "to" labels and corresponding dates if depth > 2
        if depth > 2:
            sheet.cell(row=slot['startDateRow'], column=col).value = "from:"
            sheet.cell(row=slot['endDateRow'], column=col).value = "to:"
        start_date_string = slot['start'].strftime(self.settings['dateFormat'])
        end_date_string = slot['end'].strftime(self.settings['dateFormat'])
        sheet.cell(row=slot['startDateRow'], column=col
                   + depth - 1).value = start_date_string
        sheet.cell(row=slot['endDateRow'], column=col
                   + depth - 1).value = end_date_string
        # self.logger.debug(f"Set date values: start '{start_date_string}', end '{end_date_string}'.")

        # Store slot cell references and boundaries
        slot['startRow'] = row
        slot['endRow'] = row + 2
        slot['startColumn'] = col
        slot['endColumn'] = col + depth - 1
        slot['startDateCellReference'] = f"{self._colchar(slot['endColumn'])}${slot['startDateRow']}"
        slot['endDateCellReference'] = f"{self._colchar(slot['endColumn'])}${slot['endDateRow']}"
        # self.logger.debug(f"Set slot cell references: start '{slot['startDateCellReference']}', end '{slot['endDateCellReference']}'.")

        # self.logger.debug(f"Completed adding slot header for slot '{slot['name']}'.")

    def build_group_summation_formula(self, report_row: dict, slot: dict, report: dict, cell_column: int):
        # self.logger.debug(f"Starting build_group_summation_formula for row '{report_row['name']}' with type '{report_row['type']}'.")

        max_levels = report['structureDepth'] - report['numSaldoRows']
        is_saldo_row = report_row['type'].lower() == "group" and len(
            report_row['hierarchyLocation']) == 0

        # Helper functions to calculate child rows and types
        def all_children(row: dict):
            return sum(1 + all_children(child) for child in row.get('children', []))

        def get_summation_type(row: dict):
            if 'children' in row:
                types = set()
                for child in row['children']:
                    types.update(get_summation_type(child))
                return types
            return {row['type']}

        types = get_summation_type(report_row)
        its_an_item_group = 'children' in report_row and report_row['children'][0]['type'] in [
            'expense', 'income']
        column_to_be_summed = slot['slotDetails']['startColumn'] + \
            max_levels - 1 - (1 if is_saldo_row else 0)

        # Determine the type of summation needed
        if len(types) == 1:
            if its_an_item_group:
                sub_rows = sorted(
                    report_row['children'], key=lambda x: x['rowNumber'])
                if sub_rows:
                    formula = f"=SUM({self._colchar(column_to_be_summed)}{sub_rows[0]['rowNumber']}:{self._colchar(column_to_be_summed)}{sub_rows[-1]['rowNumber']})" if len(
                        sub_rows) > 1 else f"={self._colchar(column_to_be_summed)}{sub_rows[0]['rowNumber']}"
                    # self.logger.debug(f"Generated item group summation formula: '{formula}'")
                    return formula
                else:
                    # self.logger.error(f"No sub rows for item group '{report_row['name']}'")
                    return "=0"
            else:
                # Group of groups summation
                children = report_row['children']
                formula = f"=SUM({self._colchar(column_to_be_summed)}{children[0]['rowNumber']}:{self._colchar(column_to_be_summed)}{children[-1]['rowNumber']})" if len(
                    children) > 1 else f"={self._colchar(column_to_be_summed)}{children[0]['rowNumber']}"
                # self.logger.debug(f"Generated group summation formula: '{formula}'")
                return formula

        elif len(types) > 1 and any(t in types for t in ['expense', 'income', '']):
            if 'children' in report_row:
                formula_parts = [
                    f"{'+' if 'income' in get_summation_type(child) else '-'}{self._colchar(column_to_be_summed)}{child['rowNumber']}"
                    for child in report_row['children']
                ]
                formula = "=" + "".join(formula_parts)
                # self.logger.debug(f"Generated mixed-type summation formula: '{formula}'")
                return formula
            else:
                # self.logger.error(f"Error: two sub item types with issues in '{report_row['name']}'")
                return "Error: two sub item types but problems"

        # self.logger.debug(f"Completed formula generation for row '{report_row['name']}'.")
        return ""

    def _get_column_letter_by_column_header(self, sheet, header: str):
        for cell in sheet["1:1"]:
            if cell.value == header:
                return cell.column_letter

    def fill_row_to_sheet(self, report: dict, report_row: dict, sheet: Worksheet, bookings_sheet: str, first_col: int = 1, depth: int = 1):
        # self.logger.debug(
        #     f"Starting fill_row_to_sheet for report row '{report_row['name']}' with type '{report_row['type']}'.")
        crow = report_row['rowNumber']
        col = first_col

        # Handle budget group rows
        if report_row['type'] == 'group':
            level = len(report_row['hierarchyLocation'])
            font = self.settings.get(
                f'depth{depth - level}Font', self.settings['depth5Font'])
            fill = self.settings.get(
                f'level{level}Fill', self.settings['level0Fill'])

            # Set row title and format cells
            cell = sheet.cell(row=crow, column=col + level)
            cell.value = report_row['name']
            cell.font = font
            cell.fill = fill
            # self.logger.debug(f"Set group title '{report_row['name']}' with font and fill at row {crow}, column {col + level}.")

            # Apply formatting across columns for the group row
            for i in range(level, depth + 1 + depth * len(report_row['slots'])):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = font
                cell.fill = fill

            # Fill slots for the group row
            for slot_number, slot in enumerate(report_row['slots']):
                start_column = slot['slotDetails']['startColumn']
                end_column = slot['slotDetails']['endColumn']
                cell_column = start_column + (level - 1 if level > 0 else 0)
                cell = sheet.cell(row=crow, column=cell_column)

                # Set value or formula based on slot type
                if slot['type'] == 'budget' and 'limit' in slot:
                    cell.value = slot['limit']
                    cell.number_format = self.settings['euroFormat']
                else:  # Summary row with formula
                    formula = self.build_group_summation_formula(
                        report_row, slot, report, cell_column)
                    cell.value = formula
                    cell.number_format = self.settings['euroFormat']
                sheet.merge_cells(
                    start_row=crow, start_column=cell_column, end_row=crow, end_column=end_column)

            # self.logger.debug(f"Filled slots for group row '{report_row['name']}'.")

        # Handle actual budget lines (expense or income) at the deepest level
        elif report_row['type'] in ['expense', 'income'] and bookings_sheet:
            sheet.cell(row=crow, column=col + depth
                       - 1).value = report_row['name']
            sheet.cell(row=crow, column=col
                       + depth).value = ', '.join(report_row['costLocationsStrings'])

            # Define ranges for bookings formula
            bookings_start_row = 2
            bookings_end_row = len(report['bookings']) + 100
            bsn = self.get_bookings_sheet_name(report['id'])
            amount_cell_column_letter = self._get_column_letter_by_column_header(
                bookings_sheet, self.settings['bookingsAmountColumnHeader'])
            amount_cell_range = f"{amount_cell_column_letter}{bookings_start_row}:{amount_cell_column_letter}{bookings_end_row}"
            costlocation_column_letter = self._get_column_letter_by_column_header(
                bookings_sheet, self.settings['bookingsCostlocationColumnHeader'])
            costlocation_cell_range = f"{costlocation_column_letter}{bookings_start_row}:{costlocation_column_letter}{bookings_end_row}"
            date_column_letter = self._get_column_letter_by_column_header(
                bookings_sheet, self.settings['bookingsDateColumnHeader'])
            date_cell_range = f"{date_column_letter}{bookings_start_row}:{date_column_letter}{bookings_end_row}"
            dbt2_column_letter = self._get_column_letter_by_column_header(
                bookings_sheet, self.settings['bookingsDebitBookingsType2'])
            dbt2_cell_range = f"{dbt2_column_letter}{bookings_start_row}:{dbt2_column_letter}{bookings_end_row}"
            cbt2_column_letter = self._get_column_letter_by_column_header(
                bookings_sheet, self.settings['bookingsCreditBookingsType2'])
            cbt2_cell_range = f"{cbt2_column_letter}{bookings_start_row}:{cbt2_column_letter}{bookings_end_row}"
            # self.logger.debug(
            #     f"Configured cell ranges for bookings data in '{bsn}'.")

            # Define functions to create formulas based on slot details and cost location checks
            def formula_opening_cl_cell(slot_details, current_row, cell_range, cell_range_search_term):
                return (f"SUMIFS('{bsn}'!{amount_cell_range},'{bsn}'!{costlocation_cell_range},$"
                        f"{self._colchar(col + depth)}{current_row},'{bsn}'!{date_cell_range},\">=\"&{slot_details['startDateCellReference']},"
                        f"'{bsn}'!{date_cell_range},\"<\"&{slot_details['endDateCellReference']}+1,'{bsn}'!{cell_range},\"{cell_range_search_term}\")")

            def formula_opening_cl_value(slot_details, cost_location, cell_range, cell_range_search_term):
                return (f"SUMIFS('{bsn}'!{amount_cell_range},'{bsn}'!{costlocation_cell_range},\"{cost_location}\","
                        f"'{bsn}'!{date_cell_range},\">=\"&{slot_details['startDateCellReference']},"
                        f"'{bsn}'!{date_cell_range},\"<\"&{slot_details['endDateCellReference']}+1,'{bsn}'!{cell_range},\"{cell_range_search_term}\")")

            # Generate formula for each slot in the row
            for slot in report_row['slots']:
                slot_details = slot['slotDetails']
                cell = sheet.cell(row=crow, column=slot_details['endColumn'])

                # Build formula based on cost location count
                if slot['type'] == 'budget' and 'limit' in slot:
                    cell.value = slot['limit']
                    cell.number_format = self.settings['euroFormat']
                elif slot['type'] == 'bookings' and report['bookings']:
                    formula = ""
                    if len(report_row["costLocationsStrings"]) > 1:
                        formula_parts = [
                            f"{sign}{formula_opening_cl_value(slot_details, cl, cell_range, term)}"
                            for cl in report_row["costLocationsStrings"]
                            for sign, cell_range, term in [
                                ("+", dbt2_cell_range, self.settings["expense_term"]), ("-",
                                                                                        cbt2_cell_range, self.settings["income_term"]),
                                ("+", dbt2_cell_range, self.settings["income_term"]), ("-",
                                                                                       cbt2_cell_range, self.settings["expense_term"])
                            ]
                        ]
                        formula = "+".join(formula_parts)
                    else:
                        formula = (
                            f"{formula_opening_cl_cell(slot_details, crow, dbt2_cell_range, self.settings['expense_term'])}"
                            f"-{formula_opening_cl_cell(slot_details, crow, cbt2_cell_range, self.settings['income_term'])}"
                            f"+{formula_opening_cl_cell(slot_details, crow, dbt2_cell_range, self.settings['income_term'])}"
                            f"-{formula_opening_cl_cell(slot_details, crow, cbt2_cell_range, self.settings['expense_term'])}"
                        )

                    if report_row['type'] == 'income':
                        formula = f"-({formula})"
                    cell.value = f"={formula}"
                    # print("filling bookings summation cell",
                    #       cell.value, cell.coordinate)

                    cell.number_format = self.settings['euroFormat']
                    cell.alignment = styles.Alignment(horizontal="right")
                    # self.logger.debug(
                    #     f"Set formula for slot '{slot['name']}' in row '{report_row['name']}'.")

        # self.logger.debug(f"Completed filling row '{report_row['name']}' in sheet.")

    def _set_border_to_area(self,
                            sheet, start_column=0, start_row=0, end_column=5, end_row=5,
                            side: styles.Side = styles.Side(
            border_style='thin', color='FF000000'),
            top=True, bottom=True, left=True, right=True):

        if top:
            for col in range(start_column, end_column + 1):
                sheet.cell(row=start_row, column=col).border = styles.Border(
                    top=side)
        if bottom:
            for col in range(start_column, end_column + 1):
                sheet.cell(row=end_row, column=col).border = styles.Border(
                    bottom=side)
        if left:
            for row in range(start_row, end_row + 1):
                sheet.cell(row=row, column=start_column).border = styles.Border(
                    left=side)
        if right:
            for row in range(start_row, end_row + 1):
                sheet.cell(row=row, column=end_column).border = styles.Border(
                    right=side)
        if top and left:
            sheet.cell(row=start_row, column=start_column).border = styles.Border(
                top=side, left=side)
        if top and right:
            sheet.cell(row=start_row, column=end_column).border = styles.Border(
                top=side, right=side)
        if bottom and left:
            sheet.cell(row=end_row, column=start_column).border = styles.Border(
                bottom=side, left=side)
        if bottom and right:
            sheet.cell(row=end_row, column=end_column).border = styles.Border(
                bottom=side, right=side)

    def fill_report_to_sheet(self, report: dict, sheet, bookings_sheet, listings_sheets=[]):
        self.logger.debug(
            f"Starting fill_report_to_sheet for report: '{report['name']}'.")

        # Fill bookings sheet with data
        self.fill_bookings_to_sheet(report['bookings'], bookings_sheet)
        self.logger.debug("Filled bookings sheet with report data.")

        # Fill listings sheets if they exist
        if listings_sheets:
            listings_sheets_names = [sheet.title for sheet in listings_sheets]
            if "listings" in report and report["listings"]:
                for listing in report["listings"]:
                    listing_sheet_name = self.get_listings_sheet_name(
                        report['id'], listing)
                    if listing_sheet_name in listings_sheets_names:
                        sheet_index = listings_sheets_names.index(
                            listing_sheet_name)
                        self.fill_listings_sheet(
                            report, listings_sheets[sheet_index], listing)
                        # self.logger.debug(f"Filled listings sheet '{listing_sheet_name}' with data from report '{report['name']}'.")

        # Add title and information to the main sheet
        crow, ccol = 1, 1
        cell = sheet.cell(row=crow, column=ccol)
        cell.value = report['name']
        sheet.merge_cells(start_row=crow, start_column=ccol,
                          end_row=crow, end_column=ccol + 10)
        cell.font = self.settings['titleFont']
        self.logger.debug(f"Added title '{report['name']}' to sheet.")

        # Display compilation information and last booking date
        last_date_str = sorted(
            report['bookings'], key=lambda booking: booking['date'])[-1]['date']
        last_date = datetime.datetime.strptime(
            last_date_str, '%Y-%m-%d %H:%M:%S')
        info_text = f"compiled: {datetime.datetime.today().strftime('%d.%m.%Y')} with last booking from {last_date.strftime('%d.%m.%Y')}"
        sheet.cell(row=crow + 1, column=ccol).value = info_text
        self.logger.debug(
            "Added compilation and last booking date information.")

        # Set row and column settings for report rows and slots
        crow += 2
        rows = report['rows']
        report_structure_levels = report['structureDepth'] - \
            report['numSaldoRows']
        first_slot_col = report_structure_levels + 2
        rows_forward_per_row = self.settings['rowsForwardForAllRows']

        # Add headers for each slot
        for slot_number, slot in enumerate(report['slots']):
            slot_col = first_slot_col + slot_number * report_structure_levels
            self._add_slot_header(sheet, slot, col=slot_col,
                                  row=crow, depth=report_structure_levels)
            # self.logger.debug(f"Added slot header for slot {slot_number} at column {slot_col}, row {crow}.")

        crow += 4
        first_data_row = crow

        # Set row numbers for report rows
        for row in rows:
            if row['type'].lower() == "group" and not row['hierarchyLocation']:
                crow += 1
            row['rowNumber'] = crow
            crow += rows_forward_per_row
            for slot in row['slots']:
                slot['slotDetails']['endRow'] = max(
                    slot['slotDetails']['endRow'], crow)
        crow = first_data_row

        # Adjust column widths for report structure levels
        for i in range(report_structure_levels):
            col_letter = self._colchar(i + 1)
            sheet.column_dimensions[col_letter].width = self.settings['groupTitleRowWidth']
        sheet.column_dimensions[self._colchar(
            report_structure_levels)].width = self.settings['itemTitleRowWidth']
        sheet.column_dimensions[self._colchar(
            report_structure_levels + 1)].width = self.settings['costLocationColumnWidth']

        self.logger.debug(f"report keys: {report.keys()}")

        # # Populate each row in the report
        for row in rows:
            self.fill_row_to_sheet(
                report, row, sheet, bookings_sheet, first_col=1, depth=report_structure_levels)
            # self.logger.debug(f"Filled row {row['rowNumber']} for report '{report['name']}'.")

        # Add borders around slots in the sheet
        for slot in report['slots']:
            self._set_border_to_area(sheet, slot['startColumn'], slot['startRow'],
                                     slot['endColumn'], slot['endRow'] - report['numSaldoRows'])
            # self.logger.debug(f"Set border for slot starting at column {slot['startColumn']}, row {slot['startRow']}.")
        self.logger.debug(
            f"Completed filling report '{report['name']}' to sheet.")

    def fill_personnel_bookings_to_sheet(self, bookings: list, sheet: Worksheet):
        if not bookings:
            self.logger.debug(
                "No bookings provided, exiting fill_personnel_bookings_to_sheet.")
            return

        # Extract headers and categorize personnel-related and other columns
        sample_columns_settings: dict = {'col_forward': 1, 'col_index': 0}
        headers = {
            header: sample_columns_settings.copy() for header in bookings[0].keys()}

        # get lists of accounts relevant for payroll and personnel bookings
        # clearing accounts should sum up to 0 saldo for each month
        clearing_accounts: dict[str, str] = {
            key: value for key, value in self.settings['skr49_payroll_clearing_accounts'].items()
            if key in headers
        }
        # liabilities accounts should be sum up to 0 as soon as payments to tax department and social insurance are done - normally monthly
        liabilities_accounts: dict[str,
                                   str] = {
            key: value for key, value in self.settings['skr49_payroll_liabilities_accounts'].items()
            if key in headers
        }
        # personnel expenses ideeller Bereich
        personnel_accounts_ideell: dict[str,
                                        str] = {
            key: value for key, value in self.settings['skr49_personnel_expenses_non_profit'].items()
            if key in headers
        }
        # personnel expenses purpose operations
        personnel_accounts_zweck: dict[str,
                                       str] = {
            key: value for key, value in self.settings['skr49_personnel_expenses_vat_exempt_purpose'].items()
            if key in headers
        }
        account_groups = [clearing_accounts, liabilities_accounts,
                          personnel_accounts_ideell, personnel_accounts_zweck]
        all_accounts: list[str] = [
            account for accounts in account_groups for account in accounts.keys()]
        # Update col_forward to 2 for the first key in each dictionary in account_groups
        for group in account_groups:
            if group:  # Ensure the dictionary is not empty
                first_header = next(iter(group))  # Get the first key
                if first_header in headers:  # Check if the header exists in headers
                    headers[first_header]['col_forward'] = 2

        # Add columns for cost location columns
        cost_location_groups: dict[str, str] = {
            "P" + key: value for key, value in self.settings['devint_cost_location_groups'].items()
            if "P" + key in headers
        }
        all_costlocation_groups: list[str] = [
            group for group in cost_location_groups.keys()]
        first_header = next(iter(cost_location_groups))  # Get the first key
        if first_header in headers:  # Check if the header exists in headers
            headers[first_header]['col_forward'] = 2

        # Start building table header
        check_column_width = self.settings['bookingsHeaderWidth']['amount'] - 4
        self.logger.debug("Starting to build table header.")
        crow, ccol = 1, 0
        for header, h_settings in headers.items():
            ccol += h_settings['col_forward']
            headers[header]['col_index'] = ccol
            cell = sheet.cell(row=crow, column=ccol)
            cell.value = header
            cell.font = styles.Font(bold=True)
            if header in self.settings.get('bookingsHeaderWidth', {}):
                column_width = self.settings['bookingsHeaderWidth'][header]
                sheet.column_dimensions[self._colchar(
                    ccol)].width = column_width
            elif header in all_accounts:
                sheet.column_dimensions[self._colchar(
                    ccol)].width = check_column_width
            elif header in [p for p in cost_location_groups.keys()]:
                sheet.column_dimensions[self._colchar(
                    ccol)].width = check_column_width

        crow += 1
        ccol = 1
        last_month = bookings[0]['month']
        self.logger.debug(
            f"Starting to fill booking entries with initial month set to {last_month}.")
        maxrow = 0
        last_saldo_row = 0
        for booking in bookings:
            if last_month != booking['month']:
                cell = sheet.cell(
                    row=crow, column=headers['postingtext']['col_index'])
                cell.value = "Saldo Monat"
                # New month data started, complete rows for last month
                num_entries_last_month = sum(
                    b['month'] == last_month for b in bookings)
                # for sum_header in self.settings['personnel_monthly_sum_columns']:
                for header, h_settings in headers.items():
                    if header in all_accounts or header in all_costlocation_groups:
                        col_index = headers[header]['col_index']
                        cell = sheet.cell(row=crow, column=col_index)
                        cell.value = f"=SUM({self._colchar(col_index)}{crow - num_entries_last_month}:{self._colchar(col_index)}{crow - 1})"
                        cell.font = styles.Font(bold=True)
                        cell.style = 'Comma'
                        cell.number_format = self.settings['euroFormatPrecise']
                        cell.border = styles.Border(top=styles.Side(
                            border_style='thin', color='FF000000'))

                crow += 1
                cell = sheet.cell(
                    row=crow, column=headers['postingtext']['col_index'])
                cell.value = "Rolling Saldo Total"
                cell.font = styles.Font(bold=True)
                cell.style = 'Comma'
                for header, h_settings in headers.items():
                    if header in {**clearing_accounts, **liabilities_accounts}.keys():
                        col_index = headers[header]['col_index']
                        cell = sheet.cell(row=crow, column=col_index)
                        cell.value = f"={self._colchar(col_index)}{crow - 1}"
                        if last_saldo_row:
                            cell.value += f"+{self._colchar(col_index)}{last_saldo_row}"
                        cell.font = styles.Font(bold=True)
                        cell.style = 'Comma'
                        cell.number_format = self.settings['euroFormatPrecise']
                        cell.fill = styles.PatternFill(
                            start_color="EDF4FF", end_color="E2EBFB", fill_type="solid")
                last_saldo_row = crow
                crow += 2
                last_month = booking['month']
                self.logger.debug(
                    f"Updated row for month change to {last_month}.")

            ccol = 1
            for header, h_settings in headers.items():
                cell = sheet.cell(
                    row=crow, column=headers[header]['col_index'])
                try:
                    if header in ['amount', 'vat', 'receipts_assigned_vat_rates', 'receipts_assigned_assigned_amounts']:
                        cell.style = 'Comma'
                        cell.value = float(booking[header])
                        cell.number_format = self.settings['euroFormatPrecise']
                    elif header == 'date':
                        cell.value = datetime.strptime(
                            booking[header], "%Y-%m-%d %H:%M:%S") if isinstance(booking[header], str) else booking[header]
                        cell.number_format = 'DD.MM.YY'
                    elif header == 'date_vat_effective':
                        cell.value = datetime.strptime(
                            booking[header], "%Y-%m-%d")
                        cell.number_format = 'DD.MM.YY'
                    elif header in ['id_by_customer', 'debit_postingaccount_number', 'credit_postingaccount_number', 'tax_key', 'booking_number', 'transactions_id_by_customer']:
                        cell.value = int(booking[header])
                    elif header in ['debit_booking_categories', 'credit_booking_categories']:
                        cell.value = ' - '.join(booking[header])
                    elif booking[header] and (header in all_accounts or header in all_costlocation_groups):
                        cell.value = booking[header]
                        cell.style = 'Comma'
                        cell.number_format = self.settings['euroFormatPrecise']
                    else:
                        cell.value = str(booking[header])
                except (ValueError, TypeError) as e:
                    self.logger.warning(
                        f"Error processing '{header}' in booking at row {crow}: {e}")
                    cell.value = str(booking.get(header, ''))
                ccol += 1

            crow += 1
            if crow > maxrow:
                maxrow = crow

        # add borders to groups
        for group in account_groups:
            if group:
                first_col_index = headers[next(iter(group))]['col_index']
                last_col_index = headers[next(reversed(group))]['col_index']
                self._set_border_to_area(sheet, first_col_index, 1,
                                         last_col_index, maxrow, side=styles.Side(border_style='medium'))
        first_col_index = headers[next(
            iter(cost_location_groups))]['col_index']
        last_col_index = headers[next(
            reversed(cost_location_groups))]['col_index']
        self._set_border_to_area(sheet, first_col_index, 1,
                                 last_col_index, maxrow, side=styles.Side(border_style='medium'))

        sheet.freeze_panes = sheet.cell(row=2,
                                        column=headers['booking_number']['col_index']
                                        )
        self.logger.debug(
            "Completed filling personnel bookings into the sheet.")

    def build_personnel_bookings(self, bookings: list, logger=None):
        self.logger.info("Starting to build personnel bookings DataFrame.")

        # get lists of accounts relevant for payroll and personnel bookings
        # clearing accounts should sum up to 0 saldo for each month
        clearing_accounts: dict[str,
                                str] = self.settings['skr49_payroll_clearing_accounts']
        # liabilities accounts should be sum up to 0 as soon as payments to tax department and social insurance are done - normally monthly
        liabilities_accounts: dict[str,
                                   str] = self.settings['skr49_payroll_liabilities_accounts']
        # personnel expenses ideeller Bereich
        personnel_accounts_ideell: dict[str,
                                        str] = self.settings['skr49_personnel_expenses_non_profit']
        # personnel expenses purpose operations
        personnel_accounts_zweck: dict[str,
                                       str] = self.settings['skr49_personnel_expenses_vat_exempt_purpose']
        account_groups = [clearing_accounts, liabilities_accounts,
                          personnel_accounts_ideell, personnel_accounts_zweck]

        # Create DataFrame and filter relevant columns
        p_bookings = pd.DataFrame(bookings)
        p_bookings = p_bookings[[
            col for col in p_bookings.columns if col in self.settings['keep_personnel_columns']]]
        self.logger.debug("Filtered personnel columns for DataFrame.")

        # Convert columns to appropriate data types
        p_bookings['date'] = pd.to_datetime(
            p_bookings['date'], errors='coerce')
        int_columns = ['booking_number', 'debit_postingaccount_number',
                       'credit_postingaccount_number', 'transaction_id_by_customer']
        for col in int_columns:
            p_bookings[col] = p_bookings[col].astype('Int64', errors='ignore')
        self.logger.debug("Converted date and integer columns in DataFrame.")

        cost_location_cols_nums: list[str] = [
            k for k in self.settings['devint_cost_location_groups'].keys()]

        # Define month grouping logic
        month_abbreviations = {'Jan.': 1, 'Feb.': 2, 'Mrz.': 3, 'Apr.': 4, 'Mai': 5,
                               'Juni': 6, 'Juli': 7, 'Aug.': 8, 'Sept.': 9, 'Okt.': 10, 'Nov.': 11, 'Dez.': 12}

        def month_group(row):
            result = row['date'].strftime(
                '%Y-%m') if pd.notnull(row['date']) else None
            if 'STEUERVERWALTUNG' in row.get('postingtext', '') and row['debit_postingaccount_number'] == 1700 and row['credit_postingaccount_number'] == 940:
                match = re.search(
                    r'Lohnsteuer\s+([A-Za-z]+\.?)\s?(\d{2})', row['transaction_purpose'])
                if match:
                    month = month_abbreviations.get(
                        match.group(1).strip(), None)
                    year = f"20{match.group(2).strip()}"
                    result = f"{year}-{month:02}" if month else result
            return result

        # Apply liability calculations
        def calculate_liability(row, liability):
            amount = float(str(row['amount']).replace(
                '.', '').replace(',', '.'))
            if str(row['debit_postingaccount_number']) == liability:
                return amount
            elif str(row['credit_postingaccount_number']) == liability:
                return -amount
            return 0

        def calculate_costlocation_group_expense(row, costlocation_group: str):
            if row['cost_location'].startswith(costlocation_group):
                # float(str(row['amount']) #.replace('.', '').replace(',', '.'))
                amount = float(row['amount'])
                if (row['debit_booking_type_1'] == "Erfolg") ^ (row['credit_booking_type_1'] == "Erfolg"):
                    if row['debit_booking_type_2'] == self.settings['expense_term'] or row['credit_booking_type_2'] == self.settings['income_term']:
                        return amount
                    elif row['debit_booking_type_2'] == self.settings['income_term'] or row['credit_booking_type_2'] == self.settings['expense_term']:
                        return -amount
                    else:
                        return 0.0
                else:
                    return 0.0
            else:
                return 0.0

        # Filter bookings by relevant accounts
        relevant_accounts: list[int] = [
            int(account) for accounts in account_groups for account in accounts.keys()]
        mask = p_bookings['debit_postingaccount_number'].isin(
            relevant_accounts) | p_bookings['credit_postingaccount_number'].isin(relevant_accounts)
        p_bookings = p_bookings[mask]
        self.logger.debug(
            "Filtered bookings based on relevant account numbers.")

        # Add 'month' column and sort
        p_bookings['month'] = p_bookings.apply(month_group, axis=1)
        p_bookings = p_bookings.sort_values(
            by=['month', 'date', 'debit_postingaccount_number'], ascending=[True, True, True])
        self.logger.debug("Assigned 'month' values and sorted the DataFrame.")

        for account, description in {key: value for group in account_groups for key, value in group.items()}.items():
            p_bookings[account] = p_bookings.apply(
                lambda row: calculate_liability(row, account) / 100.0, axis=1)
            # Check if all values in the column are 0, and drop the column if true
            if (p_bookings[account] == 0).all():
                p_bookings = p_bookings.drop(columns=[account])
            self.logger.debug(
                f"Calculated liability for column {account}:{description}.")

        for costlocation_group in cost_location_cols_nums:
            p_bookings[f"P{costlocation_group}"] = p_bookings.apply(
                lambda row: calculate_costlocation_group_expense(row, costlocation_group), axis=1)
            self.logger.debug(
                f"Calculated costlocation group expenses for group {costlocation_group}.")

        self.logger.debug("Completed personnel bookings DataFrame processing.")
        return p_bookings.to_dict(orient='records')

    def compile_all_xls_sheets(self, reports: dict, instructions: dict):
        self.logger.debug("Starting compile_all_xls_sheets function.")

        # Initialize the workbook and remove any default sheet
        wb = Workbook()
        if wb.worksheets:
            try:
                wb.remove(wb.worksheets[0])
                self.logger.debug(
                    "Removed default worksheet from the workbook.")
            except AttributeError:
                wb.remove(wb.worksheets[0])
                self.logger.debug(
                    "Removed default worksheet (legacy method) from the workbook.")

        # Add "Buchungen" sheet and populate with bookings data
        ws = wb.create_sheet("Buchungen")
        self.logger.debug("Created 'Buchungen' sheet.")
        self.fill_bookings_to_sheet(reports['bookings'], ws)
        self.logger.debug("Filled 'Buchungen' sheet with booking data.")

        # Add "Kostenstellen" sheet and populate with cost location data
        ws = wb.create_sheet("Kostenstellen")
        self.logger.debug("Created 'Kostenstellen' sheet.")
        self.fill_costlocations_to_sheet(reports['costlocations'], ws)
        self.logger.debug(
            "Filled 'Kostenstellen' sheet with cost location data.")

        # Add "Konten" sheet and populate with account data
        ws = wb.create_sheet("Konten")
        self.logger.debug("Created 'Konten' sheet.")
        self.fill_accounts_to_sheet(reports['accounts'], ws)
        self.logger.debug("Filled 'Konten' sheet with account data.")

        # Add "PK Buchungen" sheet for personnel bookings
        ws = wb.create_sheet("PK Buchungen")
        self.logger.debug("Created 'PK Buchungen' sheet.")
        self.fill_personnel_bookings_to_sheet(reports['personnelbookings'], ws)
        self.logger.debug(
            "Filled 'PK Buchungen' sheet with personnel booking data.")

        # Loop through each report in instructions and add relevant sheets
        report_ids = list(instructions['reports'].keys())

        # deleteme
        # report_ids = []
        for report_id in report_ids:
            if report_id in reports:
                self.logger.debug(f"Creating sheets for report '{report_id}'.")

                # Create main report sheet
                ws = wb.create_sheet(report_id)
                self.logger.debug(f"Created main report sheet '{report_id}'.")

                # Create bookings sheet for the report
                wsb = wb.create_sheet(self.get_bookings_sheet_name(report_id))
                self.logger.debug(
                    f"Created bookings sheet for report '{report_id}'.")

                # Create listings sheets if available
                listings_sheets = []
                if "listings" in instructions['reports'][report_id]:
                    for listing in instructions['reports'][report_id]['listings']:
                        wsl = wb.create_sheet(
                            self.get_listings_sheet_name(report_id, listing))
                        listings_sheets.append(wsl)
                        self.logger.debug(
                            f"Created listing sheet '{wsl.title}' for report '{report_id}'.")

                # Populate report, bookings, and listings sheets
                self.fill_report_to_sheet(
                    reports[report_id], ws, wsb, listings_sheets)
                self.logger.debug(f"Filled sheets for report '{report_id}'.")

        # Save workbook to a virtual file and return it
        self.logger.debug("Completed compiling all sheets into the workbook.")
        return save_virtual_workbook(wb)

    def build_reports(self,
                      bookings: list,
                      expected_bookings: list,
                      instructions: dict,
                      accounts: list):
        self.logger.debug("Starting build_reports function.")

        # Initialize the reports dictionary
        reports = {}

        # Add account information to bookings
        self.logger.debug("Adding account information to bookings.")
        self.add_more_account_information_to_bookings(
            bookings, instructions['kontenrahmen']['default'])

        # Mark each booking as "booked"
        for booking in bookings:
            booking['realisation'] = 'booked'
        reports['bookings'] = bookings

        # Build personnel bookings report
        reports['personnelbookings'] = self.build_personnel_bookings(bookings)

        # Collect accounts and cost locations for reports
        self.logger.debug("Collecting all accounts.")
        reports['accounts'] = self.collect_all_accounts(
            bookings, instructions['kontenrahmen']['default'])
        self.logger.debug("Collecting all cost locations.")
        reports['costlocations'] = self.collect_all_costlocations(
            bookings, instructions['kostenstellenplan'])

        # Build individual reports based on the provided instructions
        for report_id, report in instructions['reports'].items():
            self.logger.debug(f"Building report '{report['name']}'.")

            # Add relevant bookings to the report dictionary
            report_bookings = [
                booking for booking in bookings if booking['cost_location'] in report['costLocationsStrings']]
            reports[report_id] = report
            reports[report_id]['bookings'] = report_bookings
            self.logger.debug(
                f"Added {len(report_bookings)} bookings to report '{report['name']}'.")

            # Determine report structure depth
            report_structure_levels = 0
            for row in report['rows']:
                new_depth = self._check_depth(row, 1)
                report_structure_levels = max(
                    report_structure_levels, new_depth)
            report['structureDepth'] = report_structure_levels
            # self.logger.debug(
            #     f"Report '{report['name']}' structure depth set to {report_structure_levels}.")

            # Identify saldo rows (top-level groups) and count them
            saldo_rows = [row for row in report['rows'] if row['type'].lower(
            ) == "group" and len(row['hierarchyLocation']) == 0]
            report['numSaldoRows'] = len(saldo_rows)
            # self.logger.debug(
            #     f"Report '{report['name']}' has {report['numSaldoRows']} saldo rows.")

            # Add summary information to the report dictionary
            result = self.add_up_bookings(report_bookings)
            for key, value in result.items():
                reports[report_id][key] = value
            self.logger.debug(
                f"Added summary information to report '{report['name']}'.")

        # Compile all sheets into a single Excel file
        reports['allSheetsFile'] = self.compile_all_xls_sheets(
            reports, instructions)
        self.logger.debug(
            "Completed compiling all sheets into a single Excel file.")

        self.logger.debug("Completed building all reports.")
        return reports

    def get_report_summary(self, report: dict, format='text'):
        self.logger.debug(
            f"Starting get_report_summary for report '{report['name']}' with format '{format}'.")

        # Prepare summary content
        summary = report['name'] + "\n"
        data = {
            "expense": self.settings["expense_term"],
            "revenue": self.settings["income_term"],
            "asset": "Activa",
            "liability": "Verbindlichkeiten",
            "stock": "Vermögen",
            "profitAndLoss": "Erfolg"
        }

        # Append financial data to summary
        for item, description in data.items():
            summary += f"{description}: {report.get(item, 'N/A')}\n"
            # self.logger.debug(f"Added '{description}: {report.get(item, 'N/A')}' to summary.")

        # Append booking count to summary
        summary += f"Anzahl Buchungen: {len(report.get('bookings', []))}\n"
        # self.logger.debug(f"Added booking count: {len(report.get('bookings', []))} to summary.")

        # Return the formatted summary
        return summary

    def send_email_sending(self,
                           sending: dict,
                           reports: dict,
                           send: bool = True):
        self.logger.debug(
            f"Starting send_email_sending for recipient '{sending['recipient']}' with {len(sending['packages'])} packages.")
        self.logger.info(
            f"Preparing email to {sending['recipient']} with {len(sending['packages'])} packages.")

        email_body = "Report:\n"
        report_number = 1
        attachments = []
        time_zone = timezone("Europe/Berlin")
        today = time_zone.localize(
            datetime.datetime.now()).strftime("%Y-%m-%d")

        for package in sending['packages']:
            required_sheets_for_this_package = []

            # Process each item in the package content
            for item in package['content']:
                if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                    content_item = self.get_report_summary(
                        reports[item['scope']], format='html')
                    email_body += content_item
                    self.logger.debug(
                        f"Added report summary for '{item['scope']}' to email body.")

                elif package['packageType'] == 'excel':
                    if item['type'] == 'bookings' and item['scope'] == 'all':
                        required_sheets_for_this_package.append('Buchungen')
                    elif item['type'] == 'accounts' and item['scope'] == 'all':
                        required_sheets_for_this_package.append('Konten')
                    elif item['type'] == 'costlocations' and item['scope'] == 'all':
                        required_sheets_for_this_package.append(
                            'Kostenstellen')
                    elif item['type'] == 'report':
                        required_sheets_for_this_package.append(item['scope'])
                        required_sheets_for_this_package.append(
                            self.get_bookings_sheet_name(item['scope']))
                        if item['scope'] in reports and "listings" in reports[item['scope']]:
                            for listing in reports[item['scope']]['listings']:
                                required_sheets_for_this_package.append(
                                    self.get_listings_sheet_name(item['scope'], listing))
                    self.logger.debug(
                        f"Required sheets for package '{package['packageType']}': {required_sheets_for_this_package}")

            # Attach the relevant sheets
            if required_sheets_for_this_package:
                wb = load_workbook(BytesIO(reports['allSheetsFile']))
                for sheet in wb.sheetnames:
                    if sheet not in required_sheets_for_this_package:
                        wb.remove(wb[sheet])
                self.logger.debug(
                    f"Filtered workbook to include only required sheets for package {report_number}.")

                # Prepare attachment
                file_name = f"Financial Report {report_number} {today}.xlsx" if report_number > 1 else f"Financial Report {today}.xlsx"
                attachment_content = base64.b64encode(
                    save_virtual_workbook(wb)).decode()
                attachments.append(Attachment(
                    FileContent(attachment_content),
                    FileName(file_name),
                    FileType('application/xlsx'),
                    Disposition('attachment')
                ))
                report_number += 1
                self.logger.debug(f"Added attachment '{file_name}' to email.")

        # Set up the email message
        message = Mail(
            from_email='hrm@humanrightsmonitor.org',
            to_emails=sending['recipient'],
            subject=f'Updated DevInt Administration Report {today}',
            html_content=email_body
        )
        message.attachment = attachments

        # Send the email if `send` is True
        try:
            if send:
                response = self.conn_clients['sendgrid'].send(message)
                self.logger.info(
                    f"Mail sent with SendGrid response: {response.status_code}")
        except Exception as e:
            self.logger.error(f"Error when trying to send email: {e}")

    def send_teams_sending(self, sending: dict, reports: dict, send: bool = True):
        self.logger.debug(
            f"Starting send_teams_sending with {len(sending['packages'])} packages for recipient.")

        # Initialize Teams message
        message = pymsteams.connectorcard(self.conn_clients["teams_webhook"])
        message.title("DevInt Accounting Update")
        text = ""

        # Process each package and build the message text
        for package in sending['packages']:
            for item in package['content']:
                if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                    content_item = self.get_report_summary(
                        reports[item['scope']], format="html")
                    text += content_item
                    self.logger.debug(
                        f"Added report summary for '{item['scope']}' to Teams message text.")

            # Add process log in HTML format if available
            # text += self.logger.getloghtml()
            text += "<br>loging needs to be implemented for teams message<br>"

        # Set the message text and add buttons
        message.text(text)
        message.addLinkButton("Request new Report",
                              self.conn_clients["reportrequest_backlink"])
        message.addLinkButton("Visit Buchhaltungsbutler",
                              "https://app.buchhaltungsbutler.de")
        self.logger.debug("Added links to Teams message.")

        # Send the message if `send` is True
        if send:
            try:
                message.send()
                self.logger.info("Sent Teams message successfully.")
            except Exception as e:
                self.logger.error(
                    f"Error when trying to send Teams message: {e}")

        self.logger.debug("Completed Teams message sending.")

    def send_azure_blob_sending(self, sending: dict, reports: dict, logger=None, client: dict = None, send: bool = True):
        self.logger.debug(
            f"Starting send_azure_blob_sending with {len(sending['packages'])} packages for recipient '{sending['recipient']}'.")

        target = sending['recipient'].split('/')
        report_number = 1

        self.logger.debug(f"Sending to Azure Blob container '{target}'.")
        if target[0] == 'devintaccounting':
            time_zone = timezone("Europe/Berlin")
            today = time_zone.localize(
                datetime.datetime.now()).strftime("%Y-%m-%d")

            for package in sending['packages']:
                required_sheets_for_this_package = []

                # Process each item in the package content
                for item in package['content']:
                    if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                        content_item = self.get_report_summary(
                            reports[item['scope']], format='text')
                        output = BytesIOWrapper(StringIO(content_item))
                        file_name = f"Info {report_number} {today}.txt"
                        if send:
                            blob_client = self.conn_clients["blob_service"].get_blob_client(
                                container=target[1], blob=file_name)
                            blob_client.upload_blob(output, overwrite=True)
                            self.logger.info(f"Uploaded blob: {file_name}")
                        report_number += 1

                    elif package['packageType'] == 'info' and item['type'] == 'processlog':
                        self.logger.warning(
                            "Sending process log to Azure Blob is not implemented yet.")
                        # content_item = self.logger.getloglines()
                        # output = BytesIOWrapper(StringIO(content_item))
                        # file_name = f"Info {report_number} {today}.txt"
                        # if send:
                        #     blob_client = self.conn_clients["blob_service"].get_blob_client(container=target[1], blob=file_name)
                        #     blob_client.upload_blob(output, overwrite=True)
                        #     self.logger.info(f"Uploaded blob: {file_name}")
                        # report_number += 1

                    elif package['packageType'] == 'excel':
                        if item['type'] == 'bookings' and item['scope'] == 'all':
                            required_sheets_for_this_package.append(
                                'Buchungen')
                        elif item['type'] == 'accounts' and item['scope'] == 'all':
                            required_sheets_for_this_package.append('Konten')
                        elif item['type'] == 'costlocations' and item['scope'] == 'all':
                            required_sheets_for_this_package.append(
                                'Kostenstellen')
                        elif item['type'] == 'report':
                            required_sheets_for_this_package.append(
                                item['scope'])
                            required_sheets_for_this_package.append(
                                self.get_bookings_sheet_name(item['scope']))
                            if item['scope'] in reports and "listings" in reports[item['scope']]:
                                for listing in reports[item['scope']]['listings']:
                                    required_sheets_for_this_package.append(
                                        self.get_listings_sheet_name(item['scope'], listing))
                        self.logger.debug(
                            f"Required sheets for package '{package['packageType']}': {required_sheets_for_this_package}")

                # Attach and upload Excel file if required sheets are present
                if required_sheets_for_this_package:
                    wb = load_workbook(BytesIO(reports['allSheetsFile']))
                    for sheet in wb.sheetnames:
                        if sheet not in required_sheets_for_this_package:
                            wb.remove(wb[sheet])
                    file_name = f"Financial Report {report_number} {today}.xlsx" if report_number > 1 else f"Financial Report {today}.xlsx"

                    try:
                        if send:
                            blob_client = self.conn_clients["blob_service"].get_blob_client(
                                container=target[1], blob=file_name)
                            blob_client.upload_blob(
                                BytesIO(save_virtual_workbook(wb)), overwrite=True)
                            self.logger.info(
                                f"Uploaded file: {file_name} to Azure Blob.")
                    except Exception as e:
                        self.logger.error(f"Error uploading blob: {e}")

                    report_number += 1

        self.logger.debug("Completed Azure Blob sending.")

    def send_cosmosdb_sending(self, sending: dict, reports: dict, send: bool = True):
        package_count = len(sending['packages'])
        self.logger.warning(
            f"Preparing Cosmos DB posting with {package_count} packages. Functionality not yet implemented.")

        # Placeholder for future implementation
        if send:
            self.logger.info(
                "Cosmos DB sending function is currently a placeholder and will not execute any action.")

        return

    def send_reports(self,
                     reports: dict,
                     sendings: list,
                     send: bool = True,
                     trigger_selector: list = None):
        self.logger.debug(
            f"Starting send_reports function with trigger selector: {trigger_selector}.")

        self.logger.debug("Initialized client for sending reports.")

        # Log the beginning of the sending process
        self.logger.info("Beginning report sending process.")

        for sending in sendings:
            # Check if sending trigger matches the selector
            if sending['trigger'] in trigger_selector:
                self.logger.info(
                    f"Trigger found in selector, sending report to {sending['recipient']} via {sending['channelType']}, triggered by {sending['trigger']['type']}.")

                # Determine sending channel type and execute corresponding method
                if sending['channelType'] == 'email':
                    self.send_email_sending(sending, reports, send=send)
                elif sending['channelType'] == 'teams':
                    self.send_teams_sending(sending, reports, send=send)
                elif sending['channelType'] == 'azureblob':
                    self.send_azure_blob_sending(sending, reports, send=send)
                elif sending['channelType'] == 'cosmosdb':
                    self.send_cosmosdb_sending(sending, reports, send=send)
                else:
                    # Log error for unknown channel type
                    self.logger.error(
                        f"Unknown channel type: {sending['channelType']}")

        # Log completion of the sending process
        self.logger.info("Completed sending reports.")
        self.logger.debug("Finished send_reports function.")
