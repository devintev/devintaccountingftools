import logging
import os
import requests
import json
import base64
import re
from io import StringIO, BytesIO
from datetime import datetime
from pytz import timezone

import pandas as pd

import azure.functions as func
from azure.keyvault.secrets import SecretClient
from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
from azure.cosmos import CosmosClient
import azure.cosmos.exceptions as cdbexceptions

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName, FileType, Disposition, ContentId)
from openpyxl import Workbook, workbook, styles, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
import pymsteams

from tempfile import NamedTemporaryFile
from libraries import DLogger, ExcelStyles


def save_virtual_workbook(workbook):
    """Save an openpyxl workbook in memory."""
    with NamedTemporaryFile() as f:
        workbook.save(f.name)
        f.seek(0)
        return f.read()


class DateTimeEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, datetime):
            return o.isoformat()


class BytesIOWrapper:
    def __init__(self, string_buffer, encoding='utf-8'):
        self.string_buffer = string_buffer
        self.encoding = encoding

    def __getattr__(self, attr):
        return getattr(self.string_buffer, attr)

    def read(self, size=-1):
        content = self.string_buffer.read(size)
        return content.encode(self.encoding)

    def write(self, b):
        content = b.decode(self.encoding)
        return self.string_buffer.write(content)


def read_html_page_template(filename):
    template_html_string = ""
    styles_string = ""
    with open(filename, "r") as template_file:
        template_html_string = template_file.read()
    with open("assets/styles.css", "r") as styles_file:
        styles_string = styles_file.read()
    replace_data = {"styles": f"<style>{styles_string}</style>"}
    for key, value in replace_data.items():
        template_html_string = template_html_string.replace(
            "{{"+key+"}}", value)
    return template_html_string


class DevIntConnector:
    standard_method = 'keyvault'
    settings = {
        'dateFormat': "%d.%m.%Y",
        'titleFont': styles.Font(bold=True, size=18),
        'rowsForwardForAllRows': 1,
        'slotTitleRowHeight': 25,
        'slotListColumnWidth': 12,
        'slotGroupColumnWidth': 2,
        'slotTitleFont': styles.Font(bold=True, size=10),
        'slotTitleAlignment': styles.Alignment(horizontal='center', vertical='top', wrap_text=True),
        'slotDateFont': styles.Font(bold=False, size=6, color='333333'),
        'slotDateRowHeight': 10,

        'groupTitleRowWidth': 2,
        'itemTitleRowWidth': 40,
        'costLocationColumnWidth': 20,
        'bookingsHeaderWidth': {
            'default': False,
            'date': 9,
            'booking_number': 6,
            'cost_location': 6,
            'debit_postingaccount_number': 5,
            'credit_postingaccount_number': 5,
            'postingtext': 40,
            'amount': 15,
            'currency': 5,
            'vat': 5,
            'credit_type': 5,
        },

        'depth1Font': styles.Font(name='Cambria', bold=True, size=12, ),
        'depth2Font': styles.Font(name='Cambria', bold=True, size=14, ),
        'depth3Font': styles.Font(name='Cambria', bold=True, size=16, ),
        'depth4Font': styles.Font(name='Cambria', bold=True, size=18, ),
        'depth5Font': styles.Font(name='Cambria', bold=True, size=20, ),
        'level0Fill': styles.PatternFill(fill_type='solid', start_color='85bfe6', end_color='85bfe6'),
        'level1Fill': styles.PatternFill(fill_type='solid', start_color='96c9eb', end_color='96c9eb'),
        'level2Fill': styles.PatternFill(fill_type='solid', start_color='a7d4f2', end_color='a7d4f2'),
        'level3Fill': styles.PatternFill(fill_type='solid', start_color='b5dbf5', end_color='b5dbf5'),
        'listingsHeaderFill': styles.PatternFill(fill_type='solid', start_color='bee0ec', end_color='bee0ec'),

        'euroFormat': '#0 €',
        'euroFormatPrecise': '#,##0.00 €',

        'bookingsCostlocationColumnHeader': 'cost_location',
        'bookingsDateColumnHeader': 'date',
        'bookingsAmountColumnHeader': 'amount',
        'bookingsDebitBookingsType2': 'debit_booking_type_2',
        'bookingsCreditBookingsType2': 'credit_booking_type_2',
    }

    def __init__(self, method=None):
        self.method = method or self.standard_method
        self.client = None
        self.logger = None

    def get_best_choice_logger(self, logger=None):
        self_logger = None
        if hasattr(self, 'logger'):
            if self.logger:
                self_logger = self.logger
        best_logger = logger or self_logger or DLogger.DLogger(print=False)
        return best_logger

    def setup(self, method=None):
        method = method or self.method or self.standard_method
        # key_vault_name = None
        self.logger = DLogger.DLogger(print=False)
        key_vault_name = os.getenv("KEY_VAULT_NAME")
        if key_vault_name:
            self.client = self.build_accounting_clients_with_key_vault_and_az_credentials(
                key_vault_name=key_vault_name, method=method)

    def build_accounting_clients_with_key_vault_and_az_credentials(self, key_vault_name=None, method=None):
        credential = DefaultAzureCredential()

        # print(f"key_vault_name: {key_vault_name}")
        # Azure Blob Storage Information and Client
        blob_storage_account_name = 'devintaccounting'
        abs_acct_url = f'https://{blob_storage_account_name}.blob.core.windows.net/'
        blob_service_client = BlobServiceClient(
            abs_acct_url, credential=credential)
        templates_container_name = "templates"
        financialplanning_container_name = "financialplanning"
        financial_reports_container_name = "financialreports"
        working_time_reports_container_name = "workingtimereports"
        wcc_invoicing_container_name = "wccinvoicing"
        od_invoicing_container_name = "odinvoicing"

        # Azure Key Vault Information and Client
        key_vault_Uri = f"https://{key_vault_name}.vault.azure.net"
        key_vault_client = SecretClient(
            vault_url=key_vault_Uri, credential=credential)

        # Buchhaltungsbuttler API Information
        bb_api_key = key_vault_client.get_secret("bb-api-key").value
        # bb_authorization = key_vault_client.get_secret("bb-authorization").value
        # bb_cookie = key_vault_client.get_secret("bb-cookie").value

        # # Sendgrid API Information
        # sendgrid_api_key = key_vault_client.get_secret("sendgrid-api-key").value
        # sendgrid_client = SendGridAPIClient(sendgrid_api_key)

        # # Azure Cosmos DB Information and Client
        # db_access_key = key_vault_client.get_secret("devintaccounting-cosmos-db-key").value
        # cosmos_client = CosmosClient(url="https://devintaccountingdb.documents.azure.com:443/", credential=db_access_key)
        # db_id = "WorkingHoursAccounting"
        # try:
        #     db = cosmos_client.create_database(id=db_id)
        # except cdbexceptions.CosmosResourceExistsError:
        #     db = cosmos_client.get_database_client(db_id)

        client = {
            # "blob_service": blob_service_client,
            # "templates_folder": blob_service_client.get_container_client(container=templates_container_name),
            # "financialplanning_folder": blob_service_client.get_container_client(container=financialplanning_container_name),
            # "financial_reports_folder": blob_service_client.get_container_client(container=financial_reports_container_name),
            # "working_time_reports_folder": blob_service_client.get_container_client(container=working_time_reports_container_name),
            # "wcc_invoicing_folder": blob_service_client.get_container_client(container=wcc_invoicing_container_name),
            # "od_invoicing_folder": blob_service_client.get_container_client(container=od_invoicing_container_name),
            "key_vault": key_vault_client,
            # "cosmos": cosmos_client,
            # "sendgrid": sendgrid_client,
            # "teams_webhook": key_vault_client.get_secret("teams-webhook-keegan").value,
            # "reportrequest-backlink": key_vault_client.get_secret("reportrequest-backlink").value,
            # "bb_api_key": bb_api_key,
            # "bb_authorization": bb_authorization,
            # "bb_cookie": bb_cookie,
            # "db": db
            "dummy": "dummy"
        }
        return client

    def analyse_received_http_request(self, http_request, logger=None):
        logger = self.get_best_choice_logger(logger)

        def replace_post_chars(string):
            # see https://www.w3schools.com/tags/ref_urlencode.ASP
            rs = {"+": " ", "%E2%80%9C": "\"", "%2C": ",", "%3A": ":", "%2F": "/", "%3F": "?", "%3D": "=", "%26": "&", "%23": "#",
                  "%25": "%", "%22": "\"", "%27": "'", }
            for key in rs:
                string = string.replace(key, rs[key])
            return string

        data = {"method": http_request.method,
                "url": http_request.url,
                "header": type(http_request.headers),
                "data": {}
                }
        post_data = {}
        if http_request.method == "POST":
            post_data_raw = http_request.get_body().decode("utf-8").split('&')
            case_tag_options = []
            if post_data_raw and post_data_raw[0]:
                for post_data_item in post_data_raw:
                    post_data_item_parts = post_data_item.split('=')
                    post_data[post_data_item_parts[0]] = replace_post_chars(
                        post_data_item_parts[1])
        elif http_request.method == "GET":
            post_data = dict(http_request.params)
        logger.log(f"received data with method: {data['method']}<br>")
        return post_data

    def download_all_sheets(self, blob_client):
        stream = BytesIO()
        blob_downloader = blob_client.download_blob()
        blob_downloader.download_to_stream(stream)
        get_blob_data = pd.ExcelFile(stream, engine='openpyxl')
        dataframes = {}
        for sheet in get_blob_data.sheet_names:
            new_df = pd.read_excel(
                get_blob_data, engine='openpyxl', sheet_name=sheet)
            dataframes[sheet] = new_df
        return dataframes

    def read_time_slots(self, df: pd.DataFrame):
        time_slots_df = df.dropna(how='all')
        rename_list = {'Slot Start': 'start', 'Slot End': 'end',
                       'Slot Title': 'name', 'Slot ID': 'id'}
        time_slots_df = time_slots_df.rename(
            columns=rename_list, inplace=False)
        dict_list = {}
        for time_slot_ID in time_slots_df.index:
            mdict = time_slots_df.loc[time_slot_ID].to_dict()
            mdict['startString'] = mdict['start'].strftime('%d.%m.%Y')
            mdict['endString'] = mdict['end'].strftime('%d.%m.%Y')
            mdict['start'] = mdict['start']
            mdict['end'] = mdict['end']
            dict_list[str(mdict['id'])] = mdict
        return dict_list

    def read_kontenrahmen(self, dfs: dict):
        kontenrahmen = {}
        for key, df in dfs.items():
            entries = []
            for index in df.index:
                mdict = df.loc[index].to_dict()
                entry = {
                    'accountRangeStart': mdict['Start'],
                    'accountRangeEnd': mdict['Ende'],
                    'type1': mdict['Typ 1'],
                    'type2': mdict['Typ 2'],
                }
                if not pd.isna(mdict['Kategorie 1']):
                    entry['category1'] = mdict['Kategorie 1']
                if not pd.isna(mdict['Kategorie 2']):
                    entry['category2'] = mdict['Kategorie 2']
                if not pd.isna(mdict['Kategorie 3']):
                    entry['category3'] = mdict['Kategorie 3']
                entries.append(entry)
            kontenrahmen[key[13:]] = entries
        kontenrahmen['default'] = kontenrahmen[list(kontenrahmen.keys())[0]]
        return kontenrahmen

    def read_kostenstellenplan(self, df: pd.DataFrame):
        costlocations = {}
        ksp_df = df.dropna(how='all')
        ksp_df = ksp_df.rename(columns={'Unnamed: 5': 'sectionname', 'Unnamed: 6': 'groupName',
                                        'Unnamed: 7': 'subGroupName', 'Unnamed: 8': 'name', 'Unnamed: 9': 'type', }, inplace=False)
        ksp_df = ksp_df.dropna(subset=['Unnamed: 1'])
        max_costlocation_digits = 4
        for ksp_df_index in ksp_df.index:
            item = {"limits": {}}
            mdict = ksp_df.loc[ksp_df_index].to_dict()
            if (not pd.isna(mdict['type'])):
                if (mdict['type'] == 'income') or (mdict['type'] == 'Einnahmen'):
                    item['type'] = 'income'
                elif (mdict['type'] == 'expense') or (mdict['type'] == 'Ausgaben'):
                    item['type'] = 'expense'
                elif (mdict['type'] == 'budget'):
                    item['type'] = 'budget'
            if (not pd.isna(mdict['Unnamed: 10'])):
                item['limits']['max'] = float(mdict['Unnamed: 10'])
            if (not pd.isna(mdict['Unnamed: 11'])):
                item['limits']['extendedMax'] = float(mdict['Unnamed: 11'])
            if (not pd.isna(mdict['groupName'])) or (not pd.isna(mdict['subGroupName'])):
                item['type'] = 'group'
                if not pd.isna(mdict['subGroupName']):
                    item['name'] = mdict['subGroupName']
                elif not pd.isna(mdict['groupName']):
                    item['name'] = mdict['groupName']
                starting_digits = ""
                for col in ['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3']:
                    if not pd.isna(mdict[col]):
                        starting_digits += str(int(mdict[col]))
                min_digits, max_digits = starting_digits, starting_digits
                for digit in range(max_costlocation_digits - len(starting_digits)):
                    min_digits += "0"
                    max_digits += "9"
                item['id'] = starting_digits
                item['costLocations'] = {'startingDigits': starting_digits, 'min': int(
                    min_digits), 'max': int(max_digits)}
                item['children'] = {}
                costlocations[item['id']] = item
            elif not pd.isna(mdict['name']):
                item['type'] = 'item'
                item['name'] = mdict['name']
                digits = ""
                for col in ['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3']:
                    if not pd.isna(mdict[col]):
                        digits += str(int(mdict[col]))
                item['number'] = int(digits)
                item['id'] = digits
                costlocations[item['id']] = item
        for id, item in costlocations.items():
            if item['type'] == 'group':
                for child_id, child_item in costlocations.items():
                    if child_item['type'] == 'item':
                        if child_item['number'] >= item['costLocations']['min'] and child_item['number'] <= item['costLocations']['max']:
                            item['children'][child_id] = child_item
                    if child_item['type'] == 'group' and child_item['id'] != item['id']:
                        if child_item['costLocations']['min'] >= item['costLocations']['min'] and child_item['costLocations']['max'] <= item['costLocations']['max']:
                            item['children'][child_id] = child_item
        return costlocations

    def read_reports_overview(self, df: pd.DataFrame, logger=None):
        logger = self.get_best_choice_logger(logger)
        reports = {}
        ignored_reports = 0
        for report_ID in df.index:
            mdict = df.loc[report_ID].to_dict()
            report = {}
            report['id'] = mdict["Plan"]
            report['name'] = mdict['Name']
            report['costLocationsRange'] = {
                'min': mdict['Kostenstellenstart'], "max": mdict['Kostenstellenende'], }
            if mdict['Version'] == 2:
                reports[mdict["Plan"]] = report
            else:
                ignored_reports += 1
        if logger and ignored_reports > 0:
            logger.log(
                f"{ignored_reports} report plans were not read because of their deprecated version")
        return reports

    def read_report_schema_into(self, report: dict, schema_df: pd.DataFrame, time_slot_templates: dict, logger=None):
        logger = self.get_best_choice_logger(logger)
        schema_df.dropna(how='all', inplace=True)
        schema_df.dropna(axis=1, how="all", inplace=True)
        slot_columns = [
            col for col in schema_df.columns if col.startswith('s:')]
        slot_names = [col.split(':')[1] for col in slot_columns]
        section_column_names = [
            col for col in schema_df.columns if col.startswith('Section ')]
        slots_data = []
        listings_data = []
        order_number = 0
        for slot_column in slot_columns:
            slot_data = {}
            slot_data['timeSlotId'] = slot_column.split(':')[1]
            slot_data['buildListing'] = len(slot_column.split(
                ':')) > 2 and slot_column.split(':')[2] == "listing"
            if re.findall(r'\.\d$', slot_data['timeSlotId']) and slot_data['timeSlotId'][:-2] in slot_names:
                slot_data['timeSlotId'] = slot_data['timeSlotId'][:-2]
            slot_data['orderNumber'] = order_number
            order_number += 1
            slot_data['name'] = time_slot_templates[slot_data['timeSlotId']]['name']
            slot_data['start'] = time_slot_templates[slot_data['timeSlotId']]['start']
            slot_data['end'] = time_slot_templates[slot_data['timeSlotId']]['end']
            slots_data.append(slot_data)
            if slot_data['buildListing']:
                listings_data.append(slot_data)
        report['slots'] = slots_data
        report['numberSectionLevels'] = len(section_column_names)
        items = []
        order_number = 0
        for row_id in schema_df.index:
            mdict = schema_df.loc[row_id].to_dict()
            item = {
                'orderNumber': order_number,
            }
            order_number += 1
            hierarchy_location = []
            for section_number in range(len(section_column_names)):
                column_name = "Section " + str(section_number + 1)
                if not pd.isna(mdict[column_name]):
                    hierarchy_location.append(mdict[column_name])
            item['hierarchyLocation'] = hierarchy_location
            if not pd.isna(mdict['Name']):
                item['name'] = mdict['Name']
            if mdict['Type'].lower() in ['income', 'expense', 'group']:
                item['type'] = mdict['Type'].lower()
            if item['type'] in ['expense', 'income']:
                if isinstance(mdict['Cost Location'], str):
                    item['costLocationsStrings'] = mdict['Cost Location'].split(
                        ',')
                else:
                    cl = int(mdict['Cost Location'])
                    item['costLocationsStrings'] = [f"{cl:04d}"]
                item['costLocations'] = [
                    int(cl) for cl in item['costLocationsStrings']]
                item['hierarchyLocation'].append(item['name'])
            elif item['type'] == 'group' and hierarchy_location:
                item['name'] = hierarchy_location[-1]
            else:
                item['name'] = ''
            slots_info = []
            for report_slot in report['slots']:
                row_slot_info = {}
                row_slot_info['timeSlotId'] = report_slot['timeSlotId']
                row_slot_info['buildListing'] = report_slot['buildListing']
                row_slot_info['slotDetails'] = report_slot
                row_slot_info['orderNumber'] = report_slot['orderNumber']
                column_name = 's:' + report_slot['timeSlotId']
                if row_slot_info['buildListing']:
                    column_name += ":listing"
                value = mdict[column_name]
                if not pd.isna(value):
                    if isinstance(value, str):
                        # print(report['id'],"(string):",value,end=" ")
                        if value.lower() in ['fieldsum', 'cellsum', 'sum', 'summe']:
                            # print(" (f,c,s)" , end=" ")
                            row_slot_info['type'] = 'sum'
                        elif value.lower() in ['bookings', 'buchungen']:
                            # print(" (bookings)", end=" ")
                            row_slot_info['type'] = 'bookings'
                        elif value:
                            # print(" (other)", end=" ")
                            row_slot_info['type'] = 'budget'
                            parts = value.split(':')
                            row_slot_info['limit'] = float(parts[0])
                            if len(parts) > 1:
                                if parts[1] == "!":
                                    row_slot_info['limitType'] = 'absolute'
                                elif parts[1].startswith("+"):
                                    row_slot_info['limitType'] = 'exceedable'
                                    if parts[1][-1] == "%":
                                        row_slot_info['extendedLimit'] = row_slot_info['limit']*(
                                            1.0+float(parts[1][1:-1])/100.0)
                                    else:
                                        row_slot_info['extendedLimit'] = row_slot_info['limit']+float(
                                            parts[1][1:])
                            else:
                                row_slot_info['limitType'] = 'relative'
                        else:
                            # print(report['id'],"("+str(value)+"):",value,end=" ")
                            row_slot_info['type'] = 'empty'
                            row_slot_info['limit'] = 0.0
                            row_slot_info['limitType'] = 'relative'
                    elif isinstance(value, int) or isinstance(value, float):
                        # print(report['id'],"("+str(type(value))+"):",value,end=" ")
                        row_slot_info['type'] = 'budget'
                        row_slot_info['limit'] = float(value)
                        row_slot_info['limitType'] = 'relative'
                    else:
                        # print(report['id'],"("+str(type(value))+"):",value,end=" ")
                        pass
                else:
                    # print(report['id'],"(na!!):",value,end=" ")
                    row_slot_info['type'] = 'empty'
                    row_slot_info['limit'] = 0.0
                    row_slot_info['limitType'] = 'relative'

                slots_info.append(row_slot_info)
            item['slots'] = slots_info
            items.append(item)
            cost_locations = set()
            cost_locations_strings = set()
            for item in items:
                if item['type'] in ['expense', 'income']:
                    cost_locations.update(item['costLocations'])
                    cost_locations_strings.update(item['costLocationsStrings'])
            report['costLocations'] = list(cost_locations)
            report['costLocationsStrings'] = list(cost_locations_strings)
        for row in items:
            if row['type'] == 'group':
                children = []
                hierarchy_depth = len(row['hierarchyLocation'])
                for item in items:
                    if hierarchy_depth > 0:
                        if len(item['hierarchyLocation']) == hierarchy_depth+1 and item['hierarchyLocation'][:-1] == row['hierarchyLocation']:
                            children.append(item)
                    else:
                        if len(item['hierarchyLocation']) == 1:
                            children.append(item)
                if children:
                    row['children'] = children
        report['rows'] = items
        if len(listings_data) > 0:
            report['listings'] = listings_data

    def read_distribution_instructions(self, df: pd.DataFrame, logger=None):
        logger = self.get_best_choice_logger(logger)
        dist = []
        for row_id in df.index:
            mdict = df.loc[row_id].to_dict()
            entry = {'channelType': mdict['type'], 'packages': []}
            if not pd.isna(mdict['recipient']):
                entry['recipient'] = mdict['recipient']
            if not pd.isna(mdict['trigger']):
                entry['trigger'] = {'type': mdict['trigger'].split(':')[
                    0].strip()}
                if len(mdict['trigger'].split(':')) > 1:
                    entry['trigger']['condition'] = mdict['trigger'].split(':')[
                        1].strip()
                content = mdict['content'].split('),')
                for item in content:
                    # split off string before first occurance of '('
                    subitems = item.split('(')
                    job_type = subitems[0].strip()
                    position = subitems[1].rfind(")")
                    if position < 0:
                        job_list = subitems[1]
                    else:
                        job_list = subitems[1][:position]
                    content_items = []
                    for job in job_list.split(','):
                        citem = {'type': job.split(':')[0].strip(
                        ), 'scope': job.split(':')[1].strip()}
                        content_items.append(citem)
                    entry['packages'].append(
                        {'packageType': job_type.lower(), 'content': content_items})
            dist.append(entry)
        return dist

    def read_instruction_files(self, container=None, logger=None):
        logger = self.get_best_choice_logger(logger)
        container = container or self.client["templates_folder"]
        instruction_files_list = container.list_blobs()
        instructions = {"sheets": {},
                        "unprocessedSheets": {}, "processedSheets": {}}
        for instruction_file in instruction_files_list:
            blob_client = container.get_blob_client(instruction_file.name)
            if instruction_file.name.endswith(".xlsx"):
                dataframes = self.download_all_sheets(blob_client)
                for key, df in dataframes.items():
                    instructions["sheets"][key] = df
                    instructions["unprocessedSheets"][key] = df
        if "Time Slots" in instructions["unprocessedSheets"]:
            instructions["timeSlotTemplates"] = self.read_time_slots(
                instructions["unprocessedSheets"]["Time Slots"])
            instructions["processedSheets"]["Time Slots"] = instructions["unprocessedSheets"]["Time Slots"]
            del instructions["unprocessedSheets"]["Time Slots"]
        else:
            logger.log(
                "no 'Time Slots' sheets found in any excel file in Blob Storage.", type="error")
        kontenrahmen = {k: v for k, v in instructions["unprocessedSheets"].items(
        ) if k.startswith('Kontenrahmen ')}
        if len(kontenrahmen) > 0:
            logger.log(f"found {len(kontenrahmen)} kontenrahmen")
            instructions["kontenrahmen"] = self.read_kontenrahmen(kontenrahmen)
            instructions["processedSheets"].update(kontenrahmen)
            for key in kontenrahmen:
                del instructions["unprocessedSheets"][key]
        else:
            logger.log(
                "no 'Kontenrahmen' sheet found in any excel file in Blob Storage.", type="error")
        if "Kostenstellenplan" in instructions["unprocessedSheets"]:
            # logger.log("found Kostenstellenplan")
            instructions["kostenstellenplan"] = self.read_kostenstellenplan(
                instructions["unprocessedSheets"]["Kostenstellenplan"])
            instructions["processedSheets"]["Kostenstellenplan"] = instructions["unprocessedSheets"]["Kostenstellenplan"]
            del instructions["unprocessedSheets"]["Kostenstellenplan"]
        else:
            logger.log(
                "no 'Kostenstellenplan' sheet found in any excel file in Blob Storage.", type="error")
        if "Reports Overview" in instructions["unprocessedSheets"]:
            # logger.log("found Reports Overview")
            instructions["reports"] = self.read_reports_overview(
                instructions["unprocessedSheets"]["Reports Overview"])
            instructions["processedSheets"]["Reports Overview"] = instructions["unprocessedSheets"]["Reports Overview"]
            del instructions["unprocessedSheets"]["Reports Overview"]
        else:
            logger.log(
                "no 'Reports Overview' sheet found in any excel file in Blob Storage.", type="error")
        for key, report in instructions['reports'].items():
            sheet_name = "Budget Plan "+key
            if sheet_name in instructions["unprocessedSheets"]:
                self.read_report_schema_into(
                    report, instructions["unprocessedSheets"][sheet_name], instructions["timeSlotTemplates"], logger)
                instructions["processedSheets"][sheet_name] = instructions["unprocessedSheets"][sheet_name]
                del instructions["unprocessedSheets"][sheet_name]
        if "Flow Plan" in instructions["unprocessedSheets"]:
            # logger.log("found distribution instructions")
            instructions["distribution"] = self.read_distribution_instructions(
                instructions["unprocessedSheets"]["Flow Plan"])
            instructions["processedSheets"]["Flow Plan"] = instructions["unprocessedSheets"]["Flow Plan"]
            del instructions["unprocessedSheets"]["Flow Plan"]
        else:
            logger.log(
                "no 'Flow Plan' sheet found in any excel file in Blob Storage.", type="error")
        logger.log("completed reading instruction files")
        return instructions

    def get_all_bb_posts(self, logger=None, start_date="2021-01-01", end_date="2029-12-31", client=None):
        logger = self.get_best_choice_logger(logger)
        logging.info('Client inside: {}'.format(self.client))
        client = client or self.client
        logger.log("downloading all bb posts")
        # https://app.buchhaltungsbutler.de/docs/api/v1/
        base_url = "https://webapp.buchhaltungsbutler.de/api/v1"
        request = '/postings/get'
        url = base_url + request
        payload = json.dumps({
            "api_key": client['bb_api_key'],
            "date_from": start_date,
            "date_to": end_date,
        })
        headers = {
            'Content-Type': 'application/json',
            'Authorization': client['bb_authorization'],
            'Cookie': 'bbutler='+client['bb_cookie'],
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        result = json.loads(response.text)
        logger.log('downloaded '+str(result['rows'])+' bookings')
        return result['data']

    def get_all_bb_accounts(self, logger=None, client=None):
        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        logger.log("downloading all bb accounts")
        # https://app.buchhaltungsbutler.de/docs/api/v1/
        base_url = "https://webapp.buchhaltungsbutler.de/api/v1"
        request = '/accounts/get'
        url = base_url + request
        payload = json.dumps({
            "api_key": client['bb_api_key'],
        })
        headers = {
            'Content-Type': 'application/json',
            'Authorization': client['bb_authorization'],
            'Cookie': 'bbutler='+client['bb_cookie'],
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        result = json.loads(response.text)
        if logger:
            logger.log('downloaded '+str(result['rows'])+' accounts')
        return result['data']

    def read_expected_bookings(self, container, logger=None):
        logger = self.get_best_choice_logger(logger)
        if logger is None:
            logger = DLogger.DLogger(print=False)
        expected_bookings = {"sheets": {},
                             "unprocessedSheets": {}, "processedSheets": {}}
        files_list = container.list_blobs()
        filenames = [file.name for file in files_list]
        if "Expected Bookings.xlsx" in filenames:
            blob_client = container.get_blob_client("Expected Bookings.xlsx")
            logger.log("found Expected Bookings.xlsx")
            expected_bookings_dfs = self.download_all_sheets(blob_client)
            for key, df in expected_bookings_dfs.items():
                expected_bookings["sheets"][key] = df
                expected_bookings["unprocessedSheets"][key] = df
        return expected_bookings

    def add_more_account_information_to_bookings(self, bookings, kontenrahmen, logger=None):
        logger = self.get_best_choice_logger(logger)
        for booking in bookings:
            debit_account = int(booking['debit_postingaccount_number'])
            credit_account = int(booking['credit_postingaccount_number'])
            for account in kontenrahmen:
                if debit_account >= account['accountRangeStart'] and debit_account <= account['accountRangeEnd']:
                    booking['debit_booking_type_1'] = account['type1']
                    booking['debit_booking_type_2'] = account['type2']
                    categories = []
                    if "category1" in account:
                        categories.append(account['category1'])
                    if "category2" in account:
                        categories.append(account['category2'])
                    if "category3" in account:
                        categories.append(account['category3'])
                    booking['debit_booking_categories'] = categories
                    break
            for account in kontenrahmen:
                if credit_account >= account['accountRangeStart'] and credit_account <= account['accountRangeEnd']:
                    booking['credit_booking_type_1'] = account['type1']
                    booking['credit_booking_type_2'] = account['type2']
                    categories = []
                    if "category1" in account:
                        categories.append(account['category1'])
                    if "category2" in account:
                        categories.append(account['category2'])
                    if "category3" in account:
                        categories.append(account['category3'])
                    booking['credit_booking_categories'] = categories
                    break

    def collect_all_accounts(self, bookings, kontenrahmen, logger=None):
        logger = self.get_best_choice_logger(logger)
        all_accounts = {}
        for booking in bookings:
            if int(booking['debit_postingaccount_number']) not in all_accounts:
                all_accounts[int(booking['debit_postingaccount_number'])] = {
                    'bookings': [],
                    'type 1': booking['debit_booking_type_1'],
                    'type 2': booking['debit_booking_type_2'],
                }
            if int(booking['credit_postingaccount_number']) not in all_accounts:
                all_accounts[int(booking['credit_postingaccount_number'])] = {
                    'bookings': [],
                    'type 1': booking['credit_booking_type_1'],
                    'type 2': booking['credit_booking_type_2'],
                }
        all_accounts = {i: all_accounts[i] for i in sorted(all_accounts)}
        accountslist = list(all_accounts.keys())
        if logger:
            logger.log('Found ' + str(len(all_accounts)) + ' accounts from ' +
                       str(accountslist[0])+' to '+str(accountslist[-1]))
        for booking in bookings:
            all_accounts[int(booking['debit_postingaccount_number'])
                         ]['bookings'].append(booking)
            all_accounts[int(booking['credit_postingaccount_number'])
                         ]['bookings'].append(booking)
        for key, account in all_accounts.items():
            account_number = int(key)
            for account_information in kontenrahmen:
                if account_number >= account_information['accountRangeStart'] and account_number <= account_information['accountRangeEnd']:
                    categories = []
                    if "category1" in account_information:
                        categories.append(account_information['category1'])
                    if "category2" in account_information:
                        categories.append(account_information['category2'])
                    if "category3" in account_information:
                        categories.append(account_information['category3'])
                    account['booking_categories'] = categories
                    break
            account['bookings'].sort(key=lambda x: x['date'])
            account['saldo'] = 0.0
            account['soll'] = 0.0
            account['haben'] = 0.0
            account_no = str(key)
            for booking in account['bookings']:
                if booking['debit_postingaccount_number'] == account_no:
                    account['soll'] += float(booking['amount'])
                if booking['credit_postingaccount_number'] == account_no:
                    account['haben'] += float(booking['amount'])
            account['saldo'] = account['soll'] - account['haben']
        return all_accounts

    def add_up_bookings(self, bookings, logger=None):
        # logger = self.get_best_choice_logger(logger)
        result = {}
        einnahmen, ausgaben, activa, passiva = 0.0, 0.0, 0.0, 0.0
        for booking in bookings:
            if booking['debit_booking_type_2'] == 'Einnahmen':
                einnahmen -= float(booking['amount'])
            if booking['credit_booking_type_2'] == 'Einnahmen':
                einnahmen += float(booking['amount'])
            if booking['debit_booking_type_2'] == 'Ausgaben':
                ausgaben += float(booking['amount'])
            if booking['credit_booking_type_2'] == 'Ausgaben':
                ausgaben -= float(booking['amount'])
            if booking['debit_booking_type_2'] == 'Activa':
                activa += float(booking['amount'])
            if booking['credit_booking_type_2'] == 'Activa':
                activa -= float(booking['amount'])
            if booking['debit_booking_type_2'] == 'Passiva':
                passiva -= float(booking['amount'])
            if booking['credit_booking_type_2'] == 'Passiva':
                passiva += float(booking['amount'])
        result['revenue'] = einnahmen
        result['expense'] = ausgaben
        result['asset'] = activa
        result['liability'] = passiva
        result['stock'] = activa - passiva
        result['profitAndLoss'] = einnahmen - ausgaben
        return result

    def collect_all_costlocations(self, bookings: list, costlocations: dict, logger=None):
        logger = self.get_best_choice_logger(logger)
        if logger:
            logger.log('Collecting all costlocations with' + str(len(bookings)) +
                       ' bookings and ' + str(len(costlocations)) + ' costlocations')
        all_costlocations = {}
        for booking in bookings:
            cl = booking['cost_location'] if booking['cost_location'] else 'without'
            if cl not in all_costlocations:
                all_costlocations[cl] = {'bookings': []}
                if str(cl) in costlocations:
                    all_costlocations[booking['cost_location']
                                      ]["limits"] = costlocations[cl]['limits']
                    all_costlocations[booking['cost_location']
                                      ]["type"] = costlocations[cl]['type']
                    all_costlocations[booking['cost_location']
                                      ]["name"] = costlocations[cl]['name']
                    all_costlocations[booking['cost_location']
                                      ]["number"] = costlocations[booking['cost_location']]['number']
        all_costlocations = {i: all_costlocations[i]
                             for i in sorted(all_costlocations)}
        costlocations_list = list(all_costlocations.keys())
        for booking in bookings:
            if booking['cost_location']:
                all_costlocations[booking['cost_location']
                                  ]['bookings'].append(booking)
            else:
                all_costlocations['without']['bookings'].append(booking)
        if logger:
            logger.log('Found ' + str(len(all_costlocations)) + ' cost locations from "' +
                       str(costlocations_list[0])+'" to "'+str(costlocations_list[-2])+'"')
            logger.log(
                'Found ' + str(len(all_costlocations["without"]["bookings"])) + ' bookings without a cost location')
        for key, costlocation in all_costlocations.items():
            costlocation['bookings'].sort(key=lambda x: x['date'])
            result = self.add_up_bookings(costlocation['bookings'], logger)
            for key, value in result.items():
                costlocation[key] = value
        return all_costlocations

    def fill_bookings_to_sheet(self, bookings: list, sheet):
        def move_a_after_b(list, a, b):
            if a in list and b in list:
                list.remove(a)
                list.insert(list.index(b)+1, a)
        if len(bookings) > 0:
            headers = list(bookings[0].keys())
            crow, ccol = 1, 1
            move_a_after_b(headers, 'debit_booking_type_1',
                           'debit_postingaccount_number')
            move_a_after_b(headers, 'debit_booking_type_2',
                           'debit_booking_type_1')
            move_a_after_b(headers, 'debit_booking_categories',
                           'debit_booking_type_2')
            move_a_after_b(headers, 'credit_booking_type_1',
                           'credit_postingaccount_number')
            move_a_after_b(headers, 'credit_booking_type_2',
                           'credit_booking_type_1')
            move_a_after_b(headers, 'credit_booking_categories',
                           'credit_booking_type_2')
            move_a_after_b(headers, 'booking_number', 'id_by_customer')
            move_a_after_b(headers, 'date_delivery', 'comment')
            move_a_after_b(headers, 'date_vat_effective', 'comment')
            move_a_after_b(headers, 'tax_key', 'comment')
            move_a_after_b(headers, 'cost_location', 'date')
            # apply_style(cell, "top_th")
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)
                cell.value = header
                cell.font = styles.Font(bold=True)
                if header in self.settings['bookingsHeaderWidth']:
                    sheet.column_dimensions[self.colchar(
                        ccol)].width = self.settings['bookingsHeaderWidth'][header]
                ccol += 1
            crow += 1
            ccol = 1
            for booking in bookings:
                for header in headers:
                    cell = sheet.cell(row=crow, column=ccol)
                    if header in ['amount', 'vat', 'receipts_assigned_vat_rates', 'receipts_assigned_assigned_amounts']:
                        cell.style = 'Comma'
                        # cell.number_format = '#,##0.00 €'
                        try:
                            cell.value = float(booking[header])
                        except:
                            cell.value = str(booking[header])
                    elif header in ['date']:
                        cell.value = datetime.strptime(
                            booking[header], "%Y-%m-%d %H:%M:%S")
                        cell.number_format = 'DD.MM.YY'
                    elif header in ['date_vat_effective']:
                        cell.value = datetime.strptime(
                            booking[header], "%Y-%m-%d")
                        cell.number_format = 'DD.MM.YY'
                    elif header in ['id_by_customer', 'debit_postingaccount_number', 'credit_postingaccount_number', 'tax_key', 'booking_number', 'transactions_id_by_customer']:
                        try:
                            cell.value = int(booking[header])
                        except:
                            cell.value = str(booking[header])
                    elif header in ['debit_booking_categories', 'credit_booking_categories']:
                        cell.value = ' - '.join(booking[header])
                    else:
                        cell.value = str(booking[header])
                    ccol += 1
                crow += 1
                ccol = 1

    def fill_costlocations_to_sheet(self, costlocations: dict, sheet):
        headers = list(costlocations[list(costlocations.keys())[0]].keys())
        headers.remove('bookings')
        headers.remove('type')
        headers.insert(0, 'cost_location')
        crow, ccol = 1, 1
        for header in headers:
            cell = sheet.cell(row=crow, column=ccol)
            if header in ['type 1', 'type 2']:
                cell.value = header.replace('type', 'category')
            else:
                cell.value = header
            cell.font = styles.Font(bold=True)
            ccol += 1
        crow += 1
        ccol = 1
        for number, costlocation in costlocations.items():
            if ('type' in costlocation) and not (costlocation['type'] == 'item'):
                continue
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)
                if header in ['cost_location']:
                    if number == 'without':
                        cell.value = ''
                    else:
                        cell.value = int(number)
                elif header == 'limits':
                    if header in costlocation:
                        cell.value = ' - '.join([f'{k}: {v}' for k,
                                                v in costlocation[header].items()])
                    else:
                        cell.value = ''
                elif header in ['type', 'name', 'number']:
                    if header in costlocation:
                        cell.value = costlocation[header]
                    else:
                        cell.value = ''
                elif header in ['revenue', 'expense', 'asset', 'liability', 'stock', 'profitAndLoss']:
                    try:
                        cell.style = 'Comma'
                        cell.value = float(costlocation[header])
                    except:
                        cell.value = str(costlocation[header])
                else:
                    cell.value = str(costlocation[header])
                ccol += 1
            crow += 1
            ccol = 1

    def fill_accounts_to_sheet(self, accounts: dict, sheet):
        headers = list(accounts[list(accounts.keys())[0]].keys())
        # repace 'type' with 'account_type'
        headers.remove('bookings')
        headers.insert(0, 'account')
        if 'booking_categories' in headers and 'type 2' in headers:
            headers.remove('booking_categories')
            headers.insert(headers.index('type 2')+1, 'booking_categories')
        crow, ccol = 1, 1
        for header in headers:
            cell = sheet.cell(row=crow, column=ccol)
            if header in ['type 1', 'type 2']:
                cell.value = header.replace('type', 'category')
            else:
                cell.value = header
            cell.font = styles.Font(bold=True)
            ccol += 1
        crow += 1
        ccol = 1
        for number, account in accounts.items():
            for header in headers:
                cell = sheet.cell(row=crow, column=ccol)
                if header in ['account']:
                    cell.value = int(number)
                elif header in ['booking_categories']:
                    cell.value = ' - '.join(account[header])
                elif header in ['saldo', 'soll', 'haben']:
                    try:
                        cell.style = 'Comma'
                        cell.value = float(account[header])
                    except:
                        cell.value = str(account[header])
                else:
                    cell.value = str(account[header])
                ccol += 1
            crow += 1
            ccol = 1

    def check_depth(self, row, depth=1):
        mdepth = depth
        if 'children' in row and len(row['children']) > 0:
            cdepth = depth + 1
            for child in row['children']:
                ndepth = self.check_depth(child, cdepth)
                if ndepth > mdepth:
                    mdepth = ndepth
        return mdepth

    def add_slot_header(self, sheet, slot: dict, col: int = 1, row: int = 1, depth: int = 1):
        # slot_details = slot['slotDetails']
        sheet.cell(row=row, column=col).value = slot['name']
        sheet.row_dimensions[row].height = self.settings['slotTitleRowHeight']
        slot['startDateRow'] = row+1
        slot['endDateRow'] = slot['startDateRow'] + 1
        for i in range(depth):
            if i == (depth - 1):
                column_width = self.settings['slotListColumnWidth']
            else:
                column_width = self.settings['slotGroupColumnWidth']
            sheet.column_dimensions[self.colchar(col+i)].width = column_width

        if depth > 1:
            sheet.merge_cells(start_row=row, start_column=col,
                              end_row=row, end_column=col+depth-1)
        if depth > 2:
            sheet.merge_cells(start_row=slot['startDateRow'], start_column=col,
                              end_row=slot['startDateRow'], end_column=col+depth-2)
            sheet.merge_cells(start_row=slot['endDateRow'], start_column=col,
                              end_row=slot['endDateRow'], end_column=col+depth-2)
        sheet.cell(
            row=row, column=col).alignment = self.settings['slotTitleAlignment']
        sheet.cell(row=row, column=col).font = self.settings['slotTitleFont']
        for cell in sheet[str(slot['startDateRow'])+":"+str(slot['startDateRow'])]:
            cell.font = self.settings['slotDateFont']
        for cell in sheet[str(slot['endDateRow'])+":"+str(slot['endDateRow'])]:
            cell.font = self.settings['slotDateFont']
        sheet.row_dimensions[slot['startDateRow']
                             ].height = self.settings['slotDateRowHeight']
        sheet.row_dimensions[slot['endDateRow']
                             ].height = self.settings['slotDateRowHeight']
        if depth > 2:
            sheet.cell(row=slot['startDateRow'], column=col).value = "from:"
            sheet.cell(row=slot['endDateRow'], column=col).value = "to:"
        start_date_string = slot['start'].strftime(self.settings['dateFormat'])
        end_date_string = slot['end'].strftime(self.settings['dateFormat'])
        sheet.cell(row=slot['startDateRow'], column=col +
                   depth - 1).value = start_date_string
        sheet.cell(row=slot['endDateRow'], column=col +
                   depth - 1).value = end_date_string
        slot['startRow'] = row
        slot['endRow'] = row + 2
        slot['startColumn'] = col
        slot['endColumn'] = col + depth - 1
        slot['startDateCellReference'] = self.colchar(
            slot['endColumn']) + "$" + str(slot['startDateRow'])
        slot['endDateCellReference'] = self.colchar(
            slot['endColumn']) + "$" + str(slot['endDateRow'])
        return

    def get_column_letter_by_column_header(self, sheet, header: str):
        for cell in sheet["1:1"]:
            if cell.value == header:
                return cell.column_letter

    def build_group_summation_formula(self, report_row: dict, slot: dict, report: dict, cell_column: int):
        max_levels = report['structureDepth'] - report['numSaldoRows']
        is_saldo_row = report_row['type'].lower() == "group" and len(
            report_row['hierarchyLocation']) == 0

        def all_children(row: dict):
            val = 0
            if 'children' in row:
                for child in row['children']:
                    val += 1 + all_children(child)
            return val

        def get_summation_type(row: dict):
            if 'children' in row:
                types = set()
                for child in row['children']:
                    types.update(get_summation_type(child))
                return types
            else:
                return {row['type']}
        types = get_summation_type(report_row)
        its_an_item_group = ('children' in report_row) and (len(
            report_row['children']) > 0) and report_row['children'][0]['type'] in ['expense', 'income']
        if len(types) == 1:
            if its_an_item_group:
                # must be a group of final budget lines
                sub_rows = report_row['children']
                sub_rows = sorted(sub_rows, key=lambda x: x['rowNumber'])
                column_to_be_summed = slot['slotDetails']['startColumn'] + \
                    max_levels - 1
                if is_saldo_row:
                    column_to_be_summed -= 1
                if len(sub_rows) > 1:
                    return "=SUM(" + self.colchar(column_to_be_summed) + str(sub_rows[0]['rowNumber']) + ":" + self.colchar(column_to_be_summed) + str(sub_rows[-1]['rowNumber']) + ")"
                elif len(sub_rows) == 1:
                    return "=" + self.colchar(column_to_be_summed) + str(sub_rows[0]['rowNumber'])
                else:
                    print("Error: no sub rows for", report_row['name'])
                    return "=0"
            else:
                # must be a group of groups
                column_to_be_summed = cell_column + 1
                if is_saldo_row:
                    column_to_be_summed -= 1
                if len(report_row['children']) == 1:
                    return "=" + self.colchar(column_to_be_summed) + str(report_row['children'][0]['rowNumber'])
                else:
                    return "=SUM(" + self.colchar(column_to_be_summed) + str(report_row['children'][0]['rowNumber']) + ":" + self.colchar(column_to_be_summed) + str(report_row['children'][-1]['rowNumber']) + ")"
        elif len(types) > 1 and ('expense' in types or 'income' in types or '' in types):
            if 'children' in report_row:
                column_to_be_summed = cell_column + 1
                if is_saldo_row:
                    column_to_be_summed -= 1
                formula = "="
                for child in report_row['children']:
                    child_type = get_summation_type(child)
                    sign = "+" if 'income' in child_type else "-"
                    child_cell_string = self.colchar(
                        column_to_be_summed) + str(child['rowNumber'])
                    formula += sign + child_cell_string
                return formula
            else:
                return "Error: two sub item types but problems"
        return formula

    def colchar(self, col=0):
        base = ord('A')
        rounds = int((col - 1) / 26)
        letters = chr(base + ((col-1) % 26))
        if rounds > 0:
            letters = chr(base - 1 + rounds) + letters
        return letters

    def fill_row_to_sheet(self, report: dict, report_row: dict, sheet: Worksheet, bookings_sheet: str, first_col: int = 1, depth: int = 1):
        crow = report_row['rowNumber']
        col = first_col

        # dealing with rows that represent budget groups
        if report_row['type'] == 'group':
            level = len(report_row['hierarchyLocation'])

            # format all relevant cells
            font = self.settings['depth'+str(depth - level)+'Font'] if 0 < depth - \
                level < 6 else self.settings['level5Font']
            fill = self.settings['level'+str(level)+'Fill'] if - \
                1 < level < 4 else self.settings['level0Fill']
            cell = sheet.cell(row=crow, column=col + level)
            cell.value = report_row['name']
            cell.font = font
            cell.fill = fill
            for i in range(level, depth + 1 + depth * len(report_row['slots'])):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = font
                cell.fill = fill

            # fill all slots of the row
            slot_number = 0
            for slot in report_row['slots']:
                start_column = slot['slotDetails']['startColumn']
                end_column = slot['slotDetails']['endColumn']
                if level > 0:  # its a regular row budget item row
                    cell_column = start_column + level - 1
                    cell = sheet.cell(row=crow, column=cell_column)
                    if slot['type'] == 'budget':
                        cell.value = slot['limit']
                        cell.number_format = self.settings['euroFormat']
                    else:  # its the summary row
                        formula = self.build_group_summation_formula(
                            report_row, slot, report, cell_column)
                        cell.value = formula
                        cell.number_format = self.settings['euroFormat']
                    sheet.merge_cells(
                        start_row=crow, start_column=cell_column, end_row=crow, end_column=end_column)
                elif level == 0:  # its a slot row
                    cell_column = start_column
                    cell = sheet.cell(row=crow, column=start_column)
                    formula = self.build_group_summation_formula(
                        report_row, slot, report, cell_column)
                    cell.value = formula
                    cell.number_format = self.settings['euroFormat']
                    sheet.merge_cells(
                        start_row=crow, start_column=start_column, end_row=crow, end_column=end_column)
                slot_number += 1

        # dealing with rows that represent actual budget lines at the deepest level
        elif report_row['type'] in ['expense', 'income'] and bookings_sheet is not None:
            col = first_col
            sheet.cell(row=crow, column=col + depth -
                       1).value = report_row['name']
            costlocation_col = col + depth
            sheet.cell(row=crow, column=costlocation_col).value = ', '.join(
                report_row['costLocationsStrings'])
            # collecting values for the formula string
            if len(report['bookings']):
                bookings_start_row = 2
                bookings_end_row = len(report['bookings'])+100
                bookings_costlocations_col = self.get_column_letter_by_column_header(
                    bookings_sheet, self.settings['bookingsCostlocationColumnHeader'])
                bookings_date_col = self.get_column_letter_by_column_header(
                    bookings_sheet, self.settings['bookingsDateColumnHeader'])
                bookings_amount_col = self.get_column_letter_by_column_header(
                    bookings_sheet, self.settings['bookingsAmountColumnHeader'])
                bookings_dbt2_col = self.get_column_letter_by_column_header(
                    bookings_sheet, self.settings['bookingsDebitBookingsType2'])
                bookings_cbt2_col = self.get_column_letter_by_column_header(
                    bookings_sheet, self.settings['bookingsCreditBookingsType2'])
                expense_term = "Ausgaben"
                income_term = "Einnahmen"
                amount_cell_range = bookings_amount_col + \
                    str(bookings_start_row) + ":" + \
                    bookings_amount_col + str(bookings_end_row)
                costlocation_cell_range = bookings_costlocations_col + \
                    str(bookings_start_row) + ":" + \
                    bookings_costlocations_col + str(bookings_end_row)
                date_cell_range = bookings_date_col + \
                    str(bookings_start_row) + ":" + \
                    bookings_date_col + str(bookings_end_row)
                dbt2_cell_range = bookings_dbt2_col + \
                    str(bookings_start_row) + ":" + \
                    bookings_dbt2_col + str(bookings_end_row)
                cbt2_cell_range = bookings_cbt2_col + \
                    str(bookings_start_row) + ":" + \
                    bookings_cbt2_col + str(bookings_end_row)
                bsn = self.get_bookings_sheet_name(
                    report['id'])  # bookings sheet name

            # costlocation check by cell reference
            def formula_opening_cl_cell(slot_details, cell_range, cell_range_search_term):
                # here "," need to be used instead of ";" to separate the arguments, this will be autoreplaced by excel
                formula = "SUMIFS('"+bsn+"'!"+amount_cell_range
                formula += ",'"+bsn+"'!"+costlocation_cell_range + \
                    ",$"+self.colchar(costlocation_col) + str(crow)
                formula += ",'"+bsn+"'!"+date_cell_range + \
                    ",\">=\"&" + slot_details['startDateCellReference']
                formula += ",'"+bsn+"'!"+date_cell_range+",\"<\"&" + \
                    slot_details['endDateCellReference']+"+1"
                formula += ",'"+bsn+"'!"+cell_range+",\""+cell_range_search_term+"\")"
                return formula

            # costlocation check by value
            def formula_opening_cl_value(slot_details, cost_location, cell_range, cell_range_search_term):
                formula = "SUMIFS('"+bsn+"'!"+amount_cell_range
                formula += ",'"+bsn+"'!"+costlocation_cell_range+",\""+cost_location+"\""
                formula += ",'"+bsn+"'!"+date_cell_range + \
                    ",\">=\"&" + slot_details['startDateCellReference']
                formula += ",'"+bsn+"'!"+date_cell_range+",\"<\"&" + \
                    slot_details['endDateCellReference']+"+1"
                formula += ",'"+bsn+"'!"+cell_range+",\""+cell_range_search_term+"\")"
                return formula

            slot_number = 0
            for slot in report_row['slots']:
                slot_details = slot['slotDetails']
                if slot['type'] == 'budget':
                    cell = sheet.cell(
                        row=crow, column=slot_details['endColumn'])
                    cell.value = slot['limit']
                    cell.number_format = self.settings['euroFormat']
                elif slot['type'] == 'bookings' and len(report['bookings']):
                    cell = sheet.cell(
                        row=crow, column=slot_details['endColumn'])
                    formula = ""
                    if len(report_row["costLocationsStrings"]) > 1:
                        for cost_location in report_row["costLocationsStrings"]:
                            formula += "" + \
                                formula_opening_cl_value(
                                    slot_details, cost_location, dbt2_cell_range, expense_term)
                            formula += "-" + \
                                formula_opening_cl_value(
                                    slot_details, cost_location, cbt2_cell_range, income_term)
                            formula += "+" + \
                                formula_opening_cl_value(
                                    slot_details, cost_location, dbt2_cell_range, income_term)
                            formula += "-" + \
                                formula_opening_cl_value(
                                    slot_details, cost_location, cbt2_cell_range, expense_term)
                            if cost_location != report_row["costLocationsStrings"][-1]:
                                formula += "+"
                    else:
                        formula += "" + \
                            formula_opening_cl_cell(
                                slot_details, dbt2_cell_range, expense_term)
                        formula += "-" + \
                            formula_opening_cl_cell(
                                slot_details, cbt2_cell_range, income_term)
                        formula += "+" + \
                            formula_opening_cl_cell(
                                slot_details, dbt2_cell_range, income_term)
                        formula += "-" + \
                            formula_opening_cl_cell(
                                slot_details, cbt2_cell_range, expense_term)
                    if report_row['type'] == 'expense':
                        cell.value = "="+formula
                    elif report_row['type'] == 'income':
                        cell.value = "=-("+formula+")"
                    cell.number_format = self.settings['euroFormat']
                    cell.alignment = styles.Alignment(horizontal="right")
                slot_number += 1

    def fill_report_row_to_listings_sheet(self, report: dict, report_row: dict, sheet: Worksheet, selected_headers: list, current_row: int = 5, first_col: int = 1, depth: int = 1, sub_rows: int = 0):
        crow = current_row
        col = first_col

        # dealing with rows that represent budget groups
        if report_row['type'] == 'group':
            level = len(report_row['hierarchyLocation'])
            font = self.settings['depth'+str(depth - level)+'Font'] if 0 < depth - \
                level < 6 else self.settings['level5Font']
            fill = self.settings['level'+str(level)+'Fill'] if - \
                1 < level < 4 else self.settings['level0Fill']
            cell = sheet.cell(row=crow, column=col + level)
            cell.value = report_row['name']
            cell.font = font
            cell.fill = fill
            for i in range(level, depth + len(selected_headers)):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = font
                cell.fill = fill

        # dealing with rows that represent actual budget lines at the deepest level
        elif report_row['type'] in ['expense', 'income']:
            col = first_col
            for i in range(first_col + depth - 2, depth + len(selected_headers)):
                cell = sheet.cell(row=crow, column=col + i)
                cell.font = self.settings['depth1Font']
                cell.fill = self.settings['listingsHeaderFill']
            info_string = report_row['name'] + " ( cost locations: " + ', '.join(
                report_row['costLocationsStrings']) + " )" if len(report_row['costLocationsStrings']) > 0 else report_row['name']
            sheet.cell(row=crow, column=col + depth - 1).value = info_string
            amount_col = first_col + depth + len(selected_headers) - 2
            start_cell = sheet.cell(row=crow+1, column=amount_col)
            end_cell = sheet.cell(row=crow+sub_rows, column=amount_col)
            sum_formula = "=SUM("+start_cell.coordinate + \
                ":"+end_cell.coordinate+")"
            sheet.cell(row=crow, column=amount_col).value = sum_formula
            sheet.cell(
                row=crow, column=amount_col).number_format = self.settings['euroFormatPrecise']

    def set_border_to_area(self, sheet, start_column=0, start_row=0, end_column=5, end_row=5,
                           side: styles.Side = styles.Side(
                               border_style='thin', color='FF000000'),
                           top=True, bottom=True, left=True, right=True):
        if top:
            for col in range(start_column, end_column+1):
                sheet.cell(row=start_row, column=col).border = styles.Border(
                    top=side)
        if bottom:
            for col in range(start_column, end_column+1):
                sheet.cell(row=end_row, column=col).border = styles.Border(
                    bottom=side)
        if left:
            for row in range(start_row, end_row+1):
                sheet.cell(row=row, column=start_column).border = styles.Border(
                    left=side)
        if right:
            for row in range(start_row, end_row+1):
                sheet.cell(row=row, column=end_column).border = styles.Border(
                    right=side)
        if top and left:
            sheet.cell(row=start_row, column=start_column).border = styles.Border(
                top=side, left=side)
        if top and right:
            sheet.cell(row=start_row, column=end_column).border = styles.Border(
                top=side, right=side)
        if bottom and left:
            sheet.cell(row=end_row, column=start_column).border = styles.Border(
                bottom=side, left=side)
        if bottom and right:
            sheet.cell(row=end_row, column=end_column).border = styles.Border(
                bottom=side, right=side)

    def select_bookings_by_costlocation(self, cost_locations: list, bookings: list):
        selected_bookings = []
        string_cost_locations = [str(cl) for cl in cost_locations]
        for booking in bookings:
            if booking['cost_location'] in string_cost_locations or booking['cost_location'] in cost_locations:
                selected_bookings.append(booking)
        return selected_bookings

    def fill_bookings_listings_sheet(self, bookings: list, sheet: Worksheet, current_row: int, selected_headers: list, first_col: int = 1):
        crow = current_row
        ccol = first_col
        for booking in bookings:
            for header in selected_headers:
                cell = sheet.cell(row=crow, column=ccol)
                if header in ['amount', 'vat', 'receipts_assigned_vat_rates', 'receipts_assigned_assigned_amounts']:
                    cell.style = 'Comma'
                    # cell.number_format = '#,##0.00 €'
                    try:
                        cell.value = float(booking[header])
                    except:
                        cell.value = str(booking[header])
                elif header in ['date']:
                    cell.value = datetime.strptime(
                        booking[header], "%Y-%m-%d %H:%M:%S")
                    cell.number_format = 'DD.MM.YY'
                elif header in ['date_vat_effective']:
                    cell.value = datetime.strptime(booking[header], "%Y-%m-%d")
                    cell.number_format = 'DD.MM.YY'
                elif header in ['id_by_customer', 'debit_postingaccount_number', 'credit_postingaccount_number', 'tax_key', 'booking_number', 'transactions_id_by_customer']:
                    try:
                        cell.value = int(booking[header])
                    except:
                        cell.value = str(booking[header])
                elif header in ['debit_booking_categories', 'credit_booking_categories']:
                    cell.value = ' - '.join(booking[header])
                else:
                    cell.value = str(booking[header])
                ccol += 1
            crow += 1
            ccol = first_col

    def fill_listings_sheet(self, report: dict, sheet: Worksheet, listing: dict):
        first_report_column = 1
        report_structure_levels = 0
        for row in report['rows']:
            new_depth = self.check_depth(row, 1)
            if new_depth > report_structure_levels:
                report_structure_levels = new_depth
        crow = 1
        selected_headers = ['date', 'booking_number', 'cost_location', 'debit_postingaccount_number',
                            'credit_postingaccount_number', 'postingtext', 'amount', 'currency']
        header_width = self.settings['bookingsHeaderWidth']

        # add information about the time window
        cell = sheet.cell(row=crow, column=report_structure_levels)
        start_date_string = listing['start'].strftime('%Y-%m-%d')
        end_date_string = listing['end'].strftime('%Y-%m-%d')
        cell.value = "Time window: from " + start_date_string + " to " + end_date_string
        cell.font = styles.Font(bold=True)
        crow += 2

        # add table header
        ccol = report_structure_levels + 1
        for header in selected_headers:
            if header in header_width:
                sheet.column_dimensions[self.colchar(
                    ccol)].width = header_width[header]
            elif 'default' in header_width and header_width['default'] is not False:
                sheet.column_dimensions[self.colchar(
                    ccol)].width = header_width['default']
            cell = sheet.cell(row=crow, column=ccol)
            cell.value = header
            cell.font = styles.Font(bold=True)
            ccol += 1
        crow += 1
        ccol = 1

        # add rows
        for i in range(first_report_column, report_structure_levels + 1):
            sheet.column_dimensions[self.colchar(
                i)].width = self.settings['groupTitleRowWidth']
        for row in report['rows']:
            selected_bookings = []
            if row['type'] in ['expense', 'income']:
                selected_bookings = self.select_bookings_by_costlocation(
                    row['costLocations'], report['bookings'])
            if not (row['type'] == 'group' and row['name'] == ''):
                self.fill_report_row_to_listings_sheet(
                    report, row, sheet, selected_headers, crow, first_col=first_report_column, depth=report_structure_levels, sub_rows=len(selected_bookings))
                crow += 1
                if row['type'] in ['expense', 'income'] and len(selected_bookings) > 0:
                    self.fill_bookings_listings_sheet(selected_bookings, sheet, crow, selected_headers, first_col=int(
                        first_report_column + report_structure_levels))
                    crow += len(selected_bookings)

    def fill_report_to_sheet(self, report: dict, sheet, bookings_sheet, listings_sheets=[]):
        self.fill_bookings_to_sheet(report['bookings'], bookings_sheet)
        if len(listings_sheets) > 0:
            listings_sheets_names = [sheet.title for sheet in listings_sheets]
            if "listings" in report and len(report["listings"]) > 0:
                for listing in report["listings"]:
                    if self.get_listings_sheet_name(report['id'], listing) in listings_sheets_names:
                        sheet_index = listings_sheets_names.index(
                            self.get_listings_sheet_name(report['id'], listing))
                        self.fill_listings_sheet(
                            report, listings_sheets[sheet_index], listing)
        # add a title to the sheet
        crow, ccol = 1, 1
        cell = sheet.cell(row=crow, column=ccol)
        cell.value = report['name']
        sheet.merge_cells(start_row=crow, start_column=ccol,
                          end_row=crow, end_column=ccol+10)
        cell.font = self.settings['titleFont']
        date_string = sorted(
            report['bookings'], key=lambda booking: booking['date'])[-1]['date']
        last_date = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')
        info_text = "compiled: " + datetime.today().strftime("%d.%m.%Y") + \
            " with last booking from " + last_date.strftime('%d.%m.%Y')
        sheet.cell(row=crow+1, column=ccol).value = info_text
        crow += 2

        rows = report['rows']
        first_report_column = 1
        report_structure_levels = report['structureDepth'] - \
            report['numSaldoRows']

        # add headlines for the slots
        first_slot_col = report_structure_levels + 2
        rows_forward_per_row = self.settings['rowsForwardForAllRows']

        slot_number = 0
        for slot in report['slots']:
            self.add_slot_header(sheet, slot,
                                 col=first_slot_col + slot_number *
                                 report_structure_levels + first_report_column - 1,
                                 row=crow, depth=report_structure_levels)
            slot_number += 1
        crow += 4
        first_data_row = crow
        for row in rows:
            if row['type'].lower() == "group" and len(row['hierarchyLocation']) == 0:
                crow += 1
            row['rowNumber'] = crow
            crow += rows_forward_per_row
            for slot in row['slots']:
                if slot['slotDetails']['endRow'] < crow:
                    slot['slotDetails']['endRow'] = crow
        crow = first_data_row
        # add report rows
        for i in range(first_report_column, report_structure_levels):
            sheet.column_dimensions[self.colchar(
                i)].width = self.settings['groupTitleRowWidth']
        sheet.column_dimensions[self.colchar(
            first_report_column + report_structure_levels - 1)].width = self.settings['itemTitleRowWidth']
        sheet.column_dimensions[self.colchar(
            first_report_column + report_structure_levels)].width = self.settings['costLocationColumnWidth']

        for row in rows:
            self.fill_row_to_sheet(report, row, sheet, bookings_sheet,
                                   first_col=first_report_column, depth=report_structure_levels)

        # add boxes around slots
        for slot in report['slots']:
            self.set_border_to_area(sheet, slot['startColumn'], slot['startRow'],
                                    slot['endColumn'], slot['endRow'] - report['numSaldoRows'])
        return

    def get_bookings_sheet_name(self, report_id: str):
        return report_id + " Buchungen"

    def get_listings_sheet_name(self, report_id: str, listing: dict):
        return report_id + " Kst-Bericht " + listing['name']

    def compile_all_xls_sheets(self, reports: dict, instructions: dict, logger=None):
        logger = self.get_best_choice_logger(logger)
        wb = Workbook()
        if wb.worksheets:
            try:
                wb.remove(wb.worksheets[0])
            except AttributeError:
                wb.remove_sheet(wb.worksheets[0])
        ws = wb.create_sheet("Buchungen")
        self.fill_bookings_to_sheet(reports['bookings'], ws)
        ws = wb.create_sheet("Kostenstellen")
        self.fill_costlocations_to_sheet(reports['costlocations'], ws)
        ws = wb.create_sheet("Konten")
        self.fill_accounts_to_sheet(reports['accounts'], ws)
        report_ids = list(instructions['reports'].keys())
        for report_id in report_ids:
            if report_id in list(reports.keys()):
                ws = wb.create_sheet(report_id)
                wsb = wb.create_sheet(self.get_bookings_sheet_name(report_id))
                wsl = None
                listings_sheets = []
                if "listings" in instructions['reports'][report_id]:
                    for listing in instructions['reports'][report_id]['listings']:
                        wsl = wb.create_sheet(
                            self.get_listings_sheet_name(report_id, listing))
                        listings_sheets.append(wsl)
                self.fill_report_to_sheet(
                    reports[report_id], ws, wsb, listings_sheets)
        return save_virtual_workbook(wb)

    def build_reports(self,
                      bookings: list,
                      expected_bookings: list,
                      instructions: dict,
                      accounts: list,
                      logger: DLogger.DLogger = None):
        logger = self.get_best_choice_logger(logger)
        reports = {}
        self.add_more_account_information_to_bookings(
            bookings, instructions['kontenrahmen']['default'], logger=logger)
        for booking in bookings:
            booking['realisation'] = 'booked'
        reports['bookings'] = bookings
        reports['accounts'] = self.collect_all_accounts(
            bookings, instructions['kontenrahmen']['default'], logger=logger)
        reports['costlocations'] = self.collect_all_costlocations(
            bookings, instructions['kostenstellenplan'], logger=logger)
        for report_id, report in instructions['reports'].items():
            if logger:
                logger.log('building report '+report['name'])

            # add all relevant bookings to the report dictionary
            report_bookings = []
            for booking in bookings:
                if booking['cost_location'] in report['costLocationsStrings']:
                    report_bookings.append(booking)
            reports[report_id] = report
            reports[report_id]['bookings'] = report_bookings

            # check report structure depth
            report_structure_levels = 0
            for row in report['rows']:
                new_depth = self.check_depth(row, 1)
                if new_depth > report_structure_levels:
                    report_structure_levels = new_depth
            report['structureDepth'] = report_structure_levels
            saldo_rows = [row for row in report['rows'] if row['type'].lower(
            ) == "group" and len(row['hierarchyLocation']) == 0]
            report['numSaldoRows'] = len(saldo_rows)

            # add some summary information to the report dictionary
            result = self.add_up_bookings(report_bookings, logger)
            for key, value in result.items():
                reports[report_id][key] = value

        reports['allSheetsFile'] = self.compile_all_xls_sheets(
            reports,  instructions, logger=logger)
        if logger:
            logger.log("completed building reports")
        return reports

    def get_report_summary(self, report: dict, logger=None, format='text'):
        logger = self.get_best_choice_logger(logger)
        res = report['name']
        data = {"expense": "Ausgaben", "revenue": "Einnahmen", "asset": "Activa",
                "liability": "Verbindlichkeiten", "stock": "Vermögen", "profitAndLoss": "Erfolg"}
        for item, description in data.items():
            res += description + ": " + str(report[item]) + '\n'
        res += "Anzahl Buchungen: " + str(len(report['bookings']))+'\n'
        return res

    def send_email_sending(self,
                           sending: dict,
                           reports: dict,
                           logger: DLogger.DLogger = None,
                           client: dict = None,
                           send: bool = True):
        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        if logger:
            logger.log('preparing email to '+sending['recipient']+' with '+str(
                len(sending['packages']))+' packages', 2)
        email_body = "Report:\n"
        report_number = 1
        attachments = []
        for package in sending['packages']:
            required_sheets_for_this_package = []
            for item in package['content']:
                if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                    content_item = self.get_report_summary(
                        reports[item['scope']], format='html')
                    email_body += content_item
                elif package['packageType'] == 'excel' and item['type'] == 'bookings' and item['scope'] == 'all':
                    required_sheets_for_this_package.append('Buchungen')
                elif package['packageType'] == 'excel' and item['type'] == 'accounts' and item['scope'] == 'all':
                    required_sheets_for_this_package.append('Konten')
                elif package['packageType'] == 'excel' and item['type'] == 'costlocations' and item['scope'] == 'all':
                    required_sheets_for_this_package.append('Kostenstellen')
                elif package['packageType'] == 'excel' and item['type'] == 'report':
                    required_sheets_for_this_package.append(item['scope'])
                    required_sheets_for_this_package.append(
                        self.get_bookings_sheet_name(item['scope']))
                    if item['scope'] in reports and "listings" in reports[item['scope']]:
                        for listing in reports[item['scope']]['listings']:
                            required_sheets_for_this_package.append(
                                self.get_listings_sheet_name(item['scope'], listing))
            if required_sheets_for_this_package:
                wb = load_workbook(BytesIO(reports['allSheetsFile']))
                for sheet in wb.get_sheet_names():
                    if not (sheet in required_sheets_for_this_package):
                        wb.remove_sheet(wb[sheet])
                time_zone = timezone("Europe/Berlin")
                today = time_zone.localize(datetime.now()).strftime("%Y-%m-%d")
                if report_number > 1:
                    file_name = f"Financial Report {report_number} {today}.xlsx"
                else:
                    file_name = f"Financial Report {today}.xlsx"
                attachments.append(Attachment(
                    FileContent(base64.b64encode(
                        save_virtual_workbook(wb)).decode()),
                    FileName(file_name),
                    FileType('application/xlsx'),
                    Disposition('attachment')
                ))
                report_number += 1
        message = Mail(
            from_email='hrm@humanrightsmonitor.org',
            to_emails=sending['recipient'],
            subject='Updated DevInt Administration Report',
            html_content=email_body)
        message.attachment = attachments
        try:
            if send:
                response = client['sendgrid'].send(message)
                logger.log("Mail Sending: Sendgrid response: " +
                           str(response.status_code), 2)
        except Exception as e:
            logger.log("Error when trying to send email: "+str(e), 'error')

    def send_teams_sending(self,
                           sending: dict,
                           reports: dict,
                           logger: DLogger.DLogger = None,
                           client: dict = None,
                           send: bool = True):

        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        if logger:
            logger.log('preparing team message with ' +
                       str(len(sending['packages']))+' packages', 2)
        message = pymsteams.connectorcard(client["teams_webhook"])
        message.title("DevInt Accounting Update")
        text = ""
        for package in sending['packages']:
            for item in package['content']:
                if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                    content_item = self.get_report_summary(
                        reports[item['scope']], format="html")
                    text += content_item
            if logger:
                text += logger.getloghtml()
        message.text(text)
        message.addLinkButton("Request new Report",
                              client["reportrequest-backlink"])
        message.addLinkButton("visit Buchhaltungsbuttler",
                              "https://app.buchhaltungsbutler.de")
        if send:
            message.send()

    def send_azure_blob_sending(self,
                                sending: dict,
                                reports: dict,
                                logger: DLogger.DLogger = None,
                                client: dict = None,
                                send: bool = True):

        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        if logger:
            logger.log('preparing blob upload with ' +
                       str(len(sending['packages']))+' packages', 2)
        target = sending['recipient'].split('/')
        report_number = 1
        if target[0] == 'devintaccounting':
            time_zone = timezone("Europe/Berlin")
            today = time_zone.localize(datetime.now()).strftime("%Y-%m-%d")
            for package in sending['packages']:
                required_sheets_for_this_package = []
                for item in package['content']:
                    if package['packageType'] == 'info' and item['type'] == 'reportsummary' and item['scope'] in reports:
                        content_item = self.get_report_summary(
                            reports[item['scope']], format='text')
                        output = BytesIOWrapper(StringIO(content_item))
                        file_name = f"Info {report_number} {today}.txt"
                        if send:
                            blob_client = client["blob_service"].get_blob_client(
                                container=target[1], blob=file_name)
                            blob_client.upload_blob(output, overwrite=True)
                            if logger:
                                logger.log(
                                    "Blob Upload: uploaded "+file_name, 2)
                        report_number += 1
                    elif package['packageType'] == 'info' and item['type'] == 'processlog':
                        content_item = logger.getloglines()
                        output = BytesIOWrapper(StringIO(content_item))
                        file_name = f"Info {report_number} {today}.txt"
                        if send:
                            blob_client = client["blob_service"].get_blob_client(
                                container=target[1], blob=file_name)
                            blob_client.upload_blob(output, overwrite=True)
                            if logger:
                                logger.log(
                                    "Blob Upload: uploaded "+file_name, 2)
                        report_number += 1
                    elif package['packageType'] == 'excel' and item['type'] == 'bookings' and item['scope'] == 'all':
                        required_sheets_for_this_package.append('Buchungen')
                    elif package['packageType'] == 'excel' and item['type'] == 'accounts' and item['scope'] == 'all':
                        required_sheets_for_this_package.append('Konten')
                    elif package['packageType'] == 'excel' and item['type'] == 'costlocations' and item['scope'] == 'all':
                        required_sheets_for_this_package.append(
                            'Kostenstellen')
                    elif package['packageType'] == 'excel' and item['type'] == 'report':
                        required_sheets_for_this_package.append(item['scope'])
                        required_sheets_for_this_package.append(
                            self.get_bookings_sheet_name(item['scope']))
                        if item['scope'] in reports and "listings" in reports[item['scope']]:
                            for listing in reports[item['scope']]['listings']:
                                required_sheets_for_this_package.append(
                                    self.get_listings_sheet_name(item['scope'], listing))
                if required_sheets_for_this_package:
                    wb = load_workbook(BytesIO(reports['allSheetsFile']))
                    for sheet in wb.get_sheet_names():
                        if not (sheet in required_sheets_for_this_package):
                            wb.remove_sheet(wb[sheet])
                    if report_number > 1:
                        file_name = f"Financial Report {report_number} {today}.xlsx"
                    else:
                        file_name = f"Financial Report {today}.xlsx"
                    try:
                        if send:
                            blob_client = client["blob_service"].get_blob_client(
                                container=target[1], blob=file_name)
                            blob_client.upload_blob(
                                BytesIO(save_virtual_workbook(wb)), overwrite=True)
                            logger.log(
                                f"Uploaded file: {file_name} to the blob", 2)
                    except Exception as e:
                        logger.log(
                            "Error when trying to upload blob: "+str(e), 2, type='error')
                    report_number += 1
        return

    def send_cosmosdb_sending(self,
                              sending: dict,
                              reports: dict,
                              logger: DLogger.DLogger = None,
                              client: dict = None,
                              send: bool = True):

        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        if logger:
            logger.log('TODO!!! preparing cosmos db posting with ' +
                       str(len(sending['packages']))+' packages', 2)
        return

    def send_reports(self,
                     reports: dict,
                     sendings: list,
                     logger: DLogger.DLogger = None,
                     client: dict = None,
                     send: bool = True,
                     trigger_selector: list = None):

        logger = self.get_best_choice_logger(logger)
        client = client or self.client
        if logger:
            logger.log("sending reports")
        for sending in sendings:
            # print('checking trigger:',sending['trigger'])
            if sending['trigger'] in trigger_selector:
                # print('trigger found in selector')
                logger.log("sending to "+sending['recipient']+" via " +
                           sending['channelType']+" was triggered by "+sending['trigger']['type'])
                if sending['channelType'] == 'email':
                    self.send_email_sending(
                        sending, reports, logger=logger, client=client, send=send)
                elif sending['channelType'] == 'teams':
                    self.send_teams_sending(
                        sending, reports, logger=logger, client=client, send=send)
                elif sending['channelType'] == 'azureblob':
                    self.send_azure_blob_sending(
                        sending, reports, logger=logger, client=client, send=send)
                elif sending['channelType'] == 'cosmosdb':
                    self.send_cosmosdb_sending(
                        sending, reports, logger=logger, client=client, send=send)
                else:
                    if logger:
                        logger.log("unknown channel type: " +
                                   sending['channelType'], type='error')
        if logger:
            logger.log("completed: sending reports")
        return
