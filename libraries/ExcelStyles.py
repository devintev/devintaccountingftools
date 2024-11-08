from openpyxl import styles

cell_styles = {
    'top_th' : {
        'font': styles.Font(bold=True, size=18, color='FFFFFF'),
        'fill': styles.fills.PatternFill(patternType='solid', fgColor='4B4F58'),
        'alignment' : styles.Alignment(horizontal='center', vertical='center'),
        'borders': styles.borders.Border(
            left=styles.borders.Side(style='thin', color='000000'),
            right=styles.borders.Side(style='thin', color='000000'))
    },
    'devint_th' : {
        'font': styles.Font(bold=True, size=11, color='000000'),
        'fill': styles.fills.PatternFill(patternType='solid', fgColor='E0EBEE'),
        'alignment' : styles.Alignment(horizontal='left', vertical='top', wrap_text=True),
    },
    'HRM_th' : {
        'font': styles.Font(bold=True, size=11, color='FFFFFF'),
        'fill': styles.fills.PatternFill(patternType='solid', fgColor='004B00'),
        'alignment' : styles.Alignment(horizontal='left', vertical='top', wrap_text=True),
    },
    'white' : {
        'fill': styles.fills.PatternFill(patternType='solid', fgColor='FFFFFF'),
    },
    'left_border' : {
        'borders': styles.borders.Border(
            left=styles.borders.Side(style='thin', color='000000'))
    },
    'right_border' : {
        'borders': styles.borders.Border(
            right=styles.borders.Side(style='thin', color='000000'))
    },
}
